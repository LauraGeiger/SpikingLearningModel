from neuron import h
import matplotlib.pyplot as plt
from matplotlib.widgets import Button, Slider, TextBox
from matplotlib import rcParams
import matplotlib.gridspec as gridspec
import numpy as np
import time
import random
import serial
from openpyxl import Workbook
from datetime import datetime
from itertools import product
from matplotlib.ticker import FuncFormatter

# --- TODO --------------------------------------------------#
# determine dopamine level from rel SNc rate
# How to store weights (especially from prefrontal loop)
# Check transistion between grasping types (once learned, there shall be no dip in reward)
# Check connections (other than cor-str) need to be fully connected
# Check if the same neurons can be used across actions
# Check if 5 neurons are required or if less are sufficient
# --- TODO --------------------------------------------------#


h.load_file("stdrun.hoc")

#--- BasalGanglia ------------------------------------------------------------------------------------------------------------------------------------------------#

class BasalGanglia:

    def __init__(self, loops):
        self.loops = loops
        self.buttons = {}
        self.paused = False
        self.hw_connected = False
        self.plot_interval = 100  # ms
        self.simulation_stop_time = 10000 # ms
        self.iteration = 0
        self.ani = None

        # HW
        self.ser_sensor = None
        self.ser_exo = None
        self.recorded_sensor_data_flex = []
        self.recorded_sensor_data_touch = []
        self.performed_action = None
        self.num_flex_sensors = 6
        self.num_touch_sensors = 5
        self.actuators_flexors = [4, 1, 6, 8, 10, 12]
        if len(self.loops[0].actions_names) < 4:
            self.actuators_flexors = self.actuators_flexors[1:1+len(self.loops[0].actions_names)]
        #random.shuffle(self.actuators_flexors)
        self.actuators_extensors = [2, 7, 9, 11, 13]
        self.duration_actuators = 0.3 # s
        self.delay = 0.5 # s

        self._init_loops()
        self._init_plotting()
        self._init_simulation()

        self.run_simulation()
    
    def _init_loops(self):
        for idx, loop in enumerate(self.loops):
            loop.plot_interval = self.plot_interval
            loop.bin_width_firing_rate = self.plot_interval # ms
            if idx > 0:
                loop.set_child_loop(self.loops[idx-1])

    def _init_plotting(self):
        plt.ion()
        self.fig = plt.figure(figsize=(13, 8))
        self.fig.canvas.manager.set_window_title('BasalGangliaOverview')
        self.gs = gridspec.GridSpec(2, 1, figure=self.fig, height_ratios=[1,10])  
        self.gs_control = self.gs[0].subgridspec(1, 4, width_ratios=[1,1,1,3])
        self.gs_plot = self.gs[1].subgridspec(1, 4, width_ratios=[1,1,10,1])     
        self.gs_loops = self.gs_plot[0].subgridspec(len(self.loops), 1)
        self.gs_names = self.gs_plot[1].subgridspec(len(self.loops) + 1, 1)
        self.gs_selections = self.gs_plot[2].subgridspec(len(self.loops) + 1, 1)
        self.gs_probabilities = self.gs_plot[3].subgridspec(len(self.loops) + 1, 1)

        self.axs_control = {}
        self.axs_selections = {}
        self.axs_loops = {}
        self.axs_names = {}
        self.axs_probabilities = {}

        # Pause button
        self.axs_control[0] = self.fig.add_subplot(self.gs_control[0])
        self.axs_control[0].set_axis_off() 
        ax_pause = self.axs_control[0].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['pause'] = Button(ax_pause, 'Pause')
        self.buttons['pause'].on_clicked(self.toggle_pause)

        # Reset button
        self.axs_control[1] = self.fig.add_subplot(self.gs_control[1])
        self.axs_control[1].set_axis_off() 
        ax_reset = self.axs_control[1].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['reset'] = Button(ax_reset, 'Reset')
        self.buttons['reset'].on_clicked(self.reset)

        # HW button
        self.axs_control[2] = self.fig.add_subplot(self.gs_control[2])
        self.axs_control[2].set_axis_off() 
        ax_hw = self.axs_control[2].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['hw'] = Button(ax_hw, 'Connect\nHW')
        self.buttons['hw'].on_clicked(self.connect_hw)

        # Reward plot
        self.ax_reward = self.fig.add_subplot(self.gs_control[3])
        self.ax_reward.set_xlabel('Iteration')
        self.ax_reward.set_ylabel('Reward')
        self.ax_reward.set_xlim(0, 10)  # initial limits, will expand dynamically
        self.ax_reward.set_ylim(-0.1, 1.1)
        self.reward_lines = {}  

        for idx, loop in enumerate(self.loops):
            row = len(self.loops) - idx

            # Plot loop name
            self.axs_loops[idx] = self.fig.add_subplot(self.gs_loops[self.gs_loops.nrows - 1 - idx ])
            self.axs_loops[idx].set_axis_off() 
            ax_loops = self.axs_loops[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
            ax_loops.text(0.5, 0.5, f'{loop.name}', rotation=90, ha='center', va='center', transform=ax_loops.transAxes)
            label_text = '\n'.join(loop.name.split())
            self.buttons[f'{label_text}'] = Button(ax_loops, label='')#, textalignment='center')
            
            if idx == 0: # Motor loop
                self.buttons[f'{label_text}'].on_clicked(self.learn_from_motor_babbling)
            else: # Premotor loop
                self.buttons[f'{label_text}'].on_clicked(self.learn_from_demonstration)
            
            # Init probabilities
            self.axs_probabilities[idx] = self.fig.add_subplot(self.gs_probabilities[self.gs_probabilities.nrows - 2 - idx])
            self.axs_probabilities[idx].set_axis_off()
            self.axs_probabilities[idx].set_xlim(0, 1)
            self.axs_probabilities[idx].set_ylim(0, 1)
            
            if idx == 0:
                # Plot output group name
                self.axs_names[idx] = self.fig.add_subplot(self.gs_names[row])
                self.axs_names[idx].set_axis_off() 
                ax_names = self.axs_names[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                ax_names.set_axis_off()
                output_group = 'Actuators'
                ax_names.text(0.5, 0.5, s=output_group, ha='center', va='center', transform=ax_names.transAxes)

                # Plot outputs of loop
                col = len(loop.actions_names)
                sgs = self.gs_selections[row].subgridspec(1, col)
                self.axs_selections[idx] = [self.fig.add_subplot(sgs[i]) for i in range(col)]
                for i, name in enumerate(loop.actions_names):
                    self.axs_selections[idx][i].set_axis_off() 
                    ax_selections = self.axs_selections[idx][i].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                    label_text = '\n'.join(f"Actuator {i+1}".split())
                    #label_text = '\n'.join(name.split())
                    self.buttons[f'Actuator {i+1}'] = TextBox(ax_selections, label='', label_pad=0.05, initial=f'{label_text}', color='None', textalignment='center')

                # Plot a horizontal bar with width=prob
                self.buttons[f'probability_bar_{idx}'] = self.axs_probabilities[idx].bar(
                    x=0.5, width=0.8, height=0, color='tab:cyan')[0]  # Get BarContainer object
                ax_probabilities = self.axs_probabilities[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                ax_probabilities.set_axis_off()
                self.buttons[f'probability_{idx}'] = ax_probabilities.text(0.5, 0, s='', ha='center', va='center', transform=ax_probabilities.transAxes)

            # Plot input group name
            self.axs_names[idx+1] = self.fig.add_subplot(self.gs_names[row-1])
            self.axs_names[idx+1].set_axis_off() 
            ax_names = self.axs_names[idx+1].inset_axes([0,0,1,1]) #[x0, y0, width, height]
            ax_names.set_axis_off()
            input_group = 'Joints' if idx == 0 else 'Grasp\ntype'
            ax_names.text(0.5, 0.5, s=input_group, ha='center', va='center', transform=ax_names.transAxes)

            # Plot inputs of loop  
            col = len(loop.goals_names)
            sgs = self.gs_selections[row-1].subgridspec(1, col)
            self.axs_selections[idx+1] = [self.fig.add_subplot(sgs[i]) for i in range(col)]
            for i, name in enumerate(loop.goals_names):
                    self.axs_selections[idx+1][i].set_axis_off() 
                    ax_selections = self.axs_selections[idx+1][i].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                    label_text = '\n'.join(name.split())
                    self.buttons[f'{name}'] = Button(ax_selections, label=f'{label_text}', color='None', hovercolor='lightgray')
                    self.buttons[f'{name}'].on_clicked(lambda event, loop=loop, i=i: self.update_goals(loop, i))
                    self.buttons[f'probability_{name}'] = ax_selections.text(0.5, 0.03, '', ha='center')
                    self.buttons[f'sensor_flexion_{name}'] = ax_selections.text(0.5, 0.9, '', ha='center')
            
            if idx == 1:
                self.buttons[f'probability_bar_{idx}'] = self.axs_probabilities[idx].bar(
                    x=0.5, width=0.8, height=0, color='tab:cyan')[0]  # Get BarContainer object
                ax_probabilities = self.axs_probabilities[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                ax_probabilities.set_axis_off()
                self.buttons[f'probability_{idx}'] = ax_probabilities.text(0.5, 0, s='', ha='center', va='center', transform=ax_probabilities.transAxes)
            
            # Initialize reward lines
            line, = self.ax_reward.step([], [], label=loop.name, where='post')
            self.reward_lines[loop.name] = {'line': line, 'xdata': [], 'ydata': []}
        self.ax_reward.legend(loc='upper left')

    def _init_simulation(self):
        h.dt = 1
        h.finitialize()

    def toggle_pause(self, event=None):
        self.paused = not self.paused
        self.buttons['pause'].label.set_text('Continue' if self.paused else 'Pause')
        if not self.paused:
            self.buttons['pause'].ax.set_facecolor(rcParams['axes.facecolor'])
            self.buttons['pause'].color = '0.85'
        else:
            self.buttons['pause'].ax.set_facecolor('r')
            self.buttons['pause'].color = 'r'
            
    def reset(self, event=None):
        for loop in self.loops:
            plt.close(loop.fig)
        plt.close(self.fig)
        if self.hw_connected:
            try:
                self.ser_sensor.close()
            except Exception: None
            try:
                self.ser_exo.close()
            except Exception: None

        print("Reset of Basal Ganglia")
        create_BasalGanglia(no_of_joints=len(self.loops[0].goals_names))

    def connect_hw(self, event=None):
        successful = True

        if not self.hw_connected:
            # Define serial connection for sensor feedback
            try:
                self.ser_sensor.open()
                print(f"Serial connection to sensors established on port {self.ser_sensor.port}")

            except Exception:
                try:
                    self.ser_sensor = serial.Serial(
                        port='COM7',        # For ESP32
                        baudrate=115200,    # Baud rate for serial connection to sensors
                        timeout=1           # Timeout for read in seconds
                    )
                    print(f"Serial connection to sensors established on port {self.ser_sensor.port}")
                    
                except Exception as e:
                    successful = False
                    print(f"Serial connection to sensors failed due to exception: {e}")
            
            # Define serial connection via bluetooth to exoskeleton
            try:
                self.ser_exo.open()
                print(f"Serial connection to exoskeleton established on port {self.ser_exo.port}")
            except Exception:
                try:
                    self.ser_exo = serial.Serial(
                        port='COM11',       # For HC-06
                        baudrate=9600,      # Baud rate for connection to bluetooth module
                        write_timeout=5           # Timeout for read in seconds
                    )
                    print(f"Serial connection to exoskeleton established on port {self.ser_exo.port}")
                except Exception as e:
                    successful = False
                    print(f"Serial connection to exoskeleton failed due to exception: {e}")
        else:
            try:
                self.ser_sensor.close()
            except Exception: None
            try:
                self.ser_exo.close()
            except Exception: None
            print("Serial connections closed")

        if successful:
            self.hw_connected = not self.hw_connected
            self.buttons['hw'].label.set_text('Disconnect\nHW' if self.hw_connected else 'Connect HW')

    def learn_from_motor_babbling(self, event):
        if self.hw_connected:
            previous_action = None
            previous_performed_action = None
            
            for action in self.loops[0].actions:
                for trial in range(2):
                    h.continuerun(h.t + self.plot_interval)
                    #action = random.choice(self.loops[0].actions)
                    self.perform_action(action)
                    if self.recorded_sensor_data_flex: 
                        self.analyze_sensor_data_flex()
                    self.update_GUI_goals_actions()
                    if previous_action and previous_performed_action:
                        self.loops[0].update_weights(current_time=h.t, goal=previous_performed_action, action=previous_action)
                        self.loops[0].update_plots(current_time=h.t, goal=previous_performed_action, action=previous_action)
                    
                    self.loops[0].cortical_input_stimuli(current_time=h.t, goal=self.performed_action)
                    previous_action = action
                    previous_performed_action = self.performed_action
                time.sleep(1)

    def learn_from_demonstration(self): None

    def read_sensor_data(self, duration=5):
        self.recorded_sensor_data_flex = []  # Stores all flex sensor readings
        self.recorded_sensor_data_touch = []  # Stores all touch sensor readings
        start_time = time.time()
        while time.time() - start_time < duration:
            if self.ser_sensor.in_waiting > 0:
                line = self.ser_sensor.readline().decode('utf-8', errors='ignore').strip()
                try:
                    values = [float(x) for x in line.split(',')]
                    self.recorded_sensor_data_flex.append(values[:self.num_flex_sensors])
                    self.recorded_sensor_data_touch.append(values[self.num_flex_sensors:self.num_flex_sensors+self.num_touch_sensors])
                except ValueError:
                    print(f"Ignored malformed line: {line}")
            time.sleep(0.01)  # ~100 Hz sampling
    
    def analyze_sensor_data_flex(self, alpha=0.8, flexion_threshold=30, extension_threshold=30):
        #try:
        if len(self.recorded_sensor_data_flex[0]) < self.num_flex_sensors:
            print("Not enough data from flex sensors")
        else:
            prev_filtered = [self.recorded_sensor_data_flex[0][i] for i in range(self.num_flex_sensors)]
            start_filtered = prev_filtered.copy()
            max_filtered = prev_filtered.copy()
            min_filtered = prev_filtered.copy()

            self.flexion_detected = [False] * self.num_flex_sensors
            self.extension_detected = [False] * self.num_flex_sensors

            for sample in self.recorded_sensor_data_flex[1:]:
                for i in range(self.num_flex_sensors):
                    # Apply low-pass filter
                    filtered = alpha * prev_filtered[i] + (1 - alpha) * sample[i]

                    # Track max and min
                    max_filtered[i] = max(max_filtered[i], filtered)
                    min_filtered[i] = min(min_filtered[i], filtered)

                    # Detect flexion and extension
                    if (max_filtered[i] - start_filtered[i]) > flexion_threshold:
                        self.flexion_detected[i] = True
                    if (start_filtered[i] - min_filtered[i]) > extension_threshold:
                        self.extension_detected[i] = True

                    prev_filtered[i] = filtered

            if len(self.loops[0].actions_names) < 4:
                self.flexion_detected = self.flexion_detected[1:1+len(self.loops[0].actions_names)]
                max_filtered = max_filtered[1:1+len(self.loops[0].actions_names)]
                start_filtered = start_filtered[1:1+len(self.loops[0].actions_names)]
                min_filtered = min_filtered[1:1+len(self.loops[0].actions_names)]

            for i, name in enumerate(self.loops[0].goals_names):
                flex = self.flexion_detected[i]
                extend = self.extension_detected[i]
                delta_up = max_filtered[i] - start_filtered[i]
                delta_down = start_filtered[i] - min_filtered[i]
                text = f"{delta_up:.2f} {'Flexion' if flex else ''}"
                self.buttons[f'sensor_flexion_{name}'].set_text(text)
            '''
            # Print results
            print("\nFlexion and Extension Detection Results:")
            for i in range(len(self.loops[0].actions_names)):
                baseline = start_filtered[i]
                flex = self.flexion_detected[i]
                extend = self.extension_detected[i]
                delta_up = max_filtered[i] - start_filtered[i]
                delta_down = start_filtered[i] - min_filtered[i]
                print(
                    f"Sensor {i}: "
                    f"Baseline = {baseline} "
                    f"{'👉 Flexion' if flex else '   '} "
                    f"{'👈 Extension' if extend else ''} "
                    f"(Δ up = {delta_up:.2f}, Δ down = {delta_down:.2f})"
                )
            '''
            self.performed_action = ''.join(['1' if value else '0' for value in self.flexion_detected])
            print(f"performed action = {self.performed_action}")
        #except Exception as e: print(e)

    def analyze_sensor_data_touch(self, alpha=0.8, touch_threshold=20, window=30):
        #try:
        if len(self.recorded_sensor_data_touch[0]) < self.num_touch_sensors:
            print("Not enough data from touch sensors")
        else:
            prev_filtered = [self.recorded_sensor_data_touch[0][i] for i in range(self.num_touch_sensors)]
            start_filtered = prev_filtered.copy()
            max_filtered = prev_filtered.copy()
            min_filtered = prev_filtered.copy()

            self.touch_detected = [False] * self.num_touch_sensors

            for sample in self.recorded_sensor_data_touch[1:]:
                for i in range(self.num_touch_sensors):
                    # Apply low-pass filter
                    filtered = alpha * prev_filtered[i] + (1 - alpha) * sample[i]

                    # Track max and min
                    max_filtered[i] = max(max_filtered[i], filtered)
                    min_filtered[i] = min(min_filtered[i], filtered)

                    # Detect touch
                    if (max_filtered[i] - start_filtered[i]) > touch_threshold:
                        self.touch_detected[i] = True

                    prev_filtered[i] = filtered

            # Print results
            '''
            print("\nTouch Detection Results:")
            for i in range(self.num_touch_sensors):
                baseline = start_filtered[i]
                touch = self.touch_detected[i]
                delta_up = max_filtered[i] - start_filtered[i]
                delta_down = start_filtered[i] - min_filtered[i]
                print(
                    f"Sensor {i}: "
                    f"Baseline = {baseline} "
                    f"{'👉 Touch' if touch else '   '} "
                    f"(Δ up = {delta_up:.2f}, Δ down = {delta_down:.2f})"
                )
            '''
        #except Exception as e: print(e)

    def update_GUI_goals_actions(self, frame=None): 
        #print(f"{h.t} update_GUI_goals_actions")       
        for idx, loop in enumerate(self.loops):
            if loop.selected_goal:
                visibility = False if set(loop.selected_goal) == {'0'} else True
                self.update_probability(loop_id=idx, probability=loop.expected_reward_over_time[loop.selected_goal][-1], visibility=visibility)

            if not loop.child_loop:
                for i, name in enumerate(loop.actions_names):
                    if loop.selected_goal:
                        char = '0'
                        if loop.selected_action:
                            char = loop.selected_action[0][i]
                        selected = char == '1'
                        reward = loop.reward_over_time[loop.selected_goal][-1]
                        #self.highlight_textbox(name, selected, reward)
                        self.highlight_textbox(f"Actuator {i+1}", selected, reward)
            for i, name in enumerate(loop.goals_names):
                #if not loop.child_loop: 
                    #self.buttons[f'sensor_flexion_{name}'].set_text('') # reset sensor reading
                if loop.selected_goal:
                    char = loop.selected_goal[i]
                    selected = char == '1'
                    reward = None
                    
                    if idx < len(self.loops) - 1:
                        if self.loops[idx+1].selected_goal:
                            reward = self.loops[idx+1].reward_over_time[self.loops[idx+1].selected_goal][-1]
                    
                    self.highlight_textbox(name, selected, reward)
                    
                    if selected: self.update_goal_probability(name, loop, probability=loop.expected_reward_over_time[loop.selected_goal][-1]) 
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
        
    def update_GUI_performed_action_reward(self, frame=None):
        #print(f"{h.t} update_GUI_performed_action_reward")
        self.iteration = int(h.t / self.plot_interval)
        
        for idx, loop in enumerate(self.loops):
            if not loop.child_loop:
                for i, name in enumerate(loop.actions_names):
                    if loop.selected_goal:
                        char = '0'
                        if loop.selected_action:
                            char = loop.selected_action[0][i]
                        selected = char == '1'
                        reward = loop.reward_over_time[loop.selected_goal][-1]
                        if reward: 
                            self.highlight_textbox(f"Actuator {i+1}", selected, reward)
            for i, name in enumerate(loop.goals_names):
                if loop.selected_goal:
                    char = loop.selected_goal[i]
                    selected = char == '1'
                    reward = None
                    
                    if idx < len(self.loops) - 1:
                        if self.loops[idx+1].selected_goal:
                            reward = self.loops[idx+1].reward_over_time[self.loops[idx+1].selected_goal][-1]
                    self.highlight_textbox(name, selected, reward)
                    
            # Get reward for currently selected goal
            if loop.selected_goal:
                current_reward = loop.reward_over_time[loop.selected_goal][-1]
            else:
                current_reward = 0

            rl = self.reward_lines[loop.name]
            rl['xdata'].append(self.iteration)
            rl['ydata'].append(current_reward)
            rl['line'].set_data(rl['xdata'], rl['ydata'])
        # Update axes limits dynamically
        max_iter = max(rl['xdata'][-1] for rl in self.reward_lines.values() if rl['xdata'])
        self.ax_reward.set_xlim(0, max(10, max_iter + 1))
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
        
    def highlight_textbox(self, name, selected, reward=None):
        if selected:
            if reward:
                self.buttons[f'{name}'].ax.set_facecolor('tab:olive')
                self.buttons[f'{name}'].color = 'tab:olive'
            else:
                self.buttons[f'{name}'].ax.set_facecolor('gold')
                self.buttons[f'{name}'].color = 'gold'
        else:
            self.buttons[f'{name}'].ax.set_facecolor(rcParams['axes.facecolor'])
            self.buttons[f'{name}'].color = 'None'
   
    def update_goals(self, loop, goal_idx):
        if loop.selected_goal:
            value = loop.selected_goal[goal_idx]=='0'
        else:
            value = 1
        loop.buttons[f'cor_dur_slider{goal_idx}'].set_val(value)
        loop.update_cor_dur(val=value, goal_idx=goal_idx)
        self.update_GUI_goals_actions()

    def update_probability(self, loop_id, probability, visibility):
        if visibility:
            self.buttons[f'probability_{loop_id}'].set_text(f'{probability:.1%}')
            self.buttons[f'probability_{loop_id}'].set_position((0.5, probability + 0.03))
            self.buttons[f'probability_bar_{loop_id}'].set_height(probability)
        else:
            self.buttons[f'probability_{loop_id}'].set_text(f'')
            self.buttons[f'probability_bar_{loop_id}'].set_height(0)
    
    def update_goal_probability(self, name, loop, probability):
        if loop.single_goal:
            if self.buttons.get(f'probability_{name}'):
                self.buttons[f'probability_{name}'].set_text(f'{probability:.1%}')
        else:
            goal_idx = loop.goals_names.index(name)

            total = 0
            count = 0
            for key in loop.expected_reward_over_time:
                if key[goal_idx] == '1':
                    try:
                        total += loop.expected_reward_over_time[key][-1]
                        count += 1
                    except IndexError:
                        pass  # Skip if data is missing

            if count > 0:
                avg_prob = total / count
                if self.buttons.get(f'probability_{name}'):
                    self.buttons[f'probability_{name}'].set_text(f'{avg_prob:.1%}')
    
    def generate_action_command(self, action=None):
        action_command = ''
        factor = 2
        duration_flexors = 0

        if action:
            duration_flexors = factor
        else:
            if self.loops[0].selected_action:
                action = self.loops[0].selected_action[0]  
                
                duration_flexors = factor * self.loops[0].selected_action[1]   
        
        if duration_flexors:
            action_command_parts = []

            # Control all actuators
            #for actuator in self.actuators_extensors + self.actuators_flexors:
            #    action_command_parts.append(f"0-{actuator}-i")
            #for actuator in self.actuators_extensors + self.actuators_flexors:
            #    action_command_parts.append(f"{self.duration_actuators*1000}-{actuator}-h")

            # Control individual flexors based on selected action
            for i, bit in enumerate(action):
                if bit == '1':
                    flexor = self.actuators_flexors[i]
                    # Inlet position
                    action_command_parts.append(f"{(self.duration_actuators+self.delay)*1000}-{flexor}-i")

            for i, bit in enumerate(action):
                if bit == '1':
                    flexor = self.actuators_flexors[i]
                    # Hold position
                    action_command_parts.append(f"{(self.duration_actuators+self.delay+duration_flexors)*1000}-{flexor}-h")

            action_command = '/'.join(action_command_parts) + '/'
        
        return action_command, duration_flexors

    def perform_action(self, action=None):
        duration = 0
        if action: # used for motor babbling
            action_command, duration = self.generate_action_command(action)
            print(f"Perform action: {action} for {duration}s")
            
        else:
            if self.loops[0].selected_action:
                action_command, duration = self.generate_action_command()
                print(f"Perform action: {self.loops[0].selected_action[0]} for {duration}s")
            
        if duration: # perform action
            start_stop_command = 'S'
            self.ser_sensor.flushInput() # delete values in serial input buffer
    
            self.ser_exo.write(action_command.encode())
            time.sleep(2)

            self.ser_exo.write(start_stop_command.encode())
            self.read_sensor_data(duration=duration+2)
            self.ser_exo.write(start_stop_command.encode())
            time.sleep(2)

    def run_simulation(self):
        try:
            while True:
                if self.paused or all(all(val == 0 for val in loop.cortical_input_dur_rel) for loop in self.loops): 
                    # Simulation paused
                    time.sleep(0.1)

                    for loop in self.loops:
                        loop.fig.canvas.draw_idle()   
                        loop.fig.canvas.flush_events()
                    continue
                
                start_time = time.time()

                # Update selections in basal ganglia overview plot
                time_step = time.time()
                self.update_GUI_goals_actions()
                if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s update_GUI_goals_actions")

                for loop in self.loops:
                    if loop.selected_goal: loop.cortical_input_stimuli(current_time=h.t)
                    
                # Run simulation
                time_step = time.time()
                h.continuerun(h.t + self.plot_interval)
                if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s continuerun")

                # --- Action selection and reward update ---#                
                start_time_first_part = time.time()
                for loop in self.loops:
                    time_step = time.time()
                    loop.analyze_thalamus_activation_time(current_time=h.t)
                    if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s analyze_thalamus_activation_time {loop.name}")
                    time_step = time.time()
                    loop.select_action()
                    if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s select_action {loop.name}")
                
                self.update_GUI_goals_actions()
                if self.hw_connected:
                    self.perform_action()
                    if self.recorded_sensor_data_flex: self.analyze_sensor_data_flex()
                    if self.recorded_sensor_data_touch: self.analyze_sensor_data_touch()
                    
                    
                for loop in self.loops:
                    time_step = time.time()
                    loop.determine_reward(current_time=h.t, hw_connected=self.hw_connected, performed_action=self.performed_action)
                    if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s determine_reward {loop.name}")
                duration = time.time() - start_time_first_part
                #print(f"{duration:.6f} s first part")

                self.update_GUI_performed_action_reward()
                # --- Weight and plot update ---# 
                start_time_second_part = time.time()
                for loop in self.loops:
                    time_step = time.time()
                    loop.update_weights(current_time=h.t)
                    if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s update_weights {loop.name}")
                    time_step = time.time()
                    loop.update_plots(current_time=h.t)
                    if time.time() - time_step > 1: print(f"{(time.time() - time_step):.6f} s update_plots {loop.name}")
                duration = time.time() - start_time_second_part
                #print(f"{duration:.6f} s second part")

                # Pause simulation
                if int(h.t) % self.simulation_stop_time == 0:
                    self.toggle_pause()

                duration = time.time() - start_time
                print(f"Loop took {duration:.6f} s")

                #time.sleep(1)
                self.fig.canvas.draw_idle()
                self.fig.canvas.flush_events()
                for loop in self.loops:
                    loop.fig.canvas.draw_idle()   
                    loop.fig.canvas.flush_events()
            
        except KeyboardInterrupt:
            print("\nCtrl-C pressed. Storing data...")
            plt.close()
            if self.hw_connected:
                try:
                    self.ser_sensor.close()
                except Exception: None
                try:
                    release_command = ''
                    for actuator in range(1,14):
                        release_command += f'0-{actuator}-o/'
                    self.ser_exo.write(release_command.encode())
                    time.sleep(2)
                    self.ser_exo.write('S'.encode())
                    time.sleep(2)
                    self.ser_exo.write('S'.encode())
                    time.sleep(2)
                    self.ser_exo.close()
                except Exception: None

        finally:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            path = f"Data\{timestamp}"

            for loop in self.loops:
                loop.save_data(path) # Excel file
                loop.fig.savefig(f"{path}_{loop.name}.png", dpi=300, bbox_inches='tight') # GUI Screenshot
                        

#--- BasalGangliaLoop ------------------------------------------------------------------------------------------------------------------------------------------------#

class BasalGangliaLoop:

    def __init__(self, name, input, output, binary_input=False, single_goal=False, goal_action_table=None, actions_to_plot=None):
        self.name = name
        self.goals_names = input
        self.goals = [''.join(bits) for bits in product('10', repeat=len(input))] # binary combination of all inputs
        self.selected_goal = ''.join('0' for _ in self.goals_names)
        self.actions_names = output
        self.actions = [''.join(bits) for bits in product('10', repeat=len(output)) if any(bit == '1' for bit in bits)] # binary combination of all outputs
        self.selected_action = None
        if actions_to_plot is not None:
            self.actions_to_plot = len(self.actions) if len(self.actions) <= actions_to_plot else actions_to_plot
        self.binary_input = binary_input
        self.single_goal = single_goal
        self.goal_action_table = goal_action_table
        self.child_loop = None

        self.cell_types_numbers = {'Cor':  [len(self.goals),   1], 
                                   'SNc':  [len(self.actions), 1], 
                                   'MSNd': [len(self.actions), 5], 
                                   'MSNi': [len(self.actions), 5], 
                                   'GPe':  [len(self.actions), 5], 
                                   'GPi':  [len(self.actions), 5], 
                                   'Thal': [len(self.actions), 5]}
        self.cell_types = list(self.cell_types_numbers.keys())

        self.stim_intervals = {
            'SNc'       : 1000 / 5,   # Hz
            'MSNd'      : 1000 / 7.4, # Hz (tonic baseline)
            'MSNi'      : 1000 / 3.5, # Hz (tonic baseline)
            'GPe'       : 1000 / 48,  # Hz
            'GPi'       : 1000 / 69,  # Hz
            'Thal'      : 1000 / 14,  # Hz
            'SNc_burst' : 1000 / 50,  # Hz
            'Cor'       : 1000 / 40   # Hz (cortical input stimulation)
        }
 
        self.stim_weights = {
            'SNc'       : 0.5,
            'MSNd'      : 0.5,
            'MSNi'      : 0.5,
            'GPe'       : 0.5,
            'GPi'       : 0.5,
            'Thal'      : 0.5,
            'SNc_burst' : 0.5,
            'Cor'       : 0.4
        }

        self.stims = {}
        self.syns = {}
        self.ncs = {}

        # Define connection specifications
        self.connection_specs_cor = [# pre_group, post_group, label, e_rev, weight, tau, delay
            ('Cor', 'MSNd', 'Cor_to_MSNd',   0, 0.25,   5, 1),   # excitatory
            ('Cor', 'MSNi', 'Cor_to_MSNi',   0, 0.25,   5, 1)    # excitatory
        ]

        # Define connection specifications
        self.connection_specs = [# pre_group, post_group, label, e_rev, weight, tau, delay
            ('SNc', 'MSNd', 'SNc_to_MSNd',   0, 0,     10, 1),   # excitatory
            ('SNc', 'MSNi', 'SNc_to_MSNi', -85, 0,     10, 1),   # inhibitory
            ('MSNd', 'GPi', 'MSNd_to_GPi', -85, 0.6,   10, 1),   # inhibitory
            ('MSNi', 'GPe', 'MSNi_to_GPe', -85, 1.0,   10, 1),   # inhibitory
            ('GPe',  'GPi',  'GPe_to_GPi', -85, 0.3,   10, 1),   # inhibitory
            ('GPi', 'Thal', 'GPi_to_Thal', -85, 1.0,   10, 1)    # inhibitory
        ]
        
        self._is_updating_programmatically = False
        self.noise = 0
        self.n_spikes_SNc_burst = 5
        self.learning_rate = 0.05 #0.1
        self.reward_times = []
        self.activation_times = []
        self.expected_reward_over_time = {}
        self.activation_over_time = {}
        self.cortical_input_dur_rel = [0] * len(self.goals_names)
        self.expected_reward_value = 0.5
        self.reward_over_time = {}
        self.dopamine_over_time = {}
        self.weight_cell_types = ['MSNd', 'MSNi']
        self.weight_times = []
        self.weights_over_time = {(ct, action_id, msn_id, goal_id, cor_id): [] 
                            for ct in self.weight_cell_types
                            for action_id, action in enumerate(self.actions)
                            for msn_id in range(self.cell_types_numbers[ct][1])
                            for goal_id, goal in enumerate(self.goals)
                            for cor_id in range(self.cell_types_numbers['Cor'][1])
                            }
        self.cor_nc_index_map = {} 
        self.apc_refs = []

        self.plot_interval = None
        self.bin_width_firing_rate = None

        self.buttons = {}
        self.rates = {}
        self.rates_rel = {}

        self._init_cells()
        self._init_spike_detectors()
        self._init_stimuli()
        self._init_connections(self.connection_specs)
        self._init_connections(self.connection_specs_cor, cor=True)
        self._init_connection_stimuli()
        self._init_reward()
        self._init_activation()
        self._init_recording()
        self._init_plotting()

    def _init_cells(self):
        # Create neuron populations
        self.cells = {
            cell_type: [
                [create_cell(f'{cell_type}_{population}_{index}') for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ] 
            for cell_type in self.cell_types
        }
    
    def _init_spike_detectors(self):
        # Spike detectors and vectors
        self.spike_times = {
            cell_type: [
                [h.Vector() for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ] 
            for cell_type in self.cell_types
        }

        for cell_type in self.cell_types:
            for population in range(self.cell_types_numbers[cell_type][0]):
                for index, cell in enumerate(self.cells[cell_type][population]):
                    apc = h.APCount(cell(0.5))
                    apc.record(self.spike_times[cell_type][population][index])
                    self.apc_refs.append(apc)

    def _init_stimuli(self):
        for cell_type in self.cell_types:
            self.stims[cell_type], self.syns[cell_type], self.ncs[cell_type] = [], [], []
            for population in range(self.cell_types_numbers[cell_type][0]):
                for i, cell in enumerate(self.cells[cell_type][population]):
                    if 'SNc' in str(cell):
                        offset = self.stim_intervals['SNc']/2
                    elif 'Cor' in str(cell):
                        offset = 0
                    else:
                        offset = i*self.stim_intervals[cell_type]/self.cell_types_numbers[cell_type][1]
                    stim, syn, nc = create_stim(cell, start=offset, interval=self.stim_intervals[cell_type], weight=self.stim_weights[cell_type], noise=self.noise)
                    self.stims[cell_type].append(stim)
                    self.syns[cell_type].append(syn)
                    self.ncs[cell_type].append(nc)

    def _init_connections(self, connection_specs, cor=False):
        if cor: 
            index = 0

        # Create connections based on the specification
        for pre_group, post_group, label, e_rev, weight, tau, delay in connection_specs:
            self.ncs.update({label: []}) # Additional connections dict to store NetCons
            self.syns.update({label: []}) # Additional connections dict to store ExpSyns

            if cor:
                # Only do index mapping for cortical to striatal connections
                self.cor_nc_index_map[label] = {}
                index = 0
                for population_pre_cell in range(self.cell_types_numbers[pre_group][0]):
                    for pre_cell_id, pre_cell in enumerate(self.cells[pre_group][population_pre_cell]):
                        for population_post_cell in range(self.cell_types_numbers[post_group][0]):
                            for post_cell_id, post_cell in enumerate(self.cells[post_group][population_post_cell]):
                                syn = h.ExpSyn(post_cell(0.5))
                                syn.e = e_rev
                                syn.tau = tau
                                nc = h.NetCon(pre_cell(0.5)._ref_v, syn, sec=pre_cell)
                                nc.weight[0] = weight
                                nc.delay = delay
                                self.syns[label].append(syn)
                                self.ncs[label].append(nc)

                                # Save index mapping if cortical to striatal
                                if label.startswith('Cor_to'):
                                    key = (post_group, population_post_cell, post_cell_id, population_pre_cell, pre_cell_id)
                                    self.cor_nc_index_map[label][key] = index
                                index += 1
            else:
                for action_id, action in enumerate(self.actions):
                    '''
                    pre_cells = self.cells[pre_group][action_id]
                    post_cells = self.cells[post_group][action_id]
                    if len(pre_cells) > len(post_cells):
                        pre_cells = random.sample(pre_cells, len(post_cells))
                    elif len(post_cells) > len(pre_cells):
                        post_cells = random.sample(post_cells, len(pre_cells))
                    for pre_cell, post_cell in zip(pre_cells, post_cells):
                        syn = h.ExpSyn(post_cell(0.5))
                        syn.e = e_rev
                        syn.tau = tau
                        nc = h.NetCon(pre_cell(0.5)._ref_v, syn, sec=pre_cell)

                        nc.weight[0] = weight
                        nc.delay = delay
                        self.syns[label].append(syn)
                        self.ncs[label].append(nc)
                    '''
                    
                    for pre_cell in self.cells[pre_group][action_id]:
                        for post_cell in self.cells[post_group][action_id]:
                            syn = h.ExpSyn(post_cell(0.5))
                            syn.e = e_rev
                            syn.tau = tau
                            nc = h.NetCon(pre_cell(0.5)._ref_v, syn, sec=pre_cell)
                            nc.weight[0] = weight
                            nc.delay = delay
                            self.syns[label].append(syn)
                            self.ncs[label].append(nc)
                    
    def _init_connection_stimuli(self):
        # Additional stimuli for dopamine bursts
        self.stims.update({'SNc_burst': []})
        self.syns.update({'SNc_burst': []})
        self.ncs.update({'SNc_burst': []})

        for a, _ in enumerate(self.actions):
            for cell in self.cells['SNc'][a]:
                stim, syn, nc = create_stim(cell, start=0, interval=self.stim_intervals['SNc_burst'], weight=self.stim_weights['SNc_burst'], noise=self.noise)
                self.stims['SNc_burst'].append(stim)
                self.syns['SNc_burst'].append(syn)
                self.ncs['SNc_burst'].append(nc)
        for nc in self.ncs['SNc_burst']:
            nc.active(False)

        # Initialize weights over time
        self.weight_times.append(0)
        for ct in self.weight_cell_types:
            for action_id, action in enumerate(self.actions):
                for msn_id in range(self.cell_types_numbers[ct][1]):
                    for goal_id, goal in enumerate(self.goals):
                        for cor_id in range(self.cell_types_numbers['Cor'][1]):
                            for _, post_group, _, _, weight, _, _ in self.connection_specs_cor:
                                if ct == post_group:
                                    self.weights_over_time[(ct, action_id, msn_id, goal_id, cor_id)].append(weight)
        for nc in self.ncs[f'Cor']:
            nc.active(False)

    def _init_reward(self):
        # Reward initialization
        self.reward_times.append(0)
        for goal in self.goals:
            self.expected_reward_over_time[goal] = [self.expected_reward_value]
            self.reward_over_time[goal] = [0]
            self.dopamine_over_time[goal] = [0]
        
    def _init_activation(self):
        # Activation initialization
        self.activation_times.append(0)
        for action in self.actions:
            self.activation_over_time[action] = [0]

    def _init_recording(self):
        # Recording
        self.recordings = {ct: [[h.Vector().record(cell(0.5)._ref_v) for cell in self.cells[ct][population]] for population in range(self.cell_types_numbers[ct][0])] for ct in self.cell_types}
        self.t_vec = h.Vector().record(h._ref_t)

    def _init_plotting(self):
        plt.ion()
        self.fig = plt.figure(figsize=(13, 8))
        self.fig.canvas.manager.set_window_title(self.name)
        self.rows = 3
        self.gs = gridspec.GridSpec(2, 1, figure=self.fig, height_ratios=[1, 2 * self.rows])
        self.gs_control = self.gs[0].subgridspec(1, 2 + len(self.goals_names), width_ratios=[1] + [1] * len(self.goals_names) + [2])
        self.gs_plot = self.gs[1].subgridspec(self.rows, self.actions_to_plot)
        self.axs_control = [self.fig.add_subplot(self.gs_control[i]) for i in range(2 + len(self.goals_names))]
        self.axs_plot = []
        for i in range(self.rows):
            row = []
            for j in range(self.actions_to_plot):
                if j == 0:
                    ax = self.fig.add_subplot(self.gs_plot[i, j])
                else:
                    ax = self.fig.add_subplot(self.gs_plot[i, j], sharey=row[0])
                row.append(ax)
            self.axs_plot.append(row)
        for i in range(self.rows):
            for j in range(1, self.actions_to_plot):  # skip first in row
                plt.setp(self.axs_plot[i][j].get_yticklabels(), visible=False)

        [self.row_potential, self.row_spike, self.row_weights] = list(range(self.rows))
        
        # Deactivate axis for control axes
        for ax in self.axs_control:
            ax.set_axis_off() # deactivate axis
        self.axs_control[-1].set_axis_on() # activate axis for reward plot
        
        self._init_membrane_potential_plot()
        self._init_spike_plot()
        self._init_weight_plot()
        self._init_reward_plot()
        self._init_control_panel()

        plt.show()
        plt.tight_layout()

    def _init_membrane_potential_plot(self):
        # Membrane potential plot
        self.axs_plot[self.row_potential][0].set_ylabel('Membrane potential (mV)')
        self.mem_lines = {ct: [] for ct in self.cell_types}

        for i, ch in enumerate(range(self.actions_to_plot)):
            for j, ct in enumerate(self.cell_types):
                avg_line, = self.axs_plot[self.row_potential][i].plot([], [], f'C{j}', label=ct)
                self.mem_lines[ct].append(avg_line)

            self.axs_plot[self.row_potential][i].set_title(f'{self.actions[ch]}')
            self.axs_plot[self.row_potential][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_potential][i].set_ylim(-85, 65)
            self.axs_plot[self.row_potential][i].xaxis.set_major_formatter(ms_to_s)
        self.axs_plot[self.row_potential][-1].legend(loc='upper right')

    def _init_spike_plot(self):
        # Spike raster plot and rate lines
        self.raster_lines = {ct: [[] for _ in range(self.actions_to_plot)] for ct in self.cell_types}
        self.rate_lines = {ct: [] for ct in self.cell_types}
        self.axs_plot[self.row_spike][0].set_ylabel('Spike raster')

        for i, ch in enumerate(range(self.actions_to_plot)):
            for j, ct in enumerate(self.cell_types):
                self.raster_lines[ct][i] = []
                for k in range(self.cell_types_numbers[ct][1]):
                    line, = self.axs_plot[self.row_spike][i].plot([], [], f'C{j}.', markersize=3)
                    self.raster_lines[ct][i].append(line)

                rate_line, = self.axs_plot[self.row_spike][i].step([], [], f'C{j}')
                self.rate_lines[ct].append(rate_line)
            self.axs_plot[self.row_spike][i].plot([], [], color='black', linestyle='dotted', label=f'Spikes')
            self.axs_plot[self.row_spike][i].plot([], [], color='black', label=f'Relative\nrate')

            self.total_cells = sum(self.cell_types_numbers[ct][1] for ct in self.cell_types)
            
            y_max = self.total_cells + 1.5
            self.axs_plot[self.row_spike][i].set_ylim(0.5, y_max)
            yticks = []
            cumulative = 0
            for ct in self.cell_types:
                mid = y_max - (cumulative + (self.cell_types_numbers[ct][1]+1) / 2)
                yticks.append(mid)
                cumulative += self.cell_types_numbers[ct][1] 
            self.axs_plot[self.row_spike][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_spike][i].set_yticks(yticks)
            self.axs_plot[self.row_spike][i].set_yticklabels(self.cell_types)
            self.axs_plot[self.row_spike][i].xaxis.set_major_formatter(ms_to_s)
        self.axs_plot[self.row_spike][-1].legend(loc='upper right')

    def _init_weight_plot(self):
        # Weight plot
        self.axs_plot[self.row_weights][0].set_ylabel('Cortical input weight')
        self.weight_lines = {ct: [[] for _ in range(self.actions_to_plot)] for ct in self.weight_cell_types}
        self.activation_lines = []

        for i, ch in enumerate(range(self.actions_to_plot)):
            for j, ct in enumerate(self.cell_types):
                if ct in self.weight_cell_types:
                    self.weight_lines[ct][i] = []
                    for k in range(self.cell_types_numbers[ct][1]):
                        label = ct if k == 0 else None  # Only label the first line
                        line, = self.axs_plot[self.row_weights][i].step([], [], f'C{j}', label=label, where='post')
                        self.weight_lines[ct][i].append(line)

            activation_line, = self.axs_plot[self.row_weights][i].step([], [], 'C7', linestyle='dashed', label='Activation', where='post')
            self.activation_lines.append(activation_line)

            self.axs_plot[self.row_weights][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_weights][i].set_ylim(-0.1, 1.1)
            self.axs_plot[self.row_weights][i].xaxis.set_major_formatter(ms_to_s)
            self.axs_plot[self.row_weights][i].set_xlabel('Simulation\ntime (s)')
        self.axs_plot[self.row_weights][-1].legend(loc='upper right')

    def _init_reward_plot(self):
        # Reward plot
        self.expected_reward_lines = []
        self.reward_lines = []
        self.dopamine_lines = []

        self.expected_reward_lines, = self.axs_control[-1].plot([], [], 'C9', label='Expected\nreward')
        self.reward_lines, = self.axs_control[-1].step([], [], 'C8', label='Reward', where='post')
        self.dopamine_lines, = self.axs_control[-1].plot([], [], 'C6', label='Dopamine')

        self.axs_control[-1].xaxis.set_major_formatter(ms_to_s)
        self.axs_control[-1].set_xlabel('Simulation time (s)')
        self.axs_control[-1].set_xlim(0, self.plot_interval)
        self.axs_control[-1].set_ylim(-1.1, 1.1)
        self.axs_control[-1].legend(loc='upper left')

    def _init_control_panel(self):
        #--- Upper control panel ---#        
        ax_noise = self.axs_control[0].inset_axes([0.4,0,0.5,0.45]) #[x0, y0, width, height]
        self.buttons['noise_slider'] = Slider(ax_noise, 'Noise', 0, 1, valinit=self.noise, valstep=0.1)
        self.buttons['noise_slider'].on_changed(self.update_stim)

        for i, goal_name in enumerate(self.goals_names):
            ax_cor_dur = self.axs_control[1+i].inset_axes([0,0,0.9,0.45]) #[x0, y0, width, height]
            label_text = '\n'.join(goal_name.split())
            ax_cor_dur.set_title(label_text)
            self.buttons[f'cor_dur_slider{i}'] = Slider(ax_cor_dur, '', 0, 1, valinit=self.cortical_input_dur_rel[i], valstep=1 if self.binary_input else 0.2)
            self.buttons[f'cor_dur_slider{i}'].on_changed(lambda val, i=i: self.update_cor_dur(val=val, goal_idx=i))

    def set_child_loop(self, child_loop):
        self.child_loop = child_loop

    def update_stimulus_activation(self, cell, stimulus, index, active=True):
        i = 0
        for population in range(self.cell_types_numbers[cell][0]):
            for _ in self.cells[cell][population]:
                if population == index:
                    self.ncs[stimulus][i].active(active)
                i += 1

    def SNc_dip(self, event=None, action=None):
        self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=False) # stop SNc tonic stimulus
        h.cvode.event(h.t + self.stim_intervals['SNc'], lambda action=action: self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=True))  # start SNc tonic stimulus

    def SNc_burst(self, event=None, action=None, n_spikes=None):
        self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=False) # stop SNc tonic stimulus
        self.update_stimulus_activation(cell='SNc', stimulus='SNc_burst', index=self.actions.index(action), active=True) # start SNc burst stimulus
        if n_spikes == None:
            n_spikes = self.n_spikes_SNc_burst
        delay = self.stim_intervals['SNc_burst'] * n_spikes
        h.cvode.event(h.t + delay, lambda action=action: self.update_stimulus_activation(cell='SNc', stimulus='SNc_burst', index=self.actions.index(action), active=False))  # stop SNc burst stimulus
        h.cvode.event(h.t + delay, lambda action=action: self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=True))  # start SNc tonic stimulus

    def analyze_firing_rate(self, cell, window=None, average=True):
        """Returns a list of firing rates (Hz) for each action's cell population."""
        current_time = h.t
        rates_avg = []
        rates = []
        if window == None:
            window = self.bin_width_firing_rate
        for a, _ in enumerate(self.actions):
            spikes_avg = 0
            spikes = []
            for i in range(self.cell_types_numbers[cell][1]):
                spike_vec = self.spike_times[cell][a][i]
                # Count spikes in the last `window` ms
                recent_spikes = [t for t in spike_vec if current_time - window <= t <= current_time]
                if average:
                    spikes_avg += len(recent_spikes)
                else:
                    spikes.append(len(recent_spikes))
            if average:
                rate_avg = spikes_avg / (self.cell_types_numbers[cell][1] * (window / 1000.0))  # spikes/sec per neuron
                rates_avg.append(rate_avg)
            else:
                rate = [spike / (window / 1000.0) for spike in spikes]
                rates.append(rate)

        max_rate = 1000.0 / self.stim_intervals[cell]
        
        if average:
            rates_rel_avg = [rate_avg / max_rate for rate_avg in rates_avg]
            return rates_avg, rates_rel_avg
        else:
            rates_rel = [[r / max_rate for r in rate] for rate in rates]
            return rates, rates_rel

    def update_stim(self, val):
        self.noise = val
        for ct in self.cell_types:
            for stim in self.stims[ct]:
                stim.noise = self.noise

    def update_cor_dur(self, val, goal_idx):
        if self._is_updating_programmatically:
            return
        self.cortical_input_dur_rel[goal_idx] = val
        if self.single_goal:
            self.reset_other_durations(goal_idx)

        self.update_selected_goal()

    def reset_other_durations(self, goal_idx):
        self._is_updating_programmatically = True
        for i, _ in enumerate(self.goals_names):
            if i is not goal_idx:
                self.cortical_input_dur_rel[i] = 0
                self.buttons[f'cor_dur_slider{i}'].set_val(0)
        self._is_updating_programmatically = False

    def update_selected_goal(self):
        self.selected_goal = ''.join('1' if val != 0 else '0' for val in self.cortical_input_dur_rel)

    def cortical_input_stimuli(self, current_time, goal=None):
        selected_goal = goal if goal else self.selected_goal
        selected_goal_index = self.goals.index(selected_goal)
        
        for idx, _ in enumerate(self.goals):
            h.cvode.event(current_time + 1, lambda: self.update_stimulus_activation(cell='Cor', stimulus=f'Cor', index=idx, active=False)) # stop all cortical input stimuli
        
        if set(selected_goal) != {'0'}: # Only stimulate cortex if a goal is set
            h.cvode.event(current_time + 1, lambda: self.update_stimulus_activation(cell='Cor', stimulus=f'Cor', index=selected_goal_index, active=True)) # start particular cortical input stimulus
            h.cvode.event(current_time + self.plot_interval, lambda: self.update_stimulus_activation(cell='Cor', stimulus=f'Cor', index=selected_goal_index, active=False)) # stop particular cortical input stimulus

    def analyze_thalamus_activation_time(self, current_time):
        self.activation_times.append(int(current_time))

        if set(self.selected_goal) == {'0'}:
            # Store 0 for all actions
            for action in self.actions:
                self.activation_over_time[action].append(0)
        else:
            rates = {}
            window_start = np.array(self.t_vec.to_python())[-1] - self.plot_interval
            window_end = np.array(self.t_vec.to_python())[-1]
            window_duration_sec = (window_end - window_start) / 1000.0  # Convert ms to seconds

            for action_id, action in enumerate(self.actions):
                # Collect all thalamic spike times for this action
                all_spikes = []
                for k in range(self.cell_types_numbers['Thal'][1]):
                    spikes = np.array(self.spike_times['Thal'][action_id][k].to_python())
                    all_spikes.extend(spikes)

                # Filter spikes within the last window
                spikes_in_window = [spk for spk in all_spikes if window_start < spk <= window_end]

                # Compute firing rate: total spikes / (number of cells * window duration)
                num_cells = self.cell_types_numbers['Thal'][1]
                rate = len(spikes_in_window) / (num_cells * window_duration_sec) if window_duration_sec > 0 else 0
                rates[action] = rate

            # Find the action with the highest thalamus firing rate
            if rates:
                max_value = max(rates.values())
                candidates = [action for action, value in rates.items() if value == max_value]
                best_action = random.choice(candidates)
            else:
                best_action = None

            # Store 1 for the best action, 0 for the rest
            for action in self.actions:
                self.activation_over_time[action].append(1 if action == best_action else 0)
  
    def select_action(self):
        active_actions = [(i, activations[-1]) for i, activations in self.activation_over_time.items() if activations[-1] > 0]
        if active_actions and self.selected_goal and set(self.selected_goal) != {'0'}:
            # Pick the action with the maximum last activation value
            self.selected_action = max(active_actions, key=lambda x: x[1])
        else:
            # No active actions found
            self.selected_action = []

        if self.child_loop and self.selected_action and set(self.selected_goal) != {'0'}:
            # Set input of child loop based on output of current loop
            for id, state in enumerate(list(map(int, self.selected_action[0]))):
                self.child_loop.buttons[f'cor_dur_slider{id}'].set_val(0 if state == 0 else self.selected_action[1])
            self.child_loop.update_selected_goal()
    
    def determine_reward(self, current_time, hw_connected=None, performed_action=None):
        self.reward_times.append(int(current_time))
            
        goal_state = tuple((goal_name, dur != 0) for goal_name, dur in zip(self.goals_names, self.cortical_input_dur_rel))
        target_actions = self.goal_action_table.get(goal_state, {})
        target_action_indices = ''.join('1' if v else '0' for v in target_actions.values())
        for goal in self.goals:
            action_indices = None
            if hw_connected and not self.child_loop and performed_action:
                action_indices = performed_action
            elif self.selected_action:
                action_indices = self.selected_action[0]

            if goal == self.selected_goal and action_indices == target_action_indices:
                self.reward_over_time[goal].append(1)
            else:
                self.reward_over_time[goal].append(0)
                
            current_expected_reward = self.expected_reward_over_time[goal][-1]
            self.dopamine_over_time[goal].append(round(self.reward_over_time[goal][-1] - current_expected_reward, 4)) # TODO: determine dopamine from relative rate of SNc
            
            #'''
            for action in self.actions:
                if self.reward_over_time[goal][-1] - current_expected_reward > 0:
                    self.SNc_burst(event=None, action=action)
                elif self.reward_over_time[goal][-1] - current_expected_reward < 0:
                    self.SNc_dip(event=None, action=action)
            #'''
        
            if goal == self.selected_goal:
                #For selected goal update expected reward based on actual reward
                self.expected_reward_over_time[goal].append(round(current_expected_reward + 0.1 * (self.reward_over_time[goal][-1] - current_expected_reward), 4))
            else:
                self.expected_reward_over_time[goal].append(self.expected_reward_over_time[goal][-1])

    def update_weights(self, current_time, goal=None, action=None):
        # Analyze firing rates
        self.rates['MSNd'], self.rates_rel['MSNd'] = self.analyze_firing_rate('MSNd', window=self.plot_interval, average=False)
        self.rates['MSNi'], self.rates_rel['MSNi'] = self.analyze_firing_rate('MSNi', window=self.plot_interval, average=False)

        self.weight_times.append(int(current_time))
        goal_id = None
        action_id = None
        selected_goal = None
        selected_action = None
        if goal and action: # used for pretraining
            selected_goal = goal
            selected_action = action
        elif self.selected_goal and self.selected_action:
            selected_goal = self.selected_goal
            selected_action = self.selected_action[0]
        if selected_goal:
            goal_id = self.goals.index(selected_goal)
        if selected_action:
            action_id = self.actions.index(selected_action)

        # Only update weights for the selected goal and action
        if goal_id is not None and action_id is not None:
            #print(f"Update weights for goal {selected_goal} with id {goal_id} and action {selected_action} with id {action_id} dopamine {self.dopamine_over_time[selected_goal][-1]}")
            for cor_id in range(self.cell_types_numbers['Cor'][1]):
                for ct in self.weight_cell_types:
                    for msn_id in range(self.cell_types_numbers[ct][1]):
                        delta_w = 0
                        delta_w = self.learning_rate * (self.rates_rel[ct][action_id][msn_id] - 1) * self.dopamine_over_time[selected_goal][-1]
                        if ct == 'MSNd':
                            delta_w = delta_w
                        elif ct == 'MSNi':
                            delta_w = -delta_w
                        key = (ct, action_id, msn_id, goal_id, cor_id)
                        new_weight = min(1, max(0, self.weights_over_time[key][-1] + delta_w))
                        self.weights_over_time[key].append(round(new_weight, 4))
                        idx = self.cor_nc_index_map[f'Cor_to_{ct}'][key]
                        self.ncs[f'Cor_to_{ct}'][idx].weight[0] = new_weight
        # For all other keys, just append the previous value to keep list lengths consistent
        for key in self.weights_over_time:
            if len(self.weights_over_time[key]) < len(self.weight_times):
                self.weights_over_time[key].append(self.weights_over_time[key][-1])

    def update_plots(self, current_time, goal=None, action=None):
        if goal:
            selected_goal_index = self.goals.index(goal)
        else:
            selected_goal_index = self.goals.index(self.selected_goal)
        
        if action:
            sorted_by_recent = [action]+[a for a in self.actions if a is not action]
            actions_to_plot_now = sorted_by_recent[:self.actions_to_plot]
        else:
            # Determine actions to plot (dynamic)
            last_active_indices = { 
                action: (len(act) - 1 - next(i for i, v in enumerate(reversed(act)) if v) if any(act) else -1)
                for action, act in self.activation_over_time.items()}
            sorted_by_recent = sorted(self.actions, key=lambda a: last_active_indices[a], reverse=True)
            actions_to_plot_now = sorted_by_recent[:self.actions_to_plot]

        # Update plots
        for plot_id, action in enumerate(actions_to_plot_now):
            action_id = self.actions.index(action)
            # Membrane potential plot
            for ct in self.cell_types:
                if ct == 'Cor':
                    voltages = np.array([(self.recordings[ct][selected_goal_index][j]) for j in range(self.cell_types_numbers[ct][1])])
                else:
                    voltages = np.array([(self.recordings[ct][action_id][j]) for j in range(self.cell_types_numbers[ct][1])])
                avg_voltage = np.mean(voltages, axis=0)
                self.mem_lines[ct][plot_id].set_data(np.array(self.t_vec), avg_voltage)
                new_xlim = (max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time)))
                if self.axs_plot[self.row_potential][plot_id].get_xlim() != new_xlim:
                    self.axs_plot[self.row_potential][plot_id].set_xlim(*new_xlim)
            self.axs_plot[self.row_potential][plot_id].set_title(action)

            # Spike raster plot
            y_base = self.total_cells
            for ct in self.cell_types:
                all_spikes = []
                if ct == 'Cor':
                    spike_block = self.spike_times[ct][selected_goal_index]
                else:
                    spike_block = self.spike_times[ct][action_id]
                    
                for k in range(self.cell_types_numbers[ct][1]):
                    y_val = y_base - k
                    spikes = np.array(spike_block[k].to_python())
                    y_vals = np.ones_like(spikes) * y_val
                    self.raster_lines[ct][plot_id][k].set_data(spikes, y_vals)
                    all_spikes.extend(spikes)
                # Rate lines
                if all_spikes:
                    bins = np.arange(0, np.array(self.t_vec)[-1] + self.bin_width_firing_rate, self.bin_width_firing_rate)
                    hist, edges = np.histogram(all_spikes, bins=bins)
                    if np.any(hist):  # Only proceed if there's at least one spike
                        rate = hist / (self.cell_types_numbers[ct][1] * self.bin_width_firing_rate / 1000.0)
                        bin_ends = edges[1:]
                        if ct == 'SNc':
                            spike_rate_max = 1000.0 / self.stim_intervals['SNc_burst'] # Hz
                        elif ct == 'MSNd' or ct == 'MSNi':
                            spike_rate_max = 1000.0 / self.stim_intervals['Cor'] # Hz
                        else:
                            spike_rate_max = 1000.0 / self.stim_intervals[ct] # Hz
                        rate_scaled = (rate) / (spike_rate_max + 1e-9)
                        rate_scaled = rate_scaled * (self.cell_types_numbers[ct][1] - 1) + y_base - self.cell_types_numbers[ct][1] + 1
                        self.rate_lines[ct][plot_id].set_data(bin_ends, rate_scaled)
                    else:
                        self.rate_lines[ct][plot_id].set_data([], [])          
                y_base -= self.cell_types_numbers[ct][1]
            new_xlim = max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_spike][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_spike][plot_id].set_xlim(*new_xlim)

            # Weight plot
            for ct in self.weight_cell_types:
                for msn_id in range(self.cell_types_numbers[ct][1]):
                    for cor_id in range(self.cell_types_numbers['Cor'][1]):
                        self.weight_lines[ct][plot_id][msn_id].set_data(self.weight_times, self.weights_over_time[(ct, action_id, msn_id, selected_goal_index, cor_id)])
            self.activation_lines[plot_id].set_data(self.activation_times, self.activation_over_time[action])
            new_xlim = 0, max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_weights][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_weights][plot_id].set_xlim(*new_xlim)

        # Reward plot
        if set(self.selected_goal) != {'0'}:
            self.expected_reward_lines.set_data(self.reward_times, self.expected_reward_over_time[self.selected_goal])
            self.reward_lines.set_data(self.reward_times, self.reward_over_time[self.selected_goal])
            self.dopamine_lines.set_data(self.reward_times, self.dopamine_over_time[self.selected_goal])
        else:
            self.expected_reward_lines.set_data(self.reward_times, [0]*len(self.reward_times))
            self.reward_lines.set_data(self.reward_times, [0]*len(self.reward_times))
            self.dopamine_lines.set_data(self.reward_times, [0]*len(self.reward_times))
        new_xlim = 0, max(self.plot_interval, int(current_time))
        if self.axs_control[-1].get_xlim() != new_xlim:
            self.axs_control[-1].set_xlim(*new_xlim)
    
    def save_data(self, path):
        # Workbook
        wb = Workbook()
        path_extented = f"{path}_{self.name}"

        # Worksheet for General Details
        ws_globals = wb.active
        ws_globals.title = "GlobalVariables"

        row = 1

        # --- Dictionaries ---
        def write_dict(name, data, row):
            ws_globals.cell(row=row, column=1, value=name)
            row += 1
            for k, v in data.items():
                ws_globals.cell(row=row, column=1, value=str(k))
                ws_globals.cell(row=row, column=2, value=v)
                row += 1
            row += 1
            return row

        #row = write_dict("cell_types_numbers", self.cell_types_numbers, row)
        row = write_dict("stim_intervals", self.stim_intervals, row)
        row = write_dict("stim_weights", self.stim_weights, row)

        # --- Lists ---
        def write_list(name, lst, row):
            ws_globals.cell(row=row, column=1, value=name)
            for i, val in enumerate(lst):
                ws_globals.cell(row=row + 1 + i, column=1, value=val)
            row += len(lst) + 2
            return row

        #row = write_list("target_actions", self.target_actions, row)
        row = write_list("cortical_input_dur_rel", self.cortical_input_dur_rel, row)

        # --- Tuples/List of Tuples ---
        def write_tuples(name, tuples_list, row):
            ws_globals.cell(row=row, column=1, value=name)
            for i, tup in enumerate(tuples_list):
                for j, val in enumerate(tup):
                    ws_globals.cell(row=row + 1 + i, column=1 + j, value=val)
            row += len(tuples_list) + 2
            return row

        row = write_tuples("connection_specs", self.connection_specs, row)

        # --- Scalars ---
        scalars = {
            #"N_actions": self.N_actions,
            "plot_interval": self.plot_interval,
            "bin_width_firing_rate": self.bin_width_firing_rate,
            "n_spikes_SNc_burst": self.n_spikes_SNc_burst,
            "learning_rate": self.learning_rate,
            "expected_reward_value": self.expected_reward_value,
            "noise": self.noise

        }
        row = write_dict("Scalars", scalars, row)

        # Worksheet for Weights
        ws_weights = wb.create_sheet(title="WeightsOverTime")

        # Header
        header = ['time']
        keys = sorted(self.weights_over_time.keys())
        header.extend(f"{ct}_a{action_id}_msn{msn_id}_g{goal_id}_cor{cor_id}" for ct, action_id, msn_id, goal_id, cor_id in keys)
        ws_weights.append(header)

        # Weights
        max_len = len(self.weight_times)
        for t_idx in range(max_len):
            row = [self.weight_times[t_idx]]
            for key in keys:
                val = self.weights_over_time[key][t_idx] if t_idx < len(self.weights_over_time[key]) else None
                row.append(val)
            ws_weights.append(row)

        # Worksheets
        data_list = [
            ("ExpectedRewardOverTime", self.expected_reward_over_time), 
            ("RewardOverTime", self.reward_over_time), 
            ("DopamineOverTime", self.dopamine_over_time)
            ]
        ws_list = []
        for name, data in data_list:
            ws = wb.create_sheet(title=f"{name}")
            
            # Header
            header = ['time']
            keys = sorted(data.keys())
            header.extend(f"{key}" for key in keys)
            ws.append(header)

            # Rows
            max_len = len(self.reward_times)
            for t_idx in range(max_len):
                row = [self.reward_times[t_idx]]
                for key in keys:
                    val = data[key][t_idx] if t_idx < len(data[key]) else None
                    row.append(val)
                ws.append(row)

            ws_list.append(ws)

        # Save
        wb.save(f"{path_extented}.xlsx") # Excel
        print(path_extented)

#--- Functions ------------------------------------------------------------------------------------------------------------------------------------------------#

def create_cell(name):
    sec = h.Section(name=name)
    sec.insert('hh')
    return sec

# Stimulus helper
def create_stim(cell, start=0, number=1e9, interval=10, weight=2, noise=0):
    stim = h.NetStim()
    stim.start = start
    stim.number = number
    stim.interval = interval
    stim.noise = noise
    syn = h.ExpSyn(cell(0.5))
    syn.e = 0
    nc = h.NetCon(stim, syn)
    nc.weight[0] = weight
    nc.delay = 1
    return stim, syn, nc

def create_goal_action_table(indices_mapping, goals, actions):
    goal_action_table = {}
    all_goal_combinations = [''.join(bits) for bits in product("01", repeat=len(goals))]

    for goal_combo in all_goal_combinations:
        # Get the corresponding action string, or default to "0" * len(actions)
        action_combo = indices_mapping.get(goal_combo, "0" * len(actions))
        
        # Build the key: a tuple of (goal, True/False) tuples
        key = tuple((goal, bit == "1") for goal, bit in zip(goals, goal_combo))
        
        # Build the value: a dict of {action: True/False}
        value = {action: bit == "1" for action, bit in zip(actions, action_combo)}
        
        # Add to final mapping
        goal_action_table[key] = value

    return goal_action_table

def create_goal_action_indices_mapping(indices, goals, actions):
    goal_action_indices_mapping = {}

    for joint_bits in product('10', repeat=len(goals)):
        if '1' not in joint_bits:
            continue

        actuator_bits = ['0'] * len(actions)
        for joint_index, bit in enumerate(joint_bits):
            if bit == '1':
                for actuator_index in indices[joint_index]:
                    actuator_bits[actuator_index] = '1'

        joint_key = ''.join(joint_bits)
        actuator_value = ''.join(actuator_bits)
        goal_action_indices_mapping[joint_key] = actuator_value

    return goal_action_indices_mapping

# Formatter function: 1000 ms → 1 s
ms_to_s = FuncFormatter(lambda x, _: f'{x/1000}' if x % 100 == 0 else '')


#--- Basal Ganglia ---------------------------------------------------------------------------------------------------------------------------------------------------#

def create_BasalGanglia(no_of_joints=3):
    grasp_types = ["Precision pinch", "Power grasp"]
    if no_of_joints == 4:
        
        joints = ["Thumb opposition", "Thumb flexion", "Index finger flexion", "Remaining fingers flexion"]
        #actuators = ["Thumb oppositor", "Thumb flexor", "Index finger flexor", "Remaining fingers flexor"]
        actuators = ["Thumb oppositor", "Thumb flexor", "Index finger flexor", "Middle finger flexor", "Ring finger flexor", "Pinky finger flexor"]
        grasp_type_joint_indices_mapping = {"10": "1110", # Precision pinch
                                            "01": "1111"} # Power grasp
        joint_actuator_indices = {
            0: [0],
            1: [1],
            2: [2],
            3: [3, 4, 5]
        }

    elif no_of_joints == 6:
        joints = ["Thumb opposition", "Thumb flexion", "Index finger flexion", "Middle finger flexion", "Ring finger flexion", "Pinky finger flexion"]
        actuators = ["Thumb oppositor", "Thumb flexor", "Index finger flexor", "Middle finger flexor", "Ring finger flexor", "Pinky finger flexor"]
        grasp_type_joint_indices_mapping = {"10": "111000", # Precision pinch
                                            "01": "111111"} # Power grasp
        joint_actuator_indices = {
            0: [0],
            1: [1],
            2: [2],
            3: [3],
            4: [4],
            5: [5]
        }
        
    elif no_of_joints == 3:
        joints = ["Thumb flexion", "Index finger flexion", "Middle finger flexion"]
        actuators = ["Thumb flexor", "Index finger flexor", "Middle finger flexor"]
        grasp_type_joint_indices_mapping = {"10": "110", # Precision pinch
                                            "01": "111"} # Power grasp
        joint_actuator_indices = {
            0: [0],
            1: [1],
            2: [2]
        }

    elif no_of_joints == 2:
        joints = ["Thumb flexion", "Index finger flexion"]
        actuators = ["Thumb flexor", "Index finger flexor"]
        grasp_type_joint_indices_mapping = {"10": "11", # Precision pinch
                                            "01": "11"} # Power grasp
        joint_actuator_indices = {
            0: [0],
            1: [1]
        }

    grasp_type_joint_table = create_goal_action_table(indices_mapping=grasp_type_joint_indices_mapping, goals=grasp_types, actions=joints)
    joint_actuator_indices_mapping = create_goal_action_indices_mapping(indices=joint_actuator_indices, goals=joints, actions=actuators)
    joint_actuator_table = create_goal_action_table(indices_mapping=joint_actuator_indices_mapping, goals=joints, actions=actuators)

    bg_m = BasalGangliaLoop('MotorLoop', input=joints, output=actuators, goal_action_table=joint_actuator_table, actions_to_plot=6)
    bg_p = BasalGangliaLoop('PrefrontalLoop', input=grasp_types, output=joints, binary_input=True, single_goal=True, goal_action_table=grasp_type_joint_table, actions_to_plot=6)
    BasalGanglia(loops=[bg_m, bg_p]) # loops ordered from low-level to high-level

create_BasalGanglia(no_of_joints=3)




