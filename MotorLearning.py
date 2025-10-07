from neuron import h
import matplotlib.pyplot as plt
from matplotlib.widgets import Button, Slider, TextBox
from matplotlib import rcParams
import matplotlib.gridspec as gridspec
import numpy as np
import time
import random
import serial
import json
import sys
import pickle
import glob
import os
from openpyxl import Workbook
from datetime import datetime
from itertools import product
from matplotlib.ticker import FuncFormatter


h.load_file("stdrun.hoc")    

#--- MotorLearning ------------------------------------------------------------------------------------------------------------------------------------------------#

class MotorLearning:

    def __init__(self, no_of_joints=6, basal_ganglia_required=True, cerebellum_required=True, user_feedback=False, grasp_types=None, grasp_type_joint_indices_mapping=None, all_joints=None, all_actuators_flexors=None, all_actuators_extensors=None, shuffle_flexors=True):

        self.no_of_joints = no_of_joints
        self.basal_ganglia_required = basal_ganglia_required
        self.cerebellum_required = cerebellum_required
        self.grasp_types = grasp_types
        self.grasp_type_joint_indices_mapping = grasp_type_joint_indices_mapping
        self.all_joints = all_joints
        self.all_actuators_flexors = all_actuators_flexors
        self.all_actuators_extensors = all_actuators_extensors
        self.shuffle_flexors = shuffle_flexors
        self.user_feedback = user_feedback

        self.joints = self.all_joints[:self.no_of_joints]
        self.actuators_flexors = self.all_actuators_flexors[:self.no_of_joints]
        if self.shuffle_flexors: random.shuffle(self.actuators_flexors)
        self.actuators_extensors = self.all_actuators_extensors
        self.joint_actuator_indices = {i: [i] for i in range(self.no_of_joints)}
        
        self.grasp_type_joint_table = create_goal_action_table(indices_mapping=self.grasp_type_joint_indices_mapping, goals=self.grasp_types, actions=self.joints)
        self.joint_actuator_indices_mapping = create_goal_action_indices_mapping(indices=self.joint_actuator_indices, goals=self.joints, actions=self.actuators_flexors)
        self.joint_actuator_table = create_goal_action_table(indices_mapping=self.joint_actuator_indices_mapping, goals=self.joints, actions=self.actuators_flexors)

        self.plot_interval = 100  # ms

        self.buttons = {}
        self.paused = False
        self.hw_connected = False
        self.iteration = 0
        self.durations = {}
        self.start_time = time.time()
        self.stop_time = None
        self.ani = None
        self.interrupt = False

        # HW
        self.ser_sensor = None
        self.ser_exo = None
        self.recorded_sensor_data_flex = []
        self.recorded_sensor_data_touch = []
        self.performed_action = None
        self.num_flex_sensors = 6
        self.num_touch_sensors = 5
        self.flexion_threshold = 30
        self.touch_threshold = 8
        self.touch_overload_threshold = 25
        self.drop_threshold = -5
        self.sensor_analysis_alpha = 0.8
        self.touch_sensor_delta_grasp = [0] * self.num_touch_sensors
        self.touch_sensor_delta_hold = [0] * self.num_touch_sensors
        self.touch_sensor_expected_max_delta_grasp = [self.touch_overload_threshold] * self.num_touch_sensors
        self.touch_sensor_expected_max_delta_hold = [-self.touch_overload_threshold] * self.num_touch_sensors
        self.duration_actuators = 200 # ms
        self.duration_flexors = 1000 # ms
        self.max_duration_flexors = 6000 # ms
        self.delay = 100 # ms
        self.hold_time = 5000 # ms
        self.joint_duration_mapping = {grasp_type: {joint: self.duration_flexors for joint in self.joints} for grasp_type in self.grasp_types + ['None']}
        self.stepper_motor_low = False

        # Sensors
        self.sensor_times = []        
        self.sensor_flex_over_time = {joint_idx: [] for joint_idx in range(len(self.joints))}     
        self.sensor_touch_grasp_over_time = {joint_idx: [] for joint_idx in range(len(self.joints)-1)}       
        self.sensor_touch_hold_over_time = {joint_idx: [] for joint_idx in range(len(self.joints)-1)}
        self.feedback_over_time = {'grasp': [], 'hold': []}

        # Valve - Actuators
        #  1 - Thumb flexor
        #  2 - Thumb extensor
        #  3 - Thumb abductor
        #  4 - Thumb oppositor
        #  5 - Wrist dorsifl
        #  6 - Index finger flexor
        #  7 - Index finger extensor
        #  8 - Middle finger flexor
        #  9 - Middle finger extensor
        # 10 - Ring finger flexor
        # 11 - Ring finger extensor
        # 12 - Pinky finger flexor
        # 13 - Pinky finger extensor

        if self.basal_ganglia_required:
            bg_ml = BasalGangliaLoop('MotorLoop', input=self.joints, output=self.actuators_flexors, goal_action_table=self.joint_actuator_table, actions_to_plot=6, plot_interval=self.plot_interval)
            bg_pl = BasalGangliaLoop('PremotorLoop', input=self.grasp_types, output=self.joints, binary_input=True, single_goal=True, goal_action_table=self.grasp_type_joint_table, actions_to_plot=6, plot_interval=self.plot_interval, child_loop=bg_ml)
            self.loops = [bg_ml, bg_pl] # loops ordered from low-level to high-level
        if self.cerebellum_required:
            self.cerebellum = Cerebellum(grasp_types=self.grasp_types, joints=self.joints, actuators=self.actuators_flexors, plot_interval=self.plot_interval, touch_threshold=self.touch_threshold, touch_overload_threshold=self.touch_overload_threshold, drop_threshold=self.drop_threshold)

        self._init_plotting()
        self._init_simulation()

        self.run_simulation()

    def _init_plotting(self):
        plt.ion()
        self.fig = plt.figure(figsize=(13, 8))
        self.fig.canvas.manager.set_window_title('MotorLearning')
        self.gs = gridspec.GridSpec(2, 1, figure=self.fig, height_ratios=[1,10])  
        self.gs_control = self.gs[0].subgridspec(1, 7, width_ratios=[1,1,1,1,1,0.1,5])
        self.gs_plot = self.gs[1].subgridspec(1, 4, width_ratios=[1,1,10,1])     
        self.gs_loops = self.gs_plot[0].subgridspec(len(self.loops), 1)
        self.gs_names = self.gs_plot[1].subgridspec(len(self.loops) + 1, 1)
        self.gs_selections = self.gs_plot[2].subgridspec(len(self.loops) + 1, 1)
        self.gs_probabilities = self.gs_plot[3].subgridspec(len(self.loops) + 1, 1)

        self.axs_control = {}
        self.axs_selections = {}
        self.axs_loops = {}
        self.loops_texts = {}
        self.axs_names = {}
        self.axs_probabilities = {}
        self.arrows = {}

        # Pause button
        self.axs_control[0] = self.fig.add_subplot(self.gs_control[0])
        self.axs_control[0].set_axis_off() 
        ax_pause = self.axs_control[0].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['pause'] = Button(ax_pause, 'Pause')
        self.buttons['pause'].on_clicked(self.toggle_pause)

        # Reset button
        self.axs_control[1] = self.fig.add_subplot(self.gs_control[1])
        self.axs_control[1].set_axis_off() 
        ax_reset = self.axs_control[1].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['reset'] = Button(ax_reset, 'Reset')
        self.buttons['reset'].on_clicked(self.reset)

        # HW button
        self.axs_control[2] = self.fig.add_subplot(self.gs_control[2])
        self.axs_control[2].set_axis_off() 
        ax_hw = self.axs_control[2].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['hw'] = Button(ax_hw, 'Connect\nHW')
        self.buttons['hw'].on_clicked(self.connect_hw)

        # Save button
        self.axs_control[3] = self.fig.add_subplot(self.gs_control[3])
        self.axs_control[3].set_axis_off() 
        ax_hw = self.axs_control[3].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['save'] = Button(ax_hw, 'Save')
        self.buttons['save'].on_clicked(self.save)

        # Close button
        self.axs_control[4] = self.fig.add_subplot(self.gs_control[4])
        self.axs_control[4].set_axis_off() 
        ax_hw = self.axs_control[4].inset_axes([0,0,1,1]) #[x0, y0, width, height]
        self.buttons['close'] = Button(ax_hw, 'Close')
        self.buttons['close'].on_clicked(self.close)

        # Reward plot
        self.ax_reward = self.fig.add_subplot(self.gs_control[-1])
        self.ax_reward.set_xlabel('Iteration')
        self.ax_reward.set_ylabel('Reward')
        self.ax_reward.set_xlim(0, 10)  # initial limits, will expand dynamically
        self.ax_reward.set_ylim(-0.1, 1.1)
        self.reward_lines = {}  

        # Overlay axes for arrows
        self.ax_overlay = self.fig.add_subplot(111)
        self.ax_overlay.set_axis_off()
        self.ax_overlay.set_xlim(0, 1)
        self.ax_overlay.set_ylim(0, 1)
        self.ax_overlay.set_zorder(10)  # draw on top visually
        self.ax_overlay.set_facecolor("none")  # transparent
        self.ax_overlay.patch.set_alpha(0)     # no background fill
        self.ax_overlay.set_navigate(False)    # don't intercept pan/zoom
        self.ax_overlay.set_picker(False)
        self.ax_overlay.set_navigate(False)
        self.ax_overlay.patch.set_visible(False)

        for idx, loop in enumerate(self.loops):
            row = len(self.loops) - idx

            # Plot loop name
            self.axs_loops[idx] = self.fig.add_subplot(self.gs_loops[self.gs_loops.nrows - 1 - idx ])
            self.axs_loops[idx].set_axis_off() 
            ax_loops = self.axs_loops[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
            self.loops_texts[idx] = ax_loops.text(0.5, 0.5, f'Train\n{loop.name}', rotation=90, ha='center', va='center', transform=ax_loops.transAxes)
            self.buttons[f'{loop.name}'] = Button(ax_loops, label='')
            self.buttons[f'{loop.name}'].on_clicked(lambda event, idx=idx: self.toggle_train_loop(event=event, loop_idx=idx))
            # Load pretrained weights
            ax_loops_pretrained = self.axs_loops[idx].inset_axes([0,-0.2,1,0.2]) #[x0, y0, width, height]
            self.buttons[f'{loop.name}_pretrained'] = Button(ax_loops_pretrained, label='Use pre-\ntrained')
            self.buttons[f'{loop.name}_pretrained'].on_clicked(lambda event, idx=idx: self.load_trained_weights(event=event, loop_idx=idx))
            
            # Init probabilities
            self.axs_probabilities[idx] = self.fig.add_subplot(self.gs_probabilities[self.gs_probabilities.nrows - 2 - idx])
            self.axs_probabilities[idx].set_axis_off()
            self.axs_probabilities[idx].set_xlim(0, 1)
            self.axs_probabilities[idx].set_ylim(0, 1)

            # Init User Feedback
            self.axs_probabilities
            
            if idx == 0:
                # Plot output group name
                self.axs_names[idx] = self.fig.add_subplot(self.gs_names[row])
                self.axs_names[idx].set_axis_off() 
                ax_names = self.axs_names[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                ax_names.set_axis_off()
                output_group = 'Actuators'
                ax_names.text(0.5, 0.5, s=output_group, ha='center', va='center', transform=ax_names.transAxes)

                # Plot outputs of loop
                col = len(loop.actions_names)
                sgs = self.gs_selections[row].subgridspec(1, col)
                self.axs_selections[idx] = [self.fig.add_subplot(sgs[i]) for i in range(col)]
                for i, name in enumerate(loop.actions_names):
                    self.axs_selections[idx][i].set_axis_off() 
                    ax_selections = self.axs_selections[idx][i].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                    label_text = '\n'.join(name.split())
                    self.buttons[f'{name}'] = TextBox(ax_selections, label='', label_pad=0.05, initial=f'{label_text}', color='None', textalignment='center')
                    self.buttons[f'duration_actuator_{name}'] = ax_selections.text(0.5, -0.1, '', ha='center')

                # Plot a horizontal bar with width=prob
                self.buttons[f'probability_bar_{idx}'] = self.axs_probabilities[idx].bar(
                    x=0.5, width=0.8, height=0, color=f'C{idx}')[0]  # Get BarContainer object
                ax_probabilities = self.axs_probabilities[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                ax_probabilities.set_axis_off()
                self.buttons[f'probability_{idx}'] = ax_probabilities.text(0.5, 0, s='', ha='center', va='center', transform=ax_probabilities.transAxes)

            # Plot input group name
            self.axs_names[idx+1] = self.fig.add_subplot(self.gs_names[row-1])
            self.axs_names[idx+1].set_axis_off() 
            ax_names = self.axs_names[idx+1].inset_axes([0,0,1,1]) #[x0, y0, width, height]
            ax_names.set_axis_off()
            input_group = 'Joints' if idx == 0 else 'Grasp\ntype'
            ax_names.text(0.5, 0.5, s=input_group, ha='center', va='center', transform=ax_names.transAxes)

            # Plot inputs of loop  
            col = len(loop.goals_names)
            sgs = self.gs_selections[row-1].subgridspec(1, col)
            self.axs_selections[idx+1] = [self.fig.add_subplot(sgs[i]) for i in range(col)]
            for i, name in enumerate(loop.goals_names):
                self.axs_selections[idx+1][i].set_axis_off() 
                ax_selections = self.axs_selections[idx+1][i].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                label_text = '\n'.join(name.split())
                self.buttons[f'{name}'] = Button(ax_selections, label=f'{label_text}', color='None', hovercolor='lightgray')
                self.buttons[f'{name}'].on_clicked(lambda event, loop=loop, i=i: self.update_goals(loop, i))
                if idx == 0:
                    self.buttons[f'sensor_flex_{name}'] = ax_selections.text(0.5, 0.9, '', ha='center')
                    if i > 0: 
                        self.buttons[f'sensor_touch_{name}'] = ax_selections.text(0.5, 0.13, '', ha='center')
                        self.buttons[f'sensor_drop_{name}'] = ax_selections.text(0.5, 0.03, '', ha='center')
                
            if idx == 1:
                self.buttons[f'probability_bar_{idx}'] = self.axs_probabilities[idx].bar(
                    x=0.5, width=0.8, height=0, color=f'C{idx}')[0]  # Get BarContainer object
                ax_probabilities = self.axs_probabilities[idx].inset_axes([0,0,1,1]) #[x0, y0, width, height]
                ax_probabilities.set_axis_off()
                self.buttons[f'probability_{idx}'] = ax_probabilities.text(0.5, 0, s='', ha='center', va='center', transform=ax_probabilities.transAxes)
            
            # Initialize reward lines
            line, = self.ax_reward.step([], [], label=loop.name, where='post')
            self.reward_lines[loop.name] = {'line': line, 'xdata': [], 'ydata': []}

            # --- Initialize arrows for this loop ---
            if idx == 0:
                # Connect goals -> actions
                for goal in loop.goals_names:
                    for action in loop.actions_names:
                        bbox_start = self.buttons[goal].ax.get_position()
                        x_start, y_start = bbox_start.x0 + 0.5 * (bbox_start.x1 - bbox_start.x0), bbox_start.y0
                        bbox_end = self.buttons[action].ax.get_position()
                        x_end, y_end = bbox_end.x0 + 0.5 * (bbox_end.x1 - bbox_end.x0), bbox_end.y1
                        arrow = self.ax_overlay.annotate(
                            "",
                            xy=(x_end, y_end), xycoords='figure fraction', # arrow tip (action top)
                            xytext=(x_start, y_start), textcoords='figure fraction', # arrow tail (goal bottom)
                            arrowprops=dict(
                                arrowstyle="->", 
                                color=f"C{idx}", 
                                lw=2, alpha=0.6
                            ),
                            visible=False
                        )
                        self.arrows[(goal, action)] = arrow
            else:
                # Connect current loop goals -> previous loop goals
                prev_loop = self.loops[idx-1]
                for goal in loop.goals_names:
                    for prev_goal in prev_loop.goals_names:
                        bbox_start = self.buttons[goal].ax.get_position()
                        x_start, y_start = bbox_start.x0 + 0.5 * (bbox_start.x1 - bbox_start.x0), bbox_start.y0
                        bbox_end = self.buttons[prev_goal].ax.get_position()
                        x_end, y_end = bbox_end.x0 + 0.5 * (bbox_end.x1 - bbox_end.x0), bbox_end.y1
                        arrow = self.ax_overlay.annotate(
                            "",
                            xy=(x_end, y_end), xycoords='figure fraction',
                            xytext=(x_start, y_start), textcoords='figure fraction',
                            arrowprops=dict(
                                arrowstyle="->", 
                                color=f"C{idx}", 
                                lw=2, alpha=0.6
                            ),
                            visible=False
                        )
                        self.arrows[(goal, prev_goal)] = arrow
        
        self.ax_reward.legend(loc='upper left')
        self.update_arrows()

    def _init_simulation(self):
        h.dt = 1
        h.finitialize()

    class UserFeedback:
        def __init__(self):
            self.result = {"grasp": None, "hold": None}
            self._ask()

        def _ask(self):
            fig = plt.figure(figsize=(6, 4))
            fig.canvas.manager.set_window_title('UserFeedback')
            gs = gridspec.GridSpec(5, 2, figure=fig)

            # Question labels and result texts
            ax_grasp_label = fig.add_subplot(gs[0, 0])
            ax_grasp_label.axis('off')
            ax_grasp_label.text(0, 0.5, "Object grasped?", fontsize=12, weight='bold', verticalalignment='center')
            ax_grasp_result = fig.add_subplot(gs[0, 1])
            ax_grasp_result.axis('off')
            grasp_result_text = ax_grasp_result.text(0, 0.5, "-", fontsize=12, color='blue', verticalalignment='center')

            ax_hold_label = fig.add_subplot(gs[2, 0])
            ax_hold_label.axis('off')
            ax_hold_label.text(0, 0.5, "Object held?", fontsize=12, weight='bold', verticalalignment='center')
            ax_hold_result = fig.add_subplot(gs[2, 1])
            ax_hold_result.axis('off')
            hold_result_text = ax_hold_result.text(0, 0.5, "-", fontsize=12, color='blue', verticalalignment='center')

            # Buttons
            b_buttons = {}  # store button objects
            # Grasp buttons
            ax_grasp_yes = fig.add_subplot(gs[1, 0])
            b_grasp_yes = Button(ax_grasp_yes, "Yes")
            ax_grasp_no = fig.add_subplot(gs[1, 1])
            b_grasp_no = Button(ax_grasp_no, "No")
            b_buttons["grasp"] = {"Yes": b_grasp_yes, "No": b_grasp_no}

            # Hold buttons
            ax_hold_yes = fig.add_subplot(gs[3, 0])
            b_hold_yes = Button(ax_hold_yes, "Yes")
            ax_hold_no = fig.add_subplot(gs[3, 1])
            b_hold_no = Button(ax_hold_no, "No")
            b_buttons["hold"] = {"Yes": b_hold_yes, "No": b_hold_no}

            # Confirm button
            ax_confirm = fig.add_subplot(gs[4, :])
            b_confirm = Button(ax_confirm, "Confirm")
            b_confirm.eventson = False  # disabled initially
            ax_confirm.set_visible(False)

            # Helper to update confirm button
            def update_confirm():
                if self.result["grasp"] is not None and self.result["hold"] is not None:
                    ax_confirm.set_visible(True)
                    b_confirm.eventson = True
                else:
                    ax_confirm.set_visible(False)
                    b_confirm.eventson = False
                fig.canvas.draw_idle()

            # General command function for Yes/No buttons
            def command(question_key, answer):
                self.result[question_key] = True if answer == "Yes" else False
                # Update result text
                if question_key == "grasp":
                    grasp_result_text.set_text(answer)
                else:
                    hold_result_text.set_text(answer)
                # Color the clicked button, reset the other
                for a, b in b_buttons[question_key].items():
                    if a == answer:
                        b.color = 'green' if answer == "Yes" else 'red'
                    else:
                        b.color = '0.85'
                update_confirm()

            # Assign command to buttons
            for key in ["grasp", "hold"]:
                for ans in ["Yes", "No"]:
                    b_buttons[key][ans].on_clicked(lambda event, k=key, a=ans: command(k, a))

            # Confirm button callback
            def confirm(event):
                plt.close(fig)
            
            b_confirm.on_clicked(confirm)

            plt.show(block=False)

            # Wait here until popup closed
            while plt.fignum_exists(fig.number):
                fig.canvas.draw_idle()
                fig.canvas.flush_events()
                time.sleep(0.01)
        
        def get_result(self):
            return self.result

    def toggle_pause(self, event=None):
        self.paused = not self.paused
        self.buttons['pause'].label.set_text('Continue' if self.paused else 'Pause')
        if not self.paused:
            self.buttons['pause'].ax.set_facecolor(rcParams['axes.facecolor'])
            self.buttons['pause'].color = '0.85'
        else:
            self.buttons['pause'].ax.set_facecolor('r')
            self.buttons['pause'].color = 'r'
            
    def reset(self, event=None):
        self.buttons['reset'].label.set_text('Reseting...')
        plt.draw()
        plt.pause(0.001)

        self.close(restart=True)
        self.__init__(self.no_of_joints, self.basal_ganglia_required, self.cerebellum_required, self.grasp_types, self.grasp_type_joint_indices_mapping, self.all_joints, self.all_actuators_flexors, self.all_actuators_extensors, self.shuffle_flexors)

    def connect_hw(self, event=None):
        successful = True

        if not self.hw_connected:
            self.buttons['hw'].label.set_text('Connecting...')
            plt.draw()
            plt.pause(0.001)

            # Define serial connection for sensor feedback
            if self.ser_sensor:
                if self.ser_sensor.is_open:
                    print(f"Serial connection to sensors already open on port {self.ser_sensor.port}")
                else:
                    try:
                        self.ser_sensor.open()
                        print(f"Serial connection to sensors opened on port {self.ser_sensor.port}")
                    except Exception as e: 
                        successful = False
                        print(f"Serial connection to sensors failed due to exception: {e}")
            else:
                try:
                    self.ser_sensor = serial.Serial(
                        port='COM7',        # For ESP32
                        baudrate=115200,    # Baud rate for serial connection to sensors
                        timeout=1           # Timeout for read in seconds
                    )
                    print(f"Serial connection to sensors established on port {self.ser_sensor.port}")  
                except Exception as e:
                    successful = False
                    print(f"Serial connection to sensors failed due to exception: {e}")
            
            # Define serial connection via bluetooth to exoskeleton
            if self.ser_exo:
                if self.ser_exo.is_open:
                    print(f"Serial connection to exoskeleton already open on port {self.ser_exo.port}")
                else:
                    try:
                        self.ser_exo.open()
                        print(f"Serial connection to exoskeleton opened on port {self.ser_exo.port}")
                    except Exception:
                        successful = False
                        print(f"Serial connection to exoskeleton failed due to exception: {e}")
            else:
                try:
                    self.ser_exo = serial.Serial(
                        port='COM11',       # For HC-06
                        baudrate=9600,      # Baud rate for connection to bluetooth module
                        write_timeout=5     # Timeout for read in seconds
                    )
                    print(f"Serial connection to exoskeleton established on port {self.ser_exo.port}")
                except Exception as e:
                    successful = False
                    print(f"Serial connection to exoskeleton failed due to exception: {e}")
        else:
            self.buttons['hw'].label.set_text('Disconnecting...')
            plt.draw()
            plt.pause(0.001)

            try:
                self.ser_sensor.close()
            except Exception: None
            try:
                self.ser_exo.close()
            except Exception: None
            print("Serial connections closed")

        if successful:
            self.hw_connected = not self.hw_connected
        
        self.buttons['hw'].label.set_text('Disconnect\nHW' if self.hw_connected else 'Connect\nHW')

    def toggle_train_loop(self, event=None, loop_idx=None):
        self.loops[loop_idx].training = not self.loops[loop_idx].training
        if self.loops[loop_idx].training:
            if self.loops[loop_idx].training_start_iteration == -1:
                self.loops[loop_idx].training_start_iteration = self.iteration
        else:
            self.loops_texts[loop_idx].set_text(f'{self.loops[loop_idx].name}\ntrained')
            self.loops[loop_idx].training_stop_iteration = self.iteration
            self.loops[loop_idx].save_trained_weights()
        self.loops[loop_idx].selected_goal = ''.join('0' for _ in self.loops[loop_idx].goals_names)
        self.loops[loop_idx].selected_action = None
        for id, state in enumerate(self.loops[loop_idx].selected_goal):
            self.loops[loop_idx].buttons[f'cor_dur_slider{id}'].set_val(0 if state == '0' else 1)
        self.training_goal_idx = 0

        if not self.loops[loop_idx].training:
            self.buttons[f'{self.loops[loop_idx].name}'].ax.set_facecolor(rcParams['axes.facecolor'])
            self.buttons[f'{self.loops[loop_idx].name}'].color = '0.85'
        else:
            self.buttons[f'{self.loops[loop_idx].name}'].ax.set_facecolor(f'C{loop_idx}')
            self.buttons[f'{self.loops[loop_idx].name}'].color = f'C{loop_idx}'

    def load_trained_weights(self, event=None, loop_idx=None):
        self.buttons[f'{self.loops[loop_idx].name}_pretrained'].label.set_text('Loading\npretrained...')
        plt.draw()
        plt.pause(0.001)

        self.loops[loop_idx].load_trained_weights()
        self.buttons[f'{self.loops[loop_idx].name}_pretrained'].ax.set_visible(False)
        self.loops_texts[loop_idx].set_text(f'{self.loops[loop_idx].name}\npretrained')
        self.update_arrows()

    def read_sensor_data(self, duration=5000, reset=True):
        if reset:
            self.ser_sensor.flushInput()         # delete values in serial input buffer
            self.recorded_sensor_data_flex = []   # Stores all flex sensor readings
            self.recorded_sensor_data_touch = []  # Stores all touch sensor readings

        start_time = time.time()

        while time.time() - start_time < duration / 1000:
            if self.ser_sensor.in_waiting > 0:
                line = self.ser_sensor.readline().decode('utf-8', errors='ignore').strip()
                try:
                    values = [int(x) for x in line.split(',')]
                    self.recorded_sensor_data_flex.append(values[:self.num_flex_sensors])
                    self.recorded_sensor_data_touch.append(values[self.num_flex_sensors:self.num_flex_sensors+self.num_touch_sensors])
                except ValueError:
                    if line != "M:done":
                        print(f"Ignored malformed line: {line}")

        for key, values in self.sensor_flex_over_time.items():
            missing_values = len(self.sensor_times) - len(values)
            if missing_values > 0:
                values.extend([None] * missing_values)

        for key, values in self.sensor_touch_grasp_over_time.items():
            missing_values = len(self.sensor_times) - len(values)
            if missing_values > 0:
                values.extend([None] * missing_values)
        
        for key, values in self.sensor_touch_hold_over_time.items():
            missing_values = len(self.sensor_times) - len(values)
            if missing_values > 0:
                values.extend([None] * missing_values)

        for key, values in self.feedback_over_time.items():
            missing_values = len(self.sensor_times) - len(values)
            if missing_values > 0:
                values.extend([None] * missing_values)

        if not self.sensor_times or int(h.t) != self.sensor_times[-1]:
            self.sensor_times.append(int(h.t))
    
    def analyze_sensor_data_flex(self):
        # Remove invalid data
        self.recorded_sensor_data_flex = [row for row in self.recorded_sensor_data_flex
            if len(row) == self.num_flex_sensors and all(int(val) != 0 for val in row)]

        # Remove the first valid row
        if self.recorded_sensor_data_flex:
            self.recorded_sensor_data_flex = self.recorded_sensor_data_flex[1:]

        if self.recorded_sensor_data_flex:
            start = [self.recorded_sensor_data_flex[0][i] for i in range(self.num_flex_sensors)]
            prev_filtered = start.copy()
            max_filtered = prev_filtered.copy()

            flexion_detected = [False] * self.num_flex_sensors

            for sample in self.recorded_sensor_data_flex[1:]:
                for i in range(self.num_flex_sensors):
                    # Apply low-pass filter
                    filtered = self.sensor_analysis_alpha * prev_filtered[i] + (1 - self.sensor_analysis_alpha) * sample[i]

                    # Track max
                    max_filtered[i] = max(max_filtered[i], filtered)

                    prev_filtered[i] = filtered

            for i, name in enumerate(self.loops[0].goals_names):
                delta = int(max_filtered[i] - start[i])
                if delta > self.flexion_threshold:
                    flexion_detected[i] = True
                text = f"{delta} {'Flex' if flexion_detected[i] else ''}"
                self.buttons[f'sensor_flex_{name}'].set_text(text)

                self.sensor_flex_over_time[i].append(delta)
            
            self.performed_action = ''.join(['1' if value else '0' for value in flexion_detected])
            self.performed_action = self.performed_action[:len(self.loops[0].actions_names)]
        else:
            print("Not enough data from flex sensors")
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
        plt.pause(0.001)  # tiny pause lets GUI update

    def analyze_sensor_data_touch(self, grasp=False, hold=False):
        # Remove invalid data
        self.recorded_sensor_data_touch = [row for row in self.recorded_sensor_data_touch
            if len(row) == self.num_touch_sensors and all(int(val) != 0 for val in row)]

        if self.recorded_sensor_data_touch:
            N = min(5, len(self.recorded_sensor_data_touch))
            start = np.median(self.recorded_sensor_data_touch[:N], axis=0).tolist()
            end = np.median(self.recorded_sensor_data_touch[-N:], axis=0).tolist()

            for i, name in enumerate(self.loops[0].goals_names[1:]):
                delta = int(end[i] - start[i])
                text = f"{delta} "
                if grasp:
                    self.touch_sensor_delta_grasp[i] = delta
                    self.touch_sensor_expected_max_delta_grasp[i] = 0.5 * self.touch_sensor_expected_max_delta_grasp[i] + 0.5 * self.touch_sensor_delta_grasp[i]
                    
                    if self.touch_sensor_delta_grasp[i] >= self.touch_overload_threshold:
                        text += 'Overload'
                    elif self.touch_sensor_delta_grasp[i] >= self.touch_threshold:
                        text += 'Grasp'
                    self.buttons[f'sensor_touch_{name}'].set_text(text)

                    self.sensor_touch_grasp_over_time[i].append(delta)

                if hold:
                    self.touch_sensor_delta_hold[i] = delta
                    self.touch_sensor_expected_max_delta_hold[i] = 0.5 * self.touch_sensor_expected_max_delta_hold[i] + 0.5 * self.touch_sensor_delta_hold[i]

                    if self.touch_sensor_delta_hold[i] <= self.drop_threshold:
                        text += 'Drop' 
                    else:
                        text += 'Hold' 
                    self.buttons[f'sensor_drop_{name}'].set_text(text)

                    self.sensor_touch_hold_over_time[i].append(delta)
        else:
            print("Not enough data from touch sensors")
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
        plt.pause(0.001)  # tiny pause lets GUI update

    def update_GUI_goals_actions(self, frame=None):       
        for idx, loop in enumerate(self.loops):
            if loop.selected_goal:
                visibility = False if set(loop.selected_goal) == {'0'} else True
                self.update_probability(loop_id=idx, probability=loop.expected_reward_over_time[loop.selected_goal][-1], visibility=visibility)

            if not loop.child_loop:
                for i, name in enumerate(loop.actions_names):
                    if loop.selected_goal:
                        char = '0'
                        if loop.selected_action:
                            char = loop.selected_action[0][i]
                        selected = char == '1'
                        reward = loop.reward_over_time[loop.selected_goal][-1]
                        self.highlight_textbox(f"{name}", selected, reward)

                    joint = next((g for g, a in loop.learned_goal_action_map.items() if name in a), None)
                    grasp_type = self.grasp_types[self.loops[1].selected_goal.index('1')] if '1' in self.loops[1].selected_goal else 'None'
                    duration = self.joint_duration_mapping[grasp_type].get(joint, 0) if joint else self.duration_flexors
                    self.buttons[f'duration_actuator_{name}'].set_text(f'{duration} ms' if duration else '')
            elif loop.training:
                for i, name in enumerate(loop.actions_names):
                    if loop.selected_goal:
                        char = '0'
                        if loop.selected_action:
                            char = loop.selected_action[0][i]
                        selected = char == '1'
                        reward = loop.reward_over_time[loop.selected_goal][-1]
                        name = self.loops[idx-1].goals_names[i]
                        self.highlight_textbox(f"{name}", selected, reward)
            
            for i, name in enumerate(loop.goals_names):
                if not loop.child_loop: # reset sensor reading
                    self.buttons[f'sensor_flex_{name}'].set_text('')
                    if i > 0:
                        self.buttons[f'sensor_touch_{name}'].set_text('')
                        self.buttons[f'sensor_drop_{name}'].set_text('')

                if loop.selected_goal:
                    char = loop.selected_goal[i]
                    selected = char == '1'
                    reward = None
                    
                    if idx < len(self.loops) - 1:
                        if not self.loops[idx+1].training:
                            if self.loops[idx+1].selected_goal:
                                reward = self.loops[idx+1].reward_over_time[self.loops[idx+1].selected_goal][-1]
                            self.highlight_textbox(name, selected, reward)
                    else:
                        self.highlight_textbox(name, selected, reward)
                    
                    if selected: self.update_goal_probability(name, loop, probability=loop.expected_reward_over_time[loop.selected_goal][-1]) 
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
        
    def update_GUI_performed_action_reward(self, frame=None):
        self.iteration = int(h.t / self.plot_interval)
        
        for idx, loop in enumerate(self.loops):
            # Disable pretraining
            if self.iteration > 0 and self.buttons[f'{loop.name}_pretrained'].ax.get_visible:
                self.buttons[f'{loop.name}_pretrained'].ax.set_visible(False)

            if not loop.child_loop:
                for i, name in enumerate(loop.actions_names):
                    if loop.selected_goal:
                        char = '0'
                        if loop.selected_action:
                            char = loop.selected_action[0][i]
                        selected = char == '1'
                        reward = loop.reward_over_time[loop.selected_goal][-1]
                        if reward: 
                            self.highlight_textbox(f"{name}", selected, reward)
            elif loop.training:
                for i, name in enumerate(loop.actions_names):
                    if loop.selected_goal:
                        char = '0'
                        if loop.selected_action:
                            char = loop.selected_action[0][i]
                        selected = char == '1'
                        reward = loop.reward_over_time[loop.selected_goal][-1]
                        name = self.loops[idx-1].goals_names[i]
                        self.highlight_textbox(f"{name}", selected, reward)
            for i, name in enumerate(loop.goals_names):
                if loop.selected_goal:
                    char = loop.selected_goal[i]
                    selected = char == '1'
                    reward = None
                    
                    if idx < len(self.loops) - 1:
                        if not self.loops[idx+1].training:
                            if self.loops[idx+1].selected_goal:
                                reward = self.loops[idx+1].reward_over_time[self.loops[idx+1].selected_goal][-1]
                            self.highlight_textbox(name, selected, reward)
                    else:
                        self.highlight_textbox(name, selected, reward)
                    
            # Get reward for currently selected goal
            if loop.selected_goal:
                current_reward = loop.reward_over_time[loop.selected_goal][-1]
            else:
                current_reward = 0

            rl = self.reward_lines[loop.name]
            rl['xdata'].append(self.iteration)
            rl['ydata'].append(current_reward)
            rl['line'].set_data(rl['xdata'], rl['ydata'])
        # Update axes limits dynamically
        max_iter = max(rl['xdata'][-1] for rl in self.reward_lines.values() if rl['xdata'])
        self.ax_reward.set_xlim(0, max(10, max_iter + 1))
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
        
    def highlight_textbox(self, name, selected, reward=None):
        if selected:
            if reward:
                self.buttons[f'{name}'].ax.set_facecolor('tab:olive')
                self.buttons[f'{name}'].color = 'tab:olive'
            else:
                self.buttons[f'{name}'].ax.set_facecolor('gold')
                self.buttons[f'{name}'].color = 'gold'
        else:
            self.buttons[f'{name}'].ax.set_facecolor(rcParams['axes.facecolor'])
            self.buttons[f'{name}'].color = 'None'
        
        self.fig.canvas.draw_idle()
        self.fig.canvas.flush_events()
   
    def update_goals(self, loop, goal_idx):
        if loop.selected_goal:
            value = loop.selected_goal[goal_idx]=='0'
        else:
            value = 1
        loop.buttons[f'cor_dur_slider{goal_idx}'].set_val(value)
        loop.update_cor_dur(val=value, goal_idx=goal_idx)
        self.update_GUI_goals_actions()

    def update_probability(self, loop_id, probability, visibility):
        if visibility:
            self.buttons[f'probability_{loop_id}'].set_text(f'{probability:.1%}')
            self.buttons[f'probability_{loop_id}'].set_position((0.5, probability + 0.03))
            self.buttons[f'probability_bar_{loop_id}'].set_height(probability)
        else:
            self.buttons[f'probability_{loop_id}'].set_text(f'')
            self.buttons[f'probability_bar_{loop_id}'].set_height(0)
    
    def update_goal_probability(self, name, loop, probability):
        if loop.single_goal:
            if self.buttons.get(f'probability_{name}'):
                self.buttons[f'probability_{name}'].set_text(f'{probability:.1%}')
    
    def generate_action_command(self):
        action = self.loops[0].selected_action[0]  
        selected_actions_names = [name for bit, name in zip(action, self.actuators_flexors) if bit == "1"]
        for name in selected_actions_names:
            joint = next((g for g, a in self.loops[0].learned_goal_action_map.items() if name in a), None)
            grasp_type = self.grasp_types[self.loops[1].selected_goal.index('1')] if '1' in self.loops[1].selected_goal else 'None'
            duration = self.joint_duration_mapping[grasp_type].get(joint, 0) if joint else self.duration_flexors
            self.buttons[f'duration_actuator_{name}'].set_text(f'{duration} ms' if duration else '')
        
        action_command_parts = []

        # Control all actuators
        for actuator in self.actuators_extensors + self.actuators_flexors:
            actuator_number = actuator.split()[-1]
            action_command_parts.append(f"0-{actuator_number}-i")
        for actuator in self.actuators_extensors + self.actuators_flexors:
            actuator_number = actuator.split()[-1]
            action_command_parts.append(f"{self.duration_actuators}-{actuator_number}-h")

        # Control individual flexors based on selected action
        for i, bit in enumerate(action):
            if bit == '1':
                flexor = self.actuators_flexors[i]
                flexor_number = flexor.split()[-1]
                action_command_parts.append(f"{(self.delay + self.duration_actuators)}-{flexor_number}-i")

        max_flexion_duration = 0
        for i, bit in enumerate(action):
            if bit == '1':
                flexor = self.actuators_flexors[i]
                joint = next((g for g, a in self.loops[0].learned_goal_action_map.items() if flexor in a), None)
                grasp_type = self.grasp_types[self.loops[1].selected_goal.index('1')] if '1' in self.loops[1].selected_goal else 'None'
                duration = self.joint_duration_mapping[grasp_type].get(joint, 0) if joint else self.duration_flexors

                if duration and duration > max_flexion_duration:
                    max_flexion_duration = duration
                flexor_number = flexor.split()[-1]
                # Hold position
                action_command_parts.append(f"{(self.delay + self.duration_actuators + duration)}-{flexor_number}-h")

        action_command = '/'.join(action_command_parts) + '/'

        max_duration = self.delay + self.duration_actuators + max_flexion_duration
        
        return action_command, max_duration

    def move_object(self, down=True, length=40, blocking=False, timeout=10):
        dir = -1 if down else 1
        stepper_motor_command = f'M:{dir*length}'
        self.ser_sensor.write(stepper_motor_command.encode())

        if blocking:
            start_time = time.time()
            stepper_done = False

            while time.time() - start_time < timeout:
                if self.ser_sensor.in_waiting:
                    line = self.ser_sensor.readline().decode().strip()
                    if "M:done" in line:
                        stepper_done = True
                        break
                time.sleep(0.01)  # avoid busy waiting

            if not stepper_done:
                print("Warning: Stepper motor did not reach target position within timeout!")

        self.stepper_motor_low = down


    def perform_action(self, grasping=False):
        max_duration = 0
        
        if self.loops[0].selected_action:
            action_command, max_duration = self.generate_action_command()
            
        if max_duration: # perform action    
            # Send action command
            self.ser_exo.write(action_command.encode())

            # Send start command
            self.ser_exo.write('Start\n'.encode())
            time.sleep((self.duration_actuators+self.delay)/1000)

            # Read sensor data during grasping
            self.read_sensor_data(duration=max_duration)
            if self.recorded_sensor_data_flex: self.analyze_sensor_data_flex()
            
            if grasping:
                if self.recorded_sensor_data_touch: self.analyze_sensor_data_touch(grasp=True)

                self.move_object(down=True)

                # Read sensor data during holding
                self.read_sensor_data(duration=self.hold_time)          
                if self.recorded_sensor_data_touch: self.analyze_sensor_data_touch(hold=True)
            else:
                time.sleep(1)
            
            # Send relax command
            self.ser_exo.write('Stop\n'.encode())

            if grasping and self.stepper_motor_low == True:
                if self.user_feedback:
                    result = self.UserFeedback().get_result()
                    self.feedback_over_time['grasp'].append(int(result['grasp']))
                    self.feedback_over_time['hold'].append(int(result['hold']))
                
                self.move_object(down=False, blocking=True)
    
    def update_joint_duration_mapping(self, time_correction):
        grasp_type = self.grasp_types[self.loops[1].selected_goal.index('1')] if '1' in self.loops[1].selected_goal else 'None'
        for joint_id, corr in enumerate(time_correction):
            if abs(corr) >= 10: 
                self.joint_duration_mapping[grasp_type][self.joints[joint_id]] += int(corr) * 10
                self.joint_duration_mapping[grasp_type][self.joints[joint_id]] = min(self.max_duration_flexors, max(0, self.joint_duration_mapping[grasp_type][self.joints[joint_id]]))

    def calculate_goal_action_mapping(self, weights_over_time, selected_goal_idx=None):
        sums_counts = {}
        pre_post_pop_mapping = {}
        best_avg = {}

        for (post_group, population_post_cell, post_cell_id, population_pre_cell, pre_cell_id), weight in weights_over_time.items():
            if population_pre_cell == selected_goal_idx if selected_goal_idx is not None else True:
                if post_group != "MSNd":
                    continue
                
                key = (population_pre_cell, population_post_cell)
                
                # keep running sums and counts to compute averages on the fly
                if key not in sums_counts:
                    sums_counts[key] = [0.0, 0]  # [sum, count]
                sums_counts[key][0] += weight[-1]
                sums_counts[key][1] += 1

        # now figure out max averages in one go
        for (pre_pop, post_pop), (s, c) in sums_counts.items():
            avg_w = s / c
            if pre_pop not in pre_post_pop_mapping:
                pre_post_pop_mapping[pre_pop] = post_pop
                best_avg[pre_pop] = avg_w
            else:
                current_max = best_avg[pre_pop]
                if avg_w > current_max:
                    pre_post_pop_mapping[pre_pop] = post_pop
                    best_avg[pre_pop] = avg_w
        
        return pre_post_pop_mapping

    def update_arrows(self):
        # Hide all arrows
        for arrow in self.arrows.values():
            arrow.set_visible(False)

        # Show only learned mappings
        for idx, loop in enumerate(self.loops):
            for goal, targets in loop.learned_goal_action_map.items():
                for target in targets:
                    key = (goal, target)
                    if key in self.arrows:
                        self.arrows[key].set_visible(True)

        self.fig.canvas.draw_idle()

    def save_data(self, path):
        # Workbook
        wb = Workbook()
        path_extented = f"{path}_MotorLearning"

        # Worksheet for General Details
        ws_globals = wb.active
        ws_globals.title = "GlobalVariables"

        row = 1
        row = write_list(ws_globals, "grasp_types", self.grasp_types, row)
        row = write_dict(ws_globals, "grasp_type_joint_indices_mapping", self.grasp_type_joint_indices_mapping, row)
        row = write_list(ws_globals, "all_joints", self.all_joints, row)
        row = write_list(ws_globals, "all_actuators_flexors", self.all_actuators_flexors, row)
        row = write_list(ws_globals, "all_actuators_extensors", self.all_actuators_extensors, row)
        row = write_list(ws_globals, "joints", self.joints, row)
        row = write_list(ws_globals, "actuators_flexors", self.actuators_flexors, row)
        row = write_list(ws_globals, "actuators_extensors", self.actuators_extensors, row)
        row = write_dict(ws_globals, "joint_actuator_indices", self.joint_actuator_indices, row)
        row = write_dict(ws_globals, "grasp_type_joint_table", self.grasp_type_joint_table, row)
        row = write_dict(ws_globals, "joint_actuator_indices_mapping", self.joint_actuator_indices_mapping, row)
        row = write_dict(ws_globals, "joint_actuator_table", self.joint_actuator_table, row)
        row = write_dict(ws_globals, "joint_duration_mapping", self.joint_duration_mapping, row)
        
        # --- Scalars ---
        scalars = {
            "no_of_joints": self.no_of_joints,
            "basal_ganglia_required": self.basal_ganglia_required,
            "cerebellum_required": self.cerebellum_required,
            "plot_interval": self.plot_interval,
            "duration_actuators": self.duration_actuators,
            "duration_flexors": self.duration_flexors,
            "max_duration_flexors": self.max_duration_flexors,
            "delay": self.delay,
            "hold_time": self.hold_time,
            "iteration": self.iteration,
            "start_time": self.start_time,
            "stop_time": self.stop_time,
            "sensor_analysis_alpha": self.sensor_analysis_alpha,
            "flexion_threshold": self.flexion_threshold,
            "touch_threshold": self.touch_threshold,
            "touch_overload_threshold": self.touch_overload_threshold,
            "drop_threshold": self.drop_threshold          
        }
        row = write_dict(ws_globals, "Scalars", scalars, row)

        # Worksheets
        data_list = [
            ("durations", self.durations)
            ]
        ws_list = []
        for name, data in data_list:
            ws = wb.create_sheet(title=f"{name}")
            
            # Header
            header = ['Iteration', 'Duration']
            keys = sorted(data.keys())
            ws.append(header)

            # Rows
            for key in keys:
                row = [key, data[key]]
                ws.append(row)

            ws_list.append(ws)

        # Worksheets
        data_list = [
            ("sensor_flex_over_time", self.sensor_flex_over_time), 
            ("sensor_touch_grasp_over_time", self.sensor_touch_grasp_over_time),
            ("sensor_touch_hold_over_time", self.sensor_touch_hold_over_time),
            ("feedback_over_time", self.feedback_over_time)
            ]
        ws_list = []
        for name, data in data_list:
            ws = wb.create_sheet(title=f"{name}")
            
            # Header
            header = ['time']
            keys = sorted(data.keys())
            header.extend(f"{key}" for key in keys)
            ws.append(header)

            # Rows
            max_len = len(self.sensor_times)
            for t_idx in range(max_len):
                row = [self.sensor_times[t_idx]]
                for key in keys:
                    val = data[key][t_idx] if t_idx < len(data[key]) else None
                    row.append(val)
                ws.append(row)

            ws_list.append(ws)

        # Save
        wb.save(f"{path_extented}.xlsx") # Excel
        print(path_extented)

    def get_current_average_weight_diff(self, loop):
        current_avg_weight = {}
        current_avg_weight_diff = 0

        selected_goal = loop.selected_goal if loop.selected_goal else None
        selected_action = loop.selected_action[0] if loop.selected_action else None
        goal_id = loop.goals.index(selected_goal) if selected_goal else None
        action_id = loop.actions.index(selected_action) if selected_action else None

        if goal_id is not None and action_id is not None:
            for cor_id in range(loop.cell_types_numbers['Cor'][1]):
                for ct in loop.weight_cell_types:
                    current_avg_weight[ct] = 0
                    for msn_id in range(loop.cell_types_numbers[ct][1]):
                        key = (ct, action_id, msn_id, goal_id, cor_id)
                        current_avg_weight[ct] += loop.weights_over_time[key][-1]
                    current_avg_weight[ct] = current_avg_weight[ct] / loop.cell_types_numbers[ct][1]
        
        if 'MSNd' in current_avg_weight and 'MSNi' in current_avg_weight:
            current_avg_weight_diff = current_avg_weight['MSNd'] - current_avg_weight['MSNi']
        
        return current_avg_weight_diff 
        
                    
    def run_simulation(self):
        try:
            while not self.interrupt:
                if self.paused or all(loop.training is False for loop in self.loops):
                    if self.paused or all(all(val == 0 for val in loop.cortical_input_dur_rel) for loop in self.loops): 
                        # Simulation paused
                        time.sleep(0.1)

                        for loop in self.loops:
                            loop.fig.canvas.draw_idle()   
                            loop.fig.canvas.flush_events()
                        if self.cerebellum_required:
                            self.cerebellum.fig.canvas.draw_idle()
                            self.cerebellum.fig.canvas.flush_events()
                        self.update_GUI_goals_actions()
                        continue
                
                start_time = time.time()

                training_mode = any(loop.training for loop in self.loops)
                action_selection_completed = all(loop.reward_over_time[loop.selected_goal][-1] for loop in self.loops)
                
                for idx, loop in enumerate(self.loops):
                    
                    #if loop.selected_goal and loop.expected_reward_over_time[loop.selected_goal][-1] > 0.5:
                    if loop.selected_goal and self.get_current_average_weight_diff(loop) > 0.75:
                        if loop.training: 
                            self.training_goal_idx += 1
                        map = self.calculate_goal_action_mapping(loop.weights_over_time, loop.goals.index(loop.selected_goal))
                        for goal_idx, action_idx in map.items():
                            goals_names_list = [loop.goals_names[i] for i, val in enumerate(loop.goals[goal_idx]) if val == '1']
                            actions_names_list = [loop.actions_names[i] for i, val in enumerate(loop.actions[action_idx]) if val == '1']
                            if len(goals_names_list) == 1:
                                loop.learned_goal_action_map[goals_names_list[0]] = actions_names_list
                                self.update_arrows()
                        if loop.training and self.training_goal_idx >= len(loop.goals) - 1:
                            self.toggle_train_loop(loop_idx=idx)
                            if idx < len(self.loops) - 1:
                                self.toggle_train_loop(loop_idx=idx+1)                         
                            continue
                    if loop.training: 
                        goal = loop.goals[self.training_goal_idx]
                        for id, state in enumerate(goal):
                            loop.buttons[f'cor_dur_slider{id}'].set_val(0 if state == '0' else 1)
                        loop.update_selected_goal()

                    if loop.selected_goal: loop.cortical_input_stimuli(current_time=h.t)
                    
                # Run simulation
                h.continuerun(h.t + self.plot_interval)

                # --- Action selection ---# 
                for loop in self.loops:
                    loop.analyze_thalamus_activation_time(current_time=h.t)
                    loop.select_action()

                # Cerebellum update input
                if self.cerebellum_required and not training_mode:
                    desired_grasp_type = self.loops[1].selected_goal if self.loops[1].selected_goal else '0' * len(self.cerebellum.grasp_types)
                    desired_joints = self.loops[1].selected_action[0] if self.loops[1].selected_action else '0' * len(self.cerebellum.joints) 
                    desired_actuators = self.loops[0].selected_action[0] if self.loops[0].selected_action else '0' * len(self.cerebellum.actuators)
                    norm_touch_delta_grasp = [min(1.0, delta_grasp / max_delta_grasp) for delta_grasp, max_delta_grasp in zip(self.touch_sensor_delta_grasp, self.touch_sensor_expected_max_delta_grasp)]
                    norm_touch_delta_hold = [min(1.0, delta_hold / max_delta_hold) for delta_hold, max_delta_hold in zip(self.touch_sensor_delta_hold, self.touch_sensor_expected_max_delta_hold)]
                    self.cerebellum.update_input_stimuli(desired_grasp_type, desired_joints, desired_actuators, norm_touch_delta_grasp, norm_touch_delta_hold)

                self.update_GUI_goals_actions()

                if self.hw_connected:
                    grasping = True if not training_mode and action_selection_completed else False
                    self.perform_action(grasping=grasping)
                elif self.cerebellum_required:
                    self.touch_sensor_delta_grasp = self.cerebellum.touch_sensor_delta_grasp
                    self.touch_sensor_delta_hold = self.cerebellum.touch_sensor_delta_hold
                
                # --- Reward update ---# 
                for loop in self.loops:
                    loop.determine_reward(current_time=h.t, hw_connected=self.hw_connected, performed_action=self.performed_action)

                # Cerebellum trigger teaching signal (IO)
                if self.cerebellum_required and not training_mode:
                    self.cerebellum.update_teaching_stimuli(current_time=h.t, desired_joints=desired_joints, action_selection_completed=action_selection_completed, touch_sensor_delta_grasp=self.touch_sensor_delta_grasp, touch_sensor_delta_hold=self.touch_sensor_delta_hold) 

                self.update_GUI_performed_action_reward()

                # --- Weight and plot update ---# 
                for loop in self.loops:
                    loop.update_weights(current_time=h.t)
                    loop.update_plots(current_time=h.t)

                # Cerebellum update weights and plot
                if self.cerebellum_required and not training_mode:
                    self.cerebellum.update_weights(current_time=h.t)
                    self.cerebellum.update_plots(current_time=h.t)
                    self.cerebellum.calculate_correction(current_time=h.t)
                    self.update_joint_duration_mapping(time_correction=self.cerebellum.DCN_diff_rates)

                duration = time.time() - start_time
                self.durations[self.iteration] = round(duration, 2)
                print(f"Iteration {self.iteration} took {self.durations[self.iteration]} s")

                self.fig.canvas.draw_idle()
                self.fig.canvas.flush_events()
                for loop in self.loops:
                    loop.fig.canvas.draw_idle()   
                    loop.fig.canvas.flush_events()
                if self.cerebellum_required:
                    self.cerebellum.fig.canvas.draw_idle()
                    self.cerebellum.fig.canvas.flush_events()
            
        except KeyboardInterrupt:
            self.interrupt = True
            print("\nCtrl-C pressed.")
            self.close()
            
            
    def save(self, event=None):
        self.buttons['save'].label.set_text('Saving...')
        plt.draw()
        plt.pause(0.001)

        self.stop_time = time.time()
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path = f"Data\{timestamp}"    

        self.save_data(path)
        self.fig.savefig(f"{path}_MotorLearning.png", dpi=300, bbox_inches='tight') # GUI Screenshot
        for loop in self.loops:
            loop.save_data(path) # Excel file
            loop.fig.savefig(f"{path}_{loop.name}.png", dpi=300, bbox_inches='tight') # GUI Screenshot
        if self.cerebellum_required:
            self.cerebellum.save_data(path)
            self.cerebellum.fig.savefig(f"{path}_Cerebellum.png", dpi=300, bbox_inches='tight') # GUI Screenshot

        self.buttons['save'].label.set_text('Saved')
        self.buttons['save'].color = 'green'
        self.buttons['save'].hovercolor = 'lightgreen' 

    def close(self, event=None, restart=False):
        self.buttons['close'].label.set_text('Closing...')
        plt.draw()
        plt.pause(0.001)

        plt.close('all') 
        if self.ser_sensor and self.ser_sensor.is_open:
            self.ser_sensor.close()
        if self.ser_exo and self.ser_exo.is_open:
            self.ser_exo.close()
        if not restart: 
            sys.exit(0) 
    

#--- BasalGangliaLoop ------------------------------------------------------------------------------------------------------------------------------------------------#

class BasalGangliaLoop:

    def __init__(self, name, input, output, binary_input=False, single_goal=False, goal_action_table=None, actions_to_plot=None, plot_interval=None, child_loop=None):
        self.name = name
        self.goals_names = input
        self.goals = [''.join(bits) for bits in product('10', repeat=len(input)) # binary combination of all inputs
            if (bits.count('1') <= 1 if single_goal else True)]
        self.selected_goal = ''.join('0' for _ in self.goals_names)
        self.new_selected_goal = None
        self.actions_names = output
        self.actions = [''.join(bits) for bits in product('10', repeat=len(output)) if any(bit == '1' for bit in bits)] # binary combination of all outputs
        self.selected_action = None
        self.binary_input = binary_input
        self.single_goal = single_goal
        self.goal_action_table = goal_action_table
        if actions_to_plot is not None:
            self.actions_to_plot = len(self.actions) if len(self.actions) <= actions_to_plot else actions_to_plot
        self.plot_interval = plot_interval
        self.child_loop = child_loop
        self.training = False
        self.training_start_iteration = -1
        self.training_stop_iteration = -1
        self.learned_goal_action_map = {goal_name: [self.actions_names[min(goal_id, len(self.actions_names)-1)]] for goal_id, goal_name in enumerate(self.goals_names)}

        self.cell_types_numbers = {'Cor':  [len(self.goals),   1], 
                                   'SNc':  [len(self.actions), 1], 
                                   'MSNd': [len(self.actions), 5], 
                                   'MSNi': [len(self.actions), 5], 
                                   'GPe':  [len(self.actions), 5], 
                                   'GPi':  [len(self.actions), 5], 
                                   'Thal': [len(self.actions), 5]}
        self.cell_types = list(self.cell_types_numbers.keys())

        self.stim_intervals = {
            'SNc'       : 1000 / 5,   # Hz
            'MSNd'      : 1000 / 7.4, # Hz (tonic baseline)
            'MSNi'      : 1000 / 3.5, # Hz (tonic baseline)
            'GPe'       : 1000 / 48,  # Hz
            'GPi'       : 1000 / 69,  # Hz
            'Thal'      : 1000 / 14,  # Hz
            'SNc_burst' : 1000 / 50,  # Hz
            'Cor'       : 1000 / 40   # Hz (cortical input stimulation)
        }
 
        self.stim_weights = {
            'SNc'       : 0.5,
            'MSNd'      : 0.5,
            'MSNi'      : 0.5,
            'GPe'       : 0.5,
            'GPi'       : 0.5,
            'Thal'      : 0.5,
            'SNc_burst' : 0.5,
            'Cor'       : 0.4
        }

        self.stims = {}
        self.syns = {}
        self.ncs = {}

        # Define connection specifications
        self.connection_specs_cor = [# pre_group, post_group, label, e_rev, weight, tau, delay
            ('Cor', 'MSNd', 'Cor_to_MSNd',   0, 0.25,   5, 1),   # excitatory
            ('Cor', 'MSNi', 'Cor_to_MSNi',   0, 0.25,   5, 1)    # excitatory
        ]

        # Define connection specifications
        self.connection_specs = [# pre_group, post_group, label, e_rev, weight, tau, delay
            ('SNc', 'MSNd', 'SNc_to_MSNd',   0, 0,     10, 1),   # excitatory
            ('SNc', 'MSNi', 'SNc_to_MSNi', -85, 0,     10, 1),   # inhibitory
            ('MSNd', 'GPi', 'MSNd_to_GPi', -85, 0.6,   10, 1),   # inhibitory
            ('MSNi', 'GPe', 'MSNi_to_GPe', -85, 1.0,   10, 1),   # inhibitory
            ('GPe',  'GPi',  'GPe_to_GPi', -85, 0.3,   10, 1),   # inhibitory
            ('GPi', 'Thal', 'GPi_to_Thal', -85, 1.0,   10, 1)    # inhibitory
        ]
        
        self._is_updating_programmatically = False
        self.n_spikes_SNc_burst = 5
        self.learning_rate = 0.05 
        self.reward_times = []
        self.activation_times = []
        self.expected_reward_over_time = {}
        self.activation_over_time = {}
        self.cortical_input_dur_rel = [0] * len(self.goals_names)
        self.expected_reward_value = 0.5
        self.reward_over_time = {}
        self.dopamine_over_time = {}
        self.weight_cell_types = ['MSNd', 'MSNi']
        self.pretrained_weights_filename = None
        self.trained_weights_filename = None
        self.weight_times = []
        self.weights_over_time = {(ct, action_id, msn_id, goal_id, cor_id): [] 
                            for ct in self.weight_cell_types
                            for action_id, action in enumerate(self.actions)
                            for msn_id in range(self.cell_types_numbers[ct][1])
                            for goal_id, goal in enumerate(self.goals)
                            for cor_id in range(self.cell_types_numbers['Cor'][1])
                            }
        self.cor_nc_index_map = {} 
        self.apc_refs = []

        self.buttons = {}
        self.rates = {}
        self.rates_rel = {}

        self._init_cells()
        self._init_spike_detectors()
        self._init_stimuli()
        self._init_connections(self.connection_specs)
        self._init_connections(self.connection_specs_cor, cor=True)
        self._init_connection_stimuli()
        self._init_reward()
        self._init_activation()
        self._init_recording()
        self._init_plotting()

    def _init_cells(self):
        # Create neuron populations
        self.cells = {
            cell_type: [
                [create_cell(f'{cell_type}_{population}_{index}') for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ] 
            for cell_type in self.cell_types
        }
    
    def _init_spike_detectors(self):
        # Spike detectors and vectors
        self.spike_times = {
            cell_type: [
                [h.Vector() for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ] 
            for cell_type in self.cell_types
        }

        for cell_type in self.cell_types:
            for population in range(self.cell_types_numbers[cell_type][0]):
                for index, cell in enumerate(self.cells[cell_type][population]):
                    apc = h.APCount(cell(0.5))
                    apc.record(self.spike_times[cell_type][population][index])
                    self.apc_refs.append(apc)

    def _init_stimuli(self):
        for cell_type in self.cell_types:
            self.stims[cell_type], self.syns[cell_type], self.ncs[cell_type] = [], [], []
            for population in range(self.cell_types_numbers[cell_type][0]):
                for i, cell in enumerate(self.cells[cell_type][population]):
                    if 'SNc' in str(cell):
                        offset = self.stim_intervals['SNc']/2
                    elif 'Cor' in str(cell):
                        offset = 0
                    else:
                        offset = i*self.stim_intervals[cell_type]/self.cell_types_numbers[cell_type][1]
                    stim, syn, nc = create_stim(cell, start=offset, interval=self.stim_intervals[cell_type], weight=self.stim_weights[cell_type])
                    self.stims[cell_type].append(stim)
                    self.syns[cell_type].append(syn)
                    self.ncs[cell_type].append(nc)

    def _init_connections(self, connection_specs, cor=False):
        if cor: 
            index = 0

        # Create connections based on the specification
        for pre_group, post_group, label, e_rev, weight, tau, delay in connection_specs:
            self.ncs.update({label: []}) # Additional connections dict to store NetCons
            self.syns.update({label: []}) # Additional connections dict to store ExpSyns

            if cor:
                # Only do index mapping for cortical to striatal connections
                self.cor_nc_index_map[label] = {}
                index = 0
                for population_pre_cell in range(self.cell_types_numbers[pre_group][0]):
                    for pre_cell_id, pre_cell in enumerate(self.cells[pre_group][population_pre_cell]):
                        for population_post_cell in range(self.cell_types_numbers[post_group][0]):
                            for post_cell_id, post_cell in enumerate(self.cells[post_group][population_post_cell]):
                                syn = h.ExpSyn(post_cell(0.5))
                                syn.e = e_rev
                                syn.tau = tau
                                nc = h.NetCon(pre_cell(0.5)._ref_v, syn, sec=pre_cell)
                                nc.weight[0] = weight
                                nc.delay = delay
                                self.syns[label].append(syn)
                                self.ncs[label].append(nc)

                                # Save index mapping if cortical to striatal
                                if label.startswith('Cor_to'):
                                    key = (post_group, population_post_cell, post_cell_id, population_pre_cell, pre_cell_id)
                                    self.cor_nc_index_map[label][key] = index
                                index += 1
            else:
                for action_id, action in enumerate(self.actions):                    
                    for pre_cell in self.cells[pre_group][action_id]:
                        for post_cell in self.cells[post_group][action_id]:
                            syn = h.ExpSyn(post_cell(0.5))
                            syn.e = e_rev
                            syn.tau = tau
                            nc = h.NetCon(pre_cell(0.5)._ref_v, syn, sec=pre_cell)
                            nc.weight[0] = weight
                            nc.delay = delay
                            self.syns[label].append(syn)
                            self.ncs[label].append(nc)
                    
    def _init_connection_stimuli(self):
        # Additional stimuli for dopamine bursts
        self.stims.update({'SNc_burst': []})
        self.syns.update({'SNc_burst': []})
        self.ncs.update({'SNc_burst': []})

        for a, _ in enumerate(self.actions):
            for cell in self.cells['SNc'][a]:
                stim, syn, nc = create_stim(cell, start=0, interval=self.stim_intervals['SNc_burst'], weight=self.stim_weights['SNc_burst'])
                self.stims['SNc_burst'].append(stim)
                self.syns['SNc_burst'].append(syn)
                self.ncs['SNc_burst'].append(nc)
        for nc in self.ncs['SNc_burst']:
            nc.active(False)

        # Initialize weights over time
        self.weight_times.append(0)
        for ct in self.weight_cell_types:
            for action_id, action in enumerate(self.actions):
                for msn_id in range(self.cell_types_numbers[ct][1]):
                    for goal_id, goal in enumerate(self.goals):
                        for cor_id in range(self.cell_types_numbers['Cor'][1]):
                            for _, post_group, _, _, weight, _, _ in self.connection_specs_cor:
                                if ct == post_group:
                                    self.weights_over_time[(ct, action_id, msn_id, goal_id, cor_id)].append(weight)
        for nc in self.ncs[f'Cor']:
            nc.active(False)

    def _init_reward(self):
        # Reward initialization
        self.reward_times.append(0)
        for goal in self.goals:
            self.expected_reward_over_time[goal] = [self.expected_reward_value]
            self.reward_over_time[goal] = [0]
            self.dopamine_over_time[goal] = [0]
        
    def _init_activation(self):
        # Activation initialization
        self.activation_times.append(0)
        for action in self.actions:
            self.activation_over_time[action] = [0]

    def _init_recording(self):
        # Recording
        self.recordings = {ct: [[h.Vector().record(cell(0.5)._ref_v) for cell in self.cells[ct][population]] for population in range(self.cell_types_numbers[ct][0])] for ct in self.cell_types}
        self.t_vec = h.Vector().record(h._ref_t)

    def _init_plotting(self):
        plt.ion()
        self.fig = plt.figure(figsize=(13, 8))
        self.fig.canvas.manager.set_window_title(self.name)
        self.rows = 3
        self.gs = gridspec.GridSpec(2, 1, figure=self.fig, height_ratios=[1, 2 * self.rows])
        self.gs_control = self.gs[0].subgridspec(1, 2 + len(self.goals_names), width_ratios=[1] + [1] * len(self.goals_names) + [2])
        self.gs_plot = self.gs[1].subgridspec(self.rows, self.actions_to_plot)
        self.axs_control = [self.fig.add_subplot(self.gs_control[i]) for i in range(2 + len(self.goals_names))]
        self.axs_plot = []
        for i in range(self.rows):
            row = []
            for j in range(self.actions_to_plot):
                if j == 0:
                    ax = self.fig.add_subplot(self.gs_plot[i, j])
                else:
                    ax = self.fig.add_subplot(self.gs_plot[i, j], sharey=row[0])
                row.append(ax)
            self.axs_plot.append(row)
        for i in range(self.rows):
            for j in range(1, self.actions_to_plot):  # skip first in row
                plt.setp(self.axs_plot[i][j].get_yticklabels(), visible=False)

        [self.row_potential, self.row_spike, self.row_weights] = list(range(self.rows))
        
        # Deactivate axis for control axes
        for ax in self.axs_control:
            ax.set_axis_off() # deactivate axis
        self.axs_control[-1].set_axis_on() # activate axis for reward plot
        
        self._init_membrane_potential_plot()
        self._init_spike_plot()
        self._init_weight_plot()
        self._init_reward_plot()
        self._init_control_panel()

        plt.show()
        plt.tight_layout()

    def _init_membrane_potential_plot(self):
        # Membrane potential plot
        self.axs_plot[self.row_potential][0].set_ylabel('Membrane potential (mV)')
        self.mem_lines = {ct: [] for ct in self.cell_types}

        for i, ch in enumerate(range(self.actions_to_plot)):
            for j, ct in enumerate(self.cell_types):
                avg_line, = self.axs_plot[self.row_potential][i].plot([], [], f'C{j}', label=ct)
                self.mem_lines[ct].append(avg_line)

            self.axs_plot[self.row_potential][i].set_title(f'{self.actions[ch]}')
            self.axs_plot[self.row_potential][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_potential][i].set_ylim(-85, 65)
            self.axs_plot[self.row_potential][i].xaxis.set_major_formatter(ms_to_s)
        self.axs_plot[self.row_potential][-1].legend(loc='upper right')

    def _init_spike_plot(self):
        # Spike raster plot and rate lines
        self.raster_lines = {ct: [[] for _ in range(self.actions_to_plot)] for ct in self.cell_types}
        self.rate_lines = {ct: [] for ct in self.cell_types}
        self.axs_plot[self.row_spike][0].set_ylabel('Spike raster')

        for i, ch in enumerate(range(self.actions_to_plot)):
            for j, ct in enumerate(self.cell_types):
                self.raster_lines[ct][i] = []
                for k in range(self.cell_types_numbers[ct][1]):
                    line, = self.axs_plot[self.row_spike][i].plot([], [], f'C{j}.', markersize=3)
                    self.raster_lines[ct][i].append(line)

                rate_line, = self.axs_plot[self.row_spike][i].step([], [], f'C{j}')
                self.rate_lines[ct].append(rate_line)
            self.axs_plot[self.row_spike][i].plot([], [], color='black', linestyle='dotted', label=f'Spikes')
            self.axs_plot[self.row_spike][i].plot([], [], color='black', label=f'Relative\nrate')

            self.total_cells = sum(self.cell_types_numbers[ct][1] for ct in self.cell_types)
            
            y_max = self.total_cells + 1.5
            self.axs_plot[self.row_spike][i].set_ylim(0.5, y_max)
            yticks = []
            cumulative = 0
            for ct in self.cell_types:
                mid = y_max - (cumulative + (self.cell_types_numbers[ct][1]+1) / 2)
                yticks.append(mid)
                cumulative += self.cell_types_numbers[ct][1] 
            self.axs_plot[self.row_spike][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_spike][i].set_yticks(yticks)
            self.axs_plot[self.row_spike][i].set_yticklabels(self.cell_types)
            self.axs_plot[self.row_spike][i].xaxis.set_major_formatter(ms_to_s)
        self.axs_plot[self.row_spike][-1].legend(loc='upper right')

    def _init_weight_plot(self):
        # Weight plot
        self.axs_plot[self.row_weights][0].set_ylabel('Cortical input weight')
        self.weight_lines = {ct: [[] for _ in range(self.actions_to_plot)] for ct in self.weight_cell_types}
        self.activation_lines = []

        for i, ch in enumerate(range(self.actions_to_plot)):
            for j, ct in enumerate(self.cell_types):
                if ct in self.weight_cell_types:
                    self.weight_lines[ct][i] = []
                    for k in range(self.cell_types_numbers[ct][1]):
                        label = ct if k == 0 else None  # Only label the first line
                        line, = self.axs_plot[self.row_weights][i].step([], [], f'C{j}', label=label, where='post')
                        self.weight_lines[ct][i].append(line)

            activation_line, = self.axs_plot[self.row_weights][i].step([], [], 'C7', linestyle='dashed', label='Activation', where='post')
            self.activation_lines.append(activation_line)

            self.axs_plot[self.row_weights][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_weights][i].set_ylim(-0.1, 1.1)
            self.axs_plot[self.row_weights][i].xaxis.set_major_formatter(ms_to_s)
            self.axs_plot[self.row_weights][i].set_xlabel('Simulation\ntime (s)')
        self.axs_plot[self.row_weights][-1].legend(loc='upper right')

    def _init_reward_plot(self):
        # Reward plot
        self.expected_reward_lines = []
        self.reward_lines = []
        self.dopamine_lines = []

        self.expected_reward_lines, = self.axs_control[-1].plot([], [], 'C9', label='Expected\nreward')
        self.reward_lines, = self.axs_control[-1].step([], [], 'C8', label='Reward', where='post')
        self.dopamine_lines, = self.axs_control[-1].plot([], [], 'C6', label='Dopamine')

        self.axs_control[-1].xaxis.set_major_formatter(ms_to_s)
        self.axs_control[-1].set_xlabel('Simulation time (s)')
        self.axs_control[-1].set_xlim(0, self.plot_interval)
        self.axs_control[-1].set_ylim(-1.1, 1.1)
        self.axs_control[-1].legend(loc='upper left')

    def _init_control_panel(self):
        #--- Upper control panel ---#        
        for i, goal_name in enumerate(self.goals_names):
            ax_cor_dur = self.axs_control[i].inset_axes([0,0,0.9,0.45]) #[x0, y0, width, height]
            label_text = '\n'.join(goal_name.split())
            ax_cor_dur.set_title(label_text)
            self.buttons[f'cor_dur_slider{i}'] = Slider(ax_cor_dur, '', 0, 1, valinit=self.cortical_input_dur_rel[i], valstep=1 if self.binary_input else 0.2)
            self.buttons[f'cor_dur_slider{i}'].on_changed(lambda val, i=i: self.update_cor_dur(val=val, goal_idx=i))

    def update_stimulus_activation(self, cell, stimulus, index, active=True):
        i = 0
        for population in range(self.cell_types_numbers[cell][0]):
            for _ in self.cells[cell][population]:
                if population == index:
                    self.ncs[stimulus][i].active(active)
                i += 1

    def SNc_dip(self, event=None, action=None):
        self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=False) # stop SNc tonic stimulus
        h.cvode.event(h.t + self.stim_intervals['SNc'], lambda action=action: self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=True))  # start SNc tonic stimulus

    def SNc_burst(self, event=None, action=None, n_spikes=None):
        self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=False) # stop SNc tonic stimulus
        self.update_stimulus_activation(cell='SNc', stimulus='SNc_burst', index=self.actions.index(action), active=True) # start SNc burst stimulus
        if n_spikes == None:
            n_spikes = self.n_spikes_SNc_burst
        delay = self.stim_intervals['SNc_burst'] * n_spikes
        h.cvode.event(h.t + delay, lambda action=action: self.update_stimulus_activation(cell='SNc', stimulus='SNc_burst', index=self.actions.index(action), active=False))  # stop SNc burst stimulus
        h.cvode.event(h.t + delay, lambda action=action: self.update_stimulus_activation(cell='SNc', stimulus='SNc', index=self.actions.index(action), active=True))  # start SNc tonic stimulus

    def analyze_firing_rate(self, cell, window=None, average=True):
        """Returns a list of firing rates (Hz) for each action's cell population."""
        current_time = h.t
        rates_avg = []
        rates = []
        if window == None:
            window = self.plot_interval
        for a, _ in enumerate(self.actions):
            spikes_avg = 0
            spikes = []
            for i in range(self.cell_types_numbers[cell][1]):
                spike_vec = self.spike_times[cell][a][i]
                # Count spikes in the last `window` ms
                recent_spikes = [t for t in spike_vec if current_time - window <= t <= current_time]
                if average:
                    spikes_avg += len(recent_spikes)
                else:
                    spikes.append(len(recent_spikes))
            if average:
                rate_avg = spikes_avg / (self.cell_types_numbers[cell][1] * (window / 1000.0))  # spikes/sec per neuron
                rates_avg.append(rate_avg)
            else:
                rate = [spike / (window / 1000.0) for spike in spikes]
                rates.append(rate)

        max_rate = 1000.0 / self.stim_intervals[cell]
        
        if average:
            rates_rel_avg = [rate_avg / max_rate for rate_avg in rates_avg]
            return rates_avg, rates_rel_avg
        else:
            rates_rel = [[r / max_rate for r in rate] for rate in rates]
            return rates, rates_rel

    def update_cor_dur(self, val, goal_idx):
        if self._is_updating_programmatically:
            return
        self.cortical_input_dur_rel[goal_idx] = val
        if self.single_goal:
            if sum(value != 0 for value in [self.buttons[f'cor_dur_slider{i}'].val for i,_ in enumerate(self.goals_names)]) > 1:
                self.reset_other_durations(goal_idx)

        self.update_selected_goal()

    def reset_other_durations(self, goal_idx):
        self._is_updating_programmatically = True
        for i, _ in enumerate(self.goals_names):
            if i is not goal_idx:
                self.cortical_input_dur_rel[i] = 0
                self.buttons[f'cor_dur_slider{i}'].set_val(0)
        self._is_updating_programmatically = False

    def update_selected_goal(self):
        self.selected_goal = ''.join('1' if val != 0 else '0' for val in self.cortical_input_dur_rel)

    def cortical_input_stimuli(self, current_time, goal=None):
        selected_goal = goal if goal else self.selected_goal
        selected_goal_index = self.goals.index(selected_goal)
        
        for idx, _ in enumerate(self.goals):
            h.cvode.event(current_time + 1, lambda: self.update_stimulus_activation(cell='Cor', stimulus=f'Cor', index=idx, active=False)) # stop all cortical input stimuli
        
        if set(selected_goal) != {'0'}: # Only stimulate cortex if a goal is set
            h.cvode.event(current_time + 1, lambda: self.update_stimulus_activation(cell='Cor', stimulus=f'Cor', index=selected_goal_index, active=True)) # start particular cortical input stimulus
            h.cvode.event(current_time + self.plot_interval, lambda: self.update_stimulus_activation(cell='Cor', stimulus=f'Cor', index=selected_goal_index, active=False)) # stop particular cortical input stimulus

    def analyze_thalamus_activation_time(self, current_time):
        self.activation_times.append(int(current_time))

        if set(self.selected_goal) == {'0'}:
            # Store 0 for all actions
            for action in self.actions:
                self.activation_over_time[action].append(0)
        else:
            rates = {}
            window_start = np.array(self.t_vec.to_python())[-1] - self.plot_interval
            window_end = np.array(self.t_vec.to_python())[-1]
            window_duration_sec = (window_end - window_start) / 1000.0  # Convert ms to seconds

            for action_id, action in enumerate(self.actions):
                # Collect all thalamic spike times for this action
                all_spikes = []
                for k in range(self.cell_types_numbers['Thal'][1]):
                    spikes = np.array(self.spike_times['Thal'][action_id][k].to_python())
                    all_spikes.extend(spikes)

                # Filter spikes within the last window
                spikes_in_window = [spk for spk in all_spikes if window_start < spk <= window_end]

                # Compute firing rate: total spikes / (number of cells * window duration)
                num_cells = self.cell_types_numbers['Thal'][1]
                rate = len(spikes_in_window) / (num_cells * window_duration_sec) if window_duration_sec > 0 else 0
                rates[action] = rate

            # Find the action with the highest thalamus firing rate
            if rates:
                max_value = max(rates.values())
                candidates = [action for action, value in rates.items() if value == max_value]
                best_action = random.choice(candidates)
            else:
                best_action = None

            # Store 1 for the best action, 0 for the rest
            for action in self.actions:
                self.activation_over_time[action].append(1 if action == best_action else 0)
  
    def select_action(self):
        active_actions = [(i, activations[-1]) for i, activations in self.activation_over_time.items() if activations[-1] > 0]
        if active_actions and self.selected_goal and set(self.selected_goal) != {'0'}:
            # Pick the action with the maximum last activation value
            self.selected_action = max(active_actions, key=lambda x: x[1])
        else:
            # No active actions found
            self.selected_action = []

        if not self.training and self.child_loop and self.selected_action and set(self.selected_goal) != {'0'}:
            # Set input of child loop based on output of current loop
            for id, state in enumerate(list(map(int, self.selected_action[0]))):
                self.child_loop.buttons[f'cor_dur_slider{id}'].set_val(0 if state == 0 else self.selected_action[1])
            self.child_loop.update_selected_goal()
    
    def determine_reward(self, current_time, hw_connected=None, performed_action=None):
        self.reward_times.append(int(current_time))

        goal_state = tuple((goal_name, dur != 0) for goal_name, dur in zip(self.goals_names, self.cortical_input_dur_rel))
        target_action_dict = self.goal_action_table.get(goal_state, {})
                  
        for goal in self.goals:
            if hw_connected and not self.child_loop and performed_action:
                action_correct = performed_action == self.selected_goal
            elif self.selected_action:
                action_dict = dict(zip(self.actions_names, [c == '1' for c in self.selected_action[0]]))
                action_correct = action_dict == target_action_dict
            else:
                action_correct = False
            
            if goal == self.selected_goal and action_correct:
                self.reward_over_time[goal].append(1)
            else:
                self.reward_over_time[goal].append(0)
                
            current_expected_reward = self.expected_reward_over_time[goal][-1]
            self.dopamine_over_time[goal].append(round(self.reward_over_time[goal][-1] - current_expected_reward, 4))
            
            for action in self.actions:
                if self.reward_over_time[goal][-1] - current_expected_reward > 0:
                    self.SNc_burst(event=None, action=action)
                elif self.reward_over_time[goal][-1] - current_expected_reward < 0:
                    self.SNc_dip(event=None, action=action)
                    
            if goal == self.selected_goal:
                #For selected goal update expected reward based on actual reward
                self.expected_reward_over_time[goal].append(round(current_expected_reward + self.learning_rate * (self.reward_over_time[goal][-1] - current_expected_reward), 4))
            else:
                self.expected_reward_over_time[goal].append(self.expected_reward_over_time[goal][-1])

    def update_weights(self, current_time):
        # Analyze firing rates
        self.rates['MSNd'], self.rates_rel['MSNd'] = self.analyze_firing_rate('MSNd', window=self.plot_interval, average=False)
        self.rates['MSNi'], self.rates_rel['MSNi'] = self.analyze_firing_rate('MSNi', window=self.plot_interval, average=False)

        self.weight_times.append(int(current_time))
        selected_goal = self.selected_goal if self.selected_goal else None
        selected_action = self.selected_action[0] if self.selected_action else None
        goal_id = self.goals.index(selected_goal) if selected_goal else None
        action_id = self.actions.index(selected_action) if selected_action else None

        # Only update weights for the selected goal and action
        if goal_id is not None and action_id is not None:
            for cor_id in range(self.cell_types_numbers['Cor'][1]):
                for ct in self.weight_cell_types:
                    for msn_id in range(self.cell_types_numbers[ct][1]):
                        delta_w = 0
                        if ct == 'MSNd':
                            delta_w = self.learning_rate * (self.rates_rel[ct][action_id][msn_id] - 1) * self.dopamine_over_time[selected_goal][-1]
                            delta_w = delta_w
                        elif ct == 'MSNi':
                            factor = 2
                            delta_w = self.learning_rate * factor * self.dopamine_over_time[selected_goal][-1]
                            delta_w = -delta_w
                        key = (ct, action_id, msn_id, goal_id, cor_id)
                        new_weight = min(1, max(0, self.weights_over_time[key][-1] + delta_w))
                        self.weights_over_time[key].append(round(new_weight, 4))
                        idx = self.cor_nc_index_map[f'Cor_to_{ct}'][key]
                        self.ncs[f'Cor_to_{ct}'][idx].weight[0] = new_weight
        # For all other keys, just append the previous value to keep list lengths consistent
        for key in self.weights_over_time:
            if len(self.weights_over_time[key]) < len(self.weight_times):
                self.weights_over_time[key].append(self.weights_over_time[key][-1])

    def save_trained_weights(self):
        trained_weights = {
            key: weights[-1] 
            for key, weights in self.weights_over_time.items() 
            if key[0] in self.weight_cell_types
        }

        state = {
            'trained_weights': trained_weights,
            'learned_goal_action_map': self.learned_goal_action_map
        }
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.trained_weights_filename  = f"Data\{timestamp}_Trained_Weights_{self.name}.pkl"  
        with open(self.trained_weights_filename, 'wb') as f:
            pickle.dump(state, f)
        print(f"Trained weights saved to {self.trained_weights_filename}")

    def load_trained_weights(self):
        pattern = f"PretrainedWeights/*_Trained_Weights_{self.name}.pkl"
        matches = glob.glob(pattern)  
        
        if matches:
            filename = matches[0]  # first match
        else:
            raise FileNotFoundError(f"No pretrained weights found for {self.name}")

        with open(filename, 'rb') as f:
            state = pickle.load(f)

        # Update weights_over_time
        trained_weights = state['trained_weights']
        for key, weight in trained_weights.items():
            self.weights_over_time[key] = [weight]  # start with loaded weight

            # Update actual NetCon objects
            ct = key[0]
            idx = self.cor_nc_index_map[f'Cor_to_{ct}'][key]
            self.ncs[f'Cor_to_{ct}'][idx].weight[0] = weight
        
        # Load learned goal-action map
        self.learned_goal_action_map = state['learned_goal_action_map']
        self.pretrained_weights_filename = filename
        print(f"Cortical weights loaded from {filename}")

    def update_plots(self, current_time, goal=None, action=None):
        if goal:
            selected_goal_index = self.goals.index(goal)
        else:
            selected_goal_index = self.goals.index(self.selected_goal)
        
        if action:
            sorted_by_recent = [action]+[a for a in self.actions if a is not action]
            actions_to_plot_now = sorted_by_recent[:self.actions_to_plot]
        else:
            # Determine actions to plot (dynamic)
            last_active_indices = { 
                action: (len(act) - 1 - next(i for i, v in enumerate(reversed(act)) if v) if any(act) else -1)
                for action, act in self.activation_over_time.items()}
            sorted_by_recent = sorted(self.actions, key=lambda a: last_active_indices[a], reverse=True)
            actions_to_plot_now = sorted_by_recent[:self.actions_to_plot]

        # Update plots
        for plot_id, action in enumerate(actions_to_plot_now):
            action_id = self.actions.index(action)
            # Membrane potential plot
            for ct in self.cell_types:
                if ct == 'Cor':
                    voltages = np.array([(self.recordings[ct][selected_goal_index][j]) for j in range(self.cell_types_numbers[ct][1])])
                else:
                    voltages = np.array([(self.recordings[ct][action_id][j]) for j in range(self.cell_types_numbers[ct][1])])
                avg_voltage = np.mean(voltages, axis=0)
                self.mem_lines[ct][plot_id].set_data(np.array(self.t_vec), avg_voltage)
                new_xlim = (max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time)))
                if self.axs_plot[self.row_potential][plot_id].get_xlim() != new_xlim:
                    self.axs_plot[self.row_potential][plot_id].set_xlim(*new_xlim)
            self.axs_plot[self.row_potential][plot_id].set_title(action)

            # Spike raster plot
            y_base = self.total_cells
            for ct in self.cell_types:
                all_spikes = []
                if ct == 'Cor':
                    spike_block = self.spike_times[ct][selected_goal_index]
                else:
                    spike_block = self.spike_times[ct][action_id]
                    
                for k in range(self.cell_types_numbers[ct][1]):
                    y_val = y_base - k
                    spikes = np.array(spike_block[k].to_python())
                    y_vals = np.ones_like(spikes) * y_val
                    self.raster_lines[ct][plot_id][k].set_data(spikes, y_vals)
                    all_spikes.extend(spikes)
                # Rate lines
                if all_spikes:
                    bins = np.arange(0, np.array(self.t_vec)[-1] + self.plot_interval, self.plot_interval)
                    hist, edges = np.histogram(all_spikes, bins=bins)
                    if np.any(hist):  # Only proceed if there's at least one spike
                        rate = hist / (self.cell_types_numbers[ct][1] * self.plot_interval / 1000.0)
                        bin_ends = edges[1:]
                        if ct == 'SNc':
                            spike_rate_max = 1000.0 / self.stim_intervals['SNc_burst'] # Hz
                        elif ct == 'MSNd' or ct == 'MSNi':
                            spike_rate_max = 1000.0 / self.stim_intervals['Cor'] # Hz
                        else:
                            spike_rate_max = 1000.0 / self.stim_intervals[ct] # Hz
                        rate_scaled = (rate) / (spike_rate_max + 1e-9)
                        rate_scaled = rate_scaled * (self.cell_types_numbers[ct][1] - 1) + y_base - self.cell_types_numbers[ct][1] + 1
                        self.rate_lines[ct][plot_id].set_data(bin_ends, rate_scaled)
                    else:
                        self.rate_lines[ct][plot_id].set_data([], [])          
                y_base -= self.cell_types_numbers[ct][1]
            new_xlim = max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_spike][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_spike][plot_id].set_xlim(*new_xlim)

            # Weight plot
            for ct in self.weight_cell_types:
                for msn_id in range(self.cell_types_numbers[ct][1]):
                    for cor_id in range(self.cell_types_numbers['Cor'][1]):
                        self.weight_lines[ct][plot_id][msn_id].set_data(self.weight_times, self.weights_over_time[(ct, action_id, msn_id, selected_goal_index, cor_id)])
            self.activation_lines[plot_id].set_data(self.activation_times, self.activation_over_time[action])
            new_xlim = 0, max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_weights][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_weights][plot_id].set_xlim(*new_xlim)

        # Reward plot
        if set(self.selected_goal) != {'0'}:
            self.expected_reward_lines.set_data(self.reward_times, self.expected_reward_over_time[self.selected_goal])
            self.reward_lines.set_data(self.reward_times, self.reward_over_time[self.selected_goal])
            self.dopamine_lines.set_data(self.reward_times, self.dopamine_over_time[self.selected_goal])
        else:
            self.expected_reward_lines.set_data(self.reward_times, [0]*len(self.reward_times))
            self.reward_lines.set_data(self.reward_times, [0]*len(self.reward_times))
            self.dopamine_lines.set_data(self.reward_times, [0]*len(self.reward_times))
        new_xlim = 0, max(self.plot_interval, int(current_time))
        if self.axs_control[-1].get_xlim() != new_xlim:
            self.axs_control[-1].set_xlim(*new_xlim)
    
    def save_data(self, path):
        # Workbook
        wb = Workbook()
        path_extented = f"{path}_{self.name}"

        # Worksheet for General Details
        ws_globals = wb.active
        ws_globals.title = "GlobalVariables"

        row = 1
        row = write_list(ws_globals, "goals_names", self.goals_names, row)
        row = write_list(ws_globals, "actions_names", self.actions_names, row)
        row = write_dict(ws_globals, "goal_action_table", self.goal_action_table, row)
        row = write_dict(ws_globals, "learned_goal_action_map", self.learned_goal_action_map, row)
        row = write_dict(ws_globals, "cell_types_numbers", self.cell_types_numbers, row)
        row = write_dict(ws_globals, "stim_intervals", self.stim_intervals, row)
        row = write_dict(ws_globals, "stim_weights", self.stim_weights, row)
        row = write_tuples(ws_globals, "connection_specs_cor", self.connection_specs_cor, row)
        row = write_tuples(ws_globals, "connection_specs", self.connection_specs, row)
    
        # --- Scalars ---
        scalars = {
            "binary_input": self.binary_input,
            "single_goal": self.single_goal,
            "child_loop": True if self.child_loop else False,
            "n_spikes_SNc_burst": self.n_spikes_SNc_burst,
            "learning_rate": self.learning_rate,
            "expected_reward_value": self.expected_reward_value,
            "training_start_iteration": self.training_start_iteration,
            "training_stop_iteration": self.training_stop_iteration,
            "pretrained_weights_filename": self.pretrained_weights_filename,
            "trained_weights_filename": self.trained_weights_filename
        }
        row = write_dict(ws_globals, "Scalars", scalars, row)

        # Worksheet for Weights
        ws_weights = wb.create_sheet(title="WeightsOverTime")

        # Header
        header = ['time']
        keys = sorted(self.weights_over_time.keys())
        header.extend(f"{ct}_a{action_id}_msn{msn_id}_g{goal_id}_cor{cor_id}" for ct, action_id, msn_id, goal_id, cor_id in keys)
        ws_weights.append(header)

        # Weights
        max_len = len(self.weight_times)
        for t_idx in range(max_len):
            row = [self.weight_times[t_idx]]
            for key in keys:
                val = self.weights_over_time[key][t_idx] if t_idx < len(self.weights_over_time[key]) else None
                row.append(val)
            ws_weights.append(row)

        # Worksheets
        data_list = [
            ("expected_reward_over_time", self.expected_reward_over_time), 
            ("activation_over_time", self.activation_over_time), 
            ("reward_over_time", self.reward_over_time), 
            ("dopamine_over_time", self.dopamine_over_time)
            ]
        ws_list = []
        for name, data in data_list:
            ws = wb.create_sheet(title=f"{name}")
            
            # Header
            header = ['time']
            keys = sorted(data.keys())
            header.extend(f"{key}" for key in keys)
            ws.append(header)

            # Rows
            max_len = len(self.reward_times)
            for t_idx in range(max_len):
                row = [self.reward_times[t_idx]]
                for key in keys:
                    val = data[key][t_idx] if t_idx < len(data[key]) else None
                    row.append(val)
                ws.append(row)

            ws_list.append(ws)

        # Save
        wb.save(f"{path_extented}.xlsx") # Excel
        print(path_extented)

#--- Cerebellum ------------------------------------------------------------------------------------------------------------------------------------------------#

class Cerebellum:

    def __init__(self, grasp_types, joints, actuators, plot_interval, touch_threshold, touch_overload_threshold, drop_threshold):

        self.apc_refs = []
        self.plot_interval = plot_interval

        N_pressure = 5
        self.grasp_types = grasp_types
        self.joints = joints
        self.actuators = actuators
        self.touch_sensor_delta_grasp = [0] * N_pressure
        self.touch_sensor_delta_hold = [0] * N_pressure

        self.touch_threshold = touch_threshold
        self.touch_overload_threshold = touch_overload_threshold
        self.drop_threshold = drop_threshold

        # STDP
        self.STDP_A_pos = 0.005
        self.STDP_A_neg = 0.05
        self.STDP_center = -100.0
        self.STDP_sigma = 30.0

        self.num_pontine = len(self.grasp_types) + len(self.joints) + len(self.actuators) + len(self.touch_sensor_delta_grasp) + len(self.touch_sensor_delta_hold)
        self.num_granule = 3 * self.num_pontine
        self.num_deep_cerebellar = 2 # 1 for positive correction, 1 for negative correction
        self.num_purkinje = 4* self.num_deep_cerebellar
        self.num_inferior_olive = self.num_deep_cerebellar

        self.DCN_diff_rates = [0] * len(self.joints)
    
        self.cell_types_numbers = {'PN':  [1, self.num_pontine], 
                                   'GC':  [1, self.num_granule], 
                                   'IO':  [len(self.actuators), self.num_inferior_olive],
                                   'PC':  [len(self.actuators), self.num_purkinje], 
                                   'DCN': [len(self.actuators), self.num_deep_cerebellar]}
        self.cell_types = list(self.cell_types_numbers.keys())
        self.total_cell_numbers = {cell: val[0] * val[1] for cell, val in self.cell_types_numbers.items()}

        self.cell_types_colors = {'PN':  ['C1'], 
                                  'GC':  ['C2'], 
                                  'IO':  ['C3'],
                                  'PC':  ['C4', 'C5'], 
                                  'DCN': ['C6']}

        self.stim_intervals = {
            'PN'  : 1000 / 40, # 40 Hz
            'GC'  : 0,
            'PC'  : 0, 
            'DCN' : 0,
            'IO'  : self.plot_interval # 1 spike per plot_interval
        }
 
        self.stim_weights = {
            'PN'  : 1.0,
            'GC'  : 0.0,
            'PC'  : 0.0,
            'DCN' : 0.0,
            'IO'  : 1.0
        }

        self.stim_active = {
            'PN'  : True,
            'GC'  : False,
            'PC'  : False,
            'DCN' : False,
            'IO'  : False
        }

        self.stims = {}
        self.syns = {}
        self.ncs = {}

        # Define connection specifications
        self.connection_specs = [# pre_group, post_group, label, e_rev, weight, tau, delay, threshold, sparsity, grouped
            ('PN', 'GC',  'PN_to_GC',    0, 0.2,  2, 10,  10, 1, 0),  # excitatory
            ('GC', 'DCN', 'GC_to_DCN',   0, 0.05, 2,  6,  10, 0, 0),  # excitatory
            ('GC', 'PC',  'GC_to_PC',    0, 0.1,  1,  2,  10, 0, 0),  # excitatory
            ('IO', 'PC',  'IO_to_PC',    0, 0.0,  1,  1,  10, 0, 1),  # excitatory
            ('IO', 'DCN', 'IO_to_DCN',   0, 3.0,  1,  1,  10, 0, 1),  # excitatory
            ('PC', 'DCN', 'PC_to_DCN', -85, 5.0,  5,  1, -10, 0, 1)   # inhibitory
        ]
        
        self.num_of_plots = min(6, len(self.actuators))
        self._is_updating_programmatically = False

        self.buttons = {}
        self.rates = {}
        self.rates_rel = {}

        # Weights
        self.weight_times = [0]
        initial_weight = 0
        
        self.weights_over_time = {(gc_id, pc_id): [] 
                            for gc_id in range(self.total_cell_numbers['GC'])
                            for pc_id in range(self.total_cell_numbers['PC'])
                            }
        for spec in self.connection_specs:
            pre_group, post_group, label, e_rev, weight, tau, delay, threshold, sparsity, grouped = spec
            if label == 'GC_to_PC':
                initial_weight = weight
        self.min_weight = 0.4 * initial_weight if initial_weight else 0.05 # 40%
        self.max_weight = 1.6 * initial_weight if initial_weight else 0.2 # 160%
        self.processed_pairs = {}
        self.gc_spikes_last_interval = []
        self.nc_index_map = {}      # label -> {(pre_id, post_id): idx}
        self.pre_to_post = {}       # label -> {pre_id: [post_id1, post_id2, ...]}
        self.post_to_pre = {}       # label -> {post_id: [pre_id1, pre_id2, ...]}

        # Errors
        self.error_times = [0]        
        self.error_over_time = {joint_idx: [0] for joint_idx in range(len(self.joints))}
        
        # Corrections
        self.correction_times = [0]        
        self.correction_over_time = {joint_idx: [0] for joint_idx in range(len(self.joints))}
        
        self._init_cells()
        self._init_spike_detectors()
        self._init_stimuli()
        self._init_connections()
        self._init_recording()
        self._init_plotting()


    def _init_cells(self):
        # Create neuron populations
        self.cells = {
            cell_type: [
                [create_cell(f'{cell_type}_{population}_{index}') for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ] 
            for cell_type in self.cell_types
        }
    
    def _init_spike_detectors(self):
        # Spike detectors and vectors
        self.spike_times = {
            cell_type: [
                [h.Vector() for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ]
            for cell_type in self.cell_types
        }

        self.apc_refs = []
        self.last_index = {cell_type: [
                [0 for index in range(self.cell_types_numbers[cell_type][1])]
                for population in range(self.cell_types_numbers[cell_type][0])
            ]
            for cell_type in self.cell_types
        }

        for cell_type in self.cell_types:
            for population in range(self.cell_types_numbers[cell_type][0]):
                for index, cell in enumerate(self.cells[cell_type][population]):
                    apc = h.APCount(cell(0.5))
                    apc.record(self.spike_times[cell_type][population][index])
                    self.apc_refs.append(apc)

    def _init_stimuli(self):
        for cell_type in self.cell_types:
            self.stims[cell_type], self.syns[cell_type], self.ncs[cell_type] = [], [], []
            for population in range(self.cell_types_numbers[cell_type][0]):
                for i, cell in enumerate(self.cells[cell_type][population]):
                    if cell_type == 'IO':
                        offset = 1
                    else:
                        offset = i*self.stim_intervals[cell_type]/self.cell_types_numbers[cell_type][1]
                    stim, syn, nc = create_stim(cell, start=offset, interval=self.stim_intervals[cell_type], e=0, tau=0.5, weight=self.stim_weights[cell_type])
                    nc.active(self.stim_active[cell_type])

                    self.stims[cell_type].append(stim)
                    self.syns[cell_type].append(syn)
                    self.ncs[cell_type].append(nc)

    def _init_connections(self):
        rng = np.random.default_rng(42)

        for spec in self.connection_specs:
            pre_group, post_group, label, e_rev, weight, tau, base_delay, threshold, sparsity, grouped = spec
            self.ncs[label] = []
            self.syns[label] = []

            all_pre = [cell for group in self.cells[pre_group] for cell in group]
            all_post = [cell for group in self.cells[post_group] for cell in group]

            if grouped:
                # Determine grouping based on which side has more neurons
                if len(all_pre) >= len(all_post):
                    # More pre than post: slice pre cells for each post
                    num_pre_per_post = len(all_pre) // len(all_post)
                    pre_groups = [
                        all_pre[i*num_pre_per_post:(i+1)*num_pre_per_post]
                        for i in range(len(all_post))
                    ]
                    post_groups = [[post] for post in all_post]
                else:
                    # More post than pre: slice post cells for each pre
                    num_post_per_pre = len(all_post) // len(all_pre)
                    post_groups = [
                        all_post[i*num_post_per_pre:(i+1)*num_post_per_pre]
                        for i in range(len(all_pre))
                    ]
                    pre_groups = [[pre] for pre in all_pre]
            else:
                # Full connection (all-to-all)
                pre_groups = [all_pre for _ in range(len(all_post))]
                post_groups = [[post] for post in all_post]

            # Create NetCon and ExpSyn
            pre_id_of  = {id(cell): i for i, cell in enumerate(all_pre)}
            post_id_of = {id(cell): i for i, cell in enumerate(all_post)}
            for pre_group_cells, post_group_cells in zip(pre_groups, post_groups):
                for pre_cell in pre_group_cells:
                    pre_id = pre_id_of[id(pre_cell)]
                    for post_cell in post_group_cells:
                        post_id = post_id_of[id(post_cell)]
                        # Optional sparsity only for full connection
                        if not grouped and sparsity:
                            k = 3
                            p_connect = min(1.0, k / len(all_pre))
                            delay = rng.uniform(0.5*base_delay, 2*base_delay)
                            if rng.random() > p_connect:
                                continue
                        else:
                            if label == 'GC_to_PC': # add offset to PCs
                                delay = (post_id % (self.cell_types_numbers['PC'][1]//2))*4 + base_delay
                            else:
                                delay = base_delay
                        syn = h.ExpSyn(post_cell(0.5))
                        syn.e = e_rev
                        syn.tau = tau
                        nc = h.NetCon(pre_cell(0.5)._ref_v, syn, sec=pre_cell)
                        nc.threshold = threshold
                        nc.delay = delay
                        nc.weight[0] = weight
                        
                        if label == 'GC_to_PC':
                            self.weights_over_time[(pre_id, post_id)].append(nc.weight[0])
                        
                        self.syns[label].append(syn)
                        self.ncs[label].append(nc)

                        # Save index mapping
                        idx = len(self.ncs[label]) - 1
                        if not label in self.nc_index_map:
                            self.nc_index_map[label] = {}
                            self.pre_to_post[label] = {}
                            self.post_to_pre[label] = {}
                        self.nc_index_map[label][(pre_id, post_id)] = idx
                        self.pre_to_post[label].setdefault(pre_id, []).append(post_id)
                        self.post_to_pre[label].setdefault(post_id, []).append(pre_id)
  
    def _init_recording(self):
        # Recording
        self.recordings = {ct: [[h.Vector().record(cell(0.5)._ref_v) for cell in self.cells[ct][population]] for population in range(self.cell_types_numbers[ct][0])] for ct in self.cell_types}
        self.t_vec = h.Vector().record(h._ref_t)

    def _init_plotting(self):
        plt.ion()
        self.fig = plt.figure(figsize=(13, 8))
        self.fig.canvas.manager.set_window_title('Cerebellum')
        self.rows = 5
        self.gs = gridspec.GridSpec(3, 1, figure=self.fig, height_ratios=[1, 1, 2 * self.rows])
        self.gs_input = self.gs[0].subgridspec(1, 2)
        self.gs_control = self.gs[1].subgridspec(1, len(self.joints) - 1)
        self.gs_plot = self.gs[2].subgridspec(self.rows, self.num_of_plots, height_ratios=[2]*3 + [1]*2)
        self.axs_input = [self.fig.add_subplot(gs) for gs in self.gs_input]
        self.axs_control = [self.fig.add_subplot(gs) for gs in self.gs_control]
        self.axs_plot = []
        for i in range(self.rows):
            row = []
            for j in range(self.num_of_plots):
                if j == 0:
                    ax = self.fig.add_subplot(self.gs_plot[i, j])
                else:
                    ax = self.fig.add_subplot(self.gs_plot[i, j], sharey=row[0])
                row.append(ax)
            self.axs_plot.append(row)
        for i in range(self.rows):
            for j in range(1, self.num_of_plots):  # skip first in row
                plt.setp(self.axs_plot[i][j].get_yticklabels(), visible=False)

        # Deactivate axis for control axes
        for ax in self.axs_control:
            ax.set_axis_off() # deactivate axis

        [self.row_potential, self.row_spike, self.row_weights, self.row_errors, self.row_corrections] = list(range(self.rows))
        
        self._init_membrane_potential_plot()
        self._init_spike_plot()
        self._init_control_panel()
        self._init_weight_plot()
        self._init_error_plot()
        self._init_correction_plot()

        plt.show()
        plt.tight_layout()

    def _init_membrane_potential_plot(self):
        # Membrane potential plot
        self.axs_plot[self.row_potential][0].set_ylabel('Membrane\npotential (mV)')
        self.axs_input[0].set_ylabel('Membrane\npotential (mV)')
        self.mem_lines = {ct: [] for ct in self.cell_types}
        self.mem_lines_pos = {ct: [[] for _ in range(self.num_of_plots)] for ct in self.cell_types}
        self.mem_lines_neg = {ct: [[] for _ in range(self.num_of_plots)] for ct in self.cell_types}

        for j, ct in enumerate(self.cell_types):
            if ct == 'PN' or ct == 'GC':
                for n in range(self.cell_types_numbers[ct][1]):
                    label = ct if n==0 else ''
                    line, = self.axs_input[0].plot([], [], self.cell_types_colors[ct][0], label=label, alpha=0.6)
                    self.mem_lines[ct].append(line)
        self.axs_input[0].set_xlim(0, self.plot_interval)
        self.axs_input[0].set_ylim(-85, 65)
        self.axs_input[0].xaxis.set_major_formatter(ms_to_s)
        self.axs_input[0].legend(loc='upper right')

        for plot_id in range(self.num_of_plots):
            for j, ct in enumerate(self.cell_types):
                if ct != 'PN' and ct != 'GC':
                    for n in range(self.cell_types_numbers[ct][1] // 2):
                        label_pos = ct+'_pos' if n==0 else ''
                        label_neg = ct+'_neg' if n==0 else ''
                        line_pos, = self.axs_plot[self.row_potential][plot_id].plot([], [], self.cell_types_colors[ct][0], label=label_pos, alpha=0.6)
                        line_neg, = self.axs_plot[self.row_potential][plot_id].plot([], [], self.cell_types_colors[ct][-1], linestyle='dashed', label=label_neg, alpha=0.6)
                        self.mem_lines_pos[ct][plot_id].append(line_pos)
                        self.mem_lines_neg[ct][plot_id].append(line_neg)
            title = self.joints[plot_id]
            self.axs_plot[self.row_potential][plot_id].set_title(title)
            self.axs_plot[self.row_potential][plot_id].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_potential][plot_id].set_ylim(-85, 65)
            self.axs_plot[self.row_potential][plot_id].xaxis.set_major_formatter(ms_to_s)
        self.axs_plot[self.row_potential][-1].legend(loc='upper right')

    def _init_spike_plot(self):
        # Spike raster plot and rate lines
        self.axs_plot[self.row_spike][0].set_ylabel('Spike raster')
        self.raster_lines = {ct: [[] for _ in range(self.num_of_plots)] for ct in self.cell_types}
        
        self.axs_input[1].set_ylabel('Spike raster')
        self.raster_lines_input = {ct: [] for ct in self.cell_types if ct == 'PN' or ct == 'GC'}

        # Spikes of PN and GC
        self.total_cells_input = 0
        for j, ct in enumerate(self.cell_types):
            if ct == 'PN' or ct == 'GC':
                for _ in range(self.cell_types_numbers[ct][1]):
                    line, = self.axs_input[1].plot([], [], self.cell_types_colors[ct][0]+'.', markersize=3)
                    self.raster_lines_input[ct].append(line)
                    self.total_cells_input += 1
            
        y_max = self.total_cells_input + 1.5
        self.axs_input[1].set_ylim(0.5, y_max)
        yticks = []
        yticklabels = []
        cumulative = 0
        for ct in self.cell_types:
            if ct == 'PN' or ct == 'GC':
                mid = y_max - (cumulative + (self.cell_types_numbers[ct][1]+1) / 2)
                yticks.append(mid)
                yticklabels.append(ct)
                cumulative += self.cell_types_numbers[ct][1] 
        self.axs_input[1].set_xlim(0, self.plot_interval)
        self.axs_input[1].set_yticks(yticks)
        self.axs_input[1].set_yticklabels(yticklabels)
        self.axs_input[1].xaxis.set_major_formatter(ms_to_s)

        # Spikes of other cells
        for i, ch in enumerate(range(self.num_of_plots)):
            self.total_cells = 0
            for j, ct in enumerate(self.cell_types):
                if ct != 'PN' and ct != 'GC':
                    self.raster_lines[ct][i] = []
                    for k in range(self.cell_types_numbers[ct][1]):
                        color = self.cell_types_colors[ct][0] if k < self.cell_types_numbers[ct][1] // 2 else self.cell_types_colors[ct][-1]
                        line, = self.axs_plot[self.row_spike][i].plot([], [], color+'.', markersize=3)
                        self.raster_lines[ct][i].append(line)
                        self.total_cells += 1

            y_max = self.total_cells + 1
            self.axs_plot[self.row_spike][i].set_ylim(0.5, y_max)
            yticks = []
            yticklabels = []
            cumulative = 0
            for half_label in ["pos", "neg"]:
                for ct in self.cell_types:
                    if ct != 'PN' and ct != 'GC':
                        n_cells = self.cell_types_numbers[ct][1] // 2
                        mid = y_max - (cumulative + (n_cells + 1) / 2)
                        yticks.append(mid)
                        yticklabels.append(f"{ct}_{half_label}")
                        cumulative += n_cells
            
            # Horizontal line to split positive and negative parts
            self.axs_plot[self.row_spike][i].axhline(y=y_max/2, color="black", linestyle='solid', linewidth=1)

            self.axs_plot[self.row_spike][i].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_spike][i].set_yticks(yticks)
            self.axs_plot[self.row_spike][i].set_yticklabels(yticklabels)
            self.axs_plot[self.row_spike][i].xaxis.set_major_formatter(ms_to_s)

    def _init_control_panel(self):
        for i, name in enumerate(self.joints[1:]):
            ax_err_pos = self.axs_control[i].inset_axes([0.3,0.5,0.6,0.45]) #[x0, y0, width, height]
            ax_err_neg = self.axs_control[i].inset_axes([0.3,0,0.6,0.45]) #[x0, y0, width, height]
            label_text = ' '.join(name.split()[:-1])
            ax_err_pos.set_title(label_text)
            self.buttons[f'err_pos{i}'] = Slider(ax_err_pos, 'Grip defizit', 0, 1, valinit=0, valstep=1)
            self.buttons[f'err_pos{i}'].on_changed(lambda val, i=i: self.update_touch_sensor_values(finger_idx=i, val=val, dir='pos'))
            self.buttons[f'err_neg{i}'] = Slider(ax_err_neg, 'Grip overforce', 0, 1, valinit=0, valstep=1)
            self.buttons[f'err_neg{i}'].on_changed(lambda val, i=i: self.update_touch_sensor_values(finger_idx=i, val=val, dir='neg'))

    def _init_weight_plot(self):
        # Weight plot
        self.axs_plot[self.row_weights][0].set_ylabel('GC-PC weights')
        self.weight_lines = [[[] for _ in range(self.cell_types_numbers['PC'][1])] for _ in range(self.num_of_plots)]

        for idx in range(self.num_of_plots):
            for pc_id in range(self.cell_types_numbers['PC'][1]):
                style = 'solid' if pc_id < self.cell_types_numbers['PC'][1] // 2 else 'dashed'
                color = self.cell_types_colors['PC'][0] if pc_id < self.cell_types_numbers['PC'][1] // 2 else self.cell_types_colors['PC'][-1]
                for gc_id in range (self.total_cell_numbers['GC']):
                    label = '' if pc_id not in [0, self.cell_types_numbers['PC'][1] // 2] or gc_id != 0 else ('PC_pos' if pc_id < self.cell_types_numbers['PC'][1] // 2 else 'PC_neg')
                    line, = self.axs_plot[self.row_weights][idx].step([], [], color, linestyle=style, label=label, where='post', alpha=0.6)
                    self.weight_lines[idx][pc_id].append(line)

            self.axs_plot[self.row_weights][idx].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_weights][idx].set_ylim(0, 1.1 * self.max_weight)
            self.axs_plot[self.row_weights][idx].xaxis.set_major_formatter(ms_to_s)
        self.axs_plot[self.row_weights][-1].legend(loc='upper right')

    def _init_error_plot(self):
        # Error plot
        self.axs_plot[self.row_errors][0].set_ylabel('Error')
        self.error_lines = [[] for _ in range(self.num_of_plots)]

        for plot_idx in range(self.num_of_plots):
            line, = self.axs_plot[self.row_errors][plot_idx].step([], [], self.cell_types_colors['IO'][0], where='post')
            self.error_lines[plot_idx] = line
            self.axs_plot[self.row_errors][plot_idx].axhline(y=0, color="black", linestyle='solid', linewidth=1, alpha=0.5)
            self.axs_plot[self.row_errors][plot_idx].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_errors][plot_idx].set_ylim(-1.1, 1.1)
            self.axs_plot[self.row_errors][plot_idx].xaxis.set_major_formatter(ms_to_s)

    def _init_correction_plot(self):
        # Correction plot
        self.axs_plot[self.row_corrections][0].set_ylabel('Correction\n(ms)')
        self.correction_lines = [[] for _ in range(self.num_of_plots)]

        for plot_idx in range(self.num_of_plots):
            line, = self.axs_plot[self.row_corrections][plot_idx].step([], [], self.cell_types_colors['DCN'][0], where='post')
            self.correction_lines[plot_idx] = line

            self.axs_plot[self.row_corrections][plot_idx].set_xlim(0, self.plot_interval)
            self.axs_plot[self.row_corrections][plot_idx].set_ylim(-1.1, 1.1)
            self.axs_plot[self.row_corrections][plot_idx].xaxis.set_major_formatter(ms_to_s)
            self.axs_plot[self.row_corrections][plot_idx].set_xlabel('Simulation\ntime (s)')
            self.axs_plot[self.row_corrections][plot_idx].axhline(y=-5, color="black", linestyle='solid', linewidth=1, alpha=0.5)
            self.axs_plot[self.row_corrections][plot_idx].axhline(y=5, color="black", linestyle='solid', linewidth=1, alpha=0.5)
            self.axs_plot[self.row_corrections][plot_idx].set_ylim(-11, 11)

    def update_touch_sensor_values(self, finger_idx, val, dir='pos'):
        if self._is_updating_programmatically:
            return
        
        if dir == 'pos':
            self.touch_sensor_delta_grasp[finger_idx] = 0 if val else self.touch_threshold
            self.touch_sensor_delta_hold[finger_idx] = self.drop_threshold if val else 0
        elif dir == 'neg':
            self.touch_sensor_delta_grasp[finger_idx] = self.touch_overload_threshold if val else self.touch_threshold

    def update_stimulus_activation(self, ct, input=None):
        for k, _ in enumerate(self.ncs[ct]):
            val = input[k]
            self.ncs[ct][k].active(True if val > 0 else False)
            if val != 0: # adjust spike rate
                interval = self.stim_intervals[ct]
                stim = self.ncs[ct][k].pre()
                stim.interval = interval / val
                
    def update_input_stimuli(self, desired_grasp_type, desired_joints, desired_actuators, norm_touch_delta_grasp=None, norm_touch_delta_hold=None):
        input = []
        input.extend([int(ch) for ch in desired_grasp_type])
        input.extend([int(ch) for ch in desired_joints])
        input.extend([int(ch) for ch in desired_actuators])
        if norm_touch_delta_grasp: input.extend(norm_touch_delta_grasp)
        if norm_touch_delta_hold: input.extend(norm_touch_delta_hold)

        # Update PN stimuli
        relative_baseline = 0.4
        input_with_baseline_spiking = [(1 - relative_baseline) * val + relative_baseline for val in input]
        self.update_stimulus_activation(ct='PN', input=input_with_baseline_spiking)
    
    def set_err_button(self, name, value):
        if self.buttons[name].val != value:
            self._is_updating_programmatically = True
            self.buttons[name].set_val(value)
            self._is_updating_programmatically = False

    def update_teaching_stimuli(self, current_time, desired_joints, action_selection_completed=True, touch_sensor_delta_grasp=None, touch_sensor_delta_hold=None):
        self.error_times.append(current_time)
        if touch_sensor_delta_grasp: self.touch_sensor_delta_grasp = touch_sensor_delta_grasp
        if touch_sensor_delta_hold: self.touch_sensor_delta_hold = touch_sensor_delta_hold

        joint_to_finger_mapping = [0] + [i for i in range(len(self.joints) - 1)] # 2 joints for thumb
        teaching_input = [0] * self.total_cell_numbers['IO']
        
        for j_idx, desired_joint in enumerate(desired_joints):
            error_dir = 0
            if action_selection_completed:
                finger_idx = joint_to_finger_mapping[j_idx]
                desired_joint = int(desired_joint)

                touch_grasp = self.touch_sensor_delta_grasp[finger_idx]
                touch_hold = self.touch_sensor_delta_hold[finger_idx]

                # Check mismatches
                if desired_joint == 1:
                    if touch_grasp >= self.touch_overload_threshold:
                        # too much pressure applied
                        error_dir = -1 # need negative correction
                    elif (touch_grasp < self.touch_threshold or touch_hold <= self.drop_threshold):
                        # wanted to grip but failed / released early
                        error_dir = 1 # need positive correction
                
                err_pos_name = f"err_pos{finger_idx}"
                err_neg_name = f"err_neg{finger_idx}"
                if error_dir == 1:  # positive correction
                    teaching_input[2 * j_idx] = 1
                    self.set_err_button(err_pos_name, 1)
                    self.set_err_button(err_neg_name, 0)
                elif error_dir == -1:  # negative correction
                    teaching_input[2 * j_idx + 1] = 1
                    self.set_err_button(err_pos_name, 0)
                    self.set_err_button(err_neg_name, 1)
                else:  # neutral
                    self.set_err_button(err_pos_name, 0)
                    self.set_err_button(err_neg_name, 0)
                self.error_over_time[j_idx].append(error_dir)
            else:
                self.error_over_time[j_idx].append(error_dir)
            
        # Update IO stimuli
        self.update_stimulus_activation(ct='IO', input=teaching_input)

    def get_new_spikes(self, cell_type, population, index):
        vec = self.spike_times[cell_type][population][index]
        start = self.last_index[cell_type][population][index]
        all_spikes = vec.to_python()
        new_spikes = all_spikes[start:]
        self.last_index[cell_type][population][index] = len(all_spikes)
        return new_spikes

    def STDP_kernel(self, delta_t):
        """
        Biphasic kernel for GC->PC plasticity:
        - negative peak at delta_t ~ center (e.g. -100 ms)
        - positive at large |delta_t|
        """
        return self.STDP_A_pos - self.STDP_A_neg * np.exp(-((delta_t - self.STDP_center) ** 2) / (2 * self.STDP_sigma ** 2))
    
    def update_weights(self, current_time, action_selection_completed=True):
        self.weight_times.append(int(current_time))

        # --- Previous GC spikes for plasticity ---
        gc_spikes_prev = self.gc_spikes_last_interval if self.gc_spikes_last_interval else {}

        # --- Gather IO spikes (current interval only) ---
        io_spikes = {}
        for pop in range(self.cell_types_numbers['IO'][0]):
            for io_id in range(self.cell_types_numbers['IO'][1]):
                new_spikes = self.get_new_spikes('IO', pop, io_id)
                if new_spikes:
                    io_spikes[(pop, io_id)] = new_spikes

        # --- Update GC->PC weights ---
        for gc_id in range(self.total_cell_numbers['GC']):
            for pc_id in range(self.total_cell_numbers['PC']):
                key = (gc_id, pc_id)

                # Get last weight
                if key not in self.weights_over_time:
                    last_weight = self.ncs['GC_to_PC'][self.nc_index_map['GC_to_PC'][key]].weight[0]
                    self.weights_over_time[key] = [last_weight]
                else:
                    last_weight = self.weights_over_time[key][-1]

                delta_w = 0.0

                if action_selection_completed:
                    # Loop over all IO neurons
                    for pop in range(self.cell_types_numbers['IO'][0]):
                        for io_id in range(self.cell_types_numbers['IO'][1]):
                            io_times = io_spikes.get((pop, io_id), [])  # empty if no spikes

                            # Assign IO population to correct PC range
                            pc_start = pop * self.cell_types_numbers['PC'][1]
                            pc_end   = (pop + 1) * self.cell_types_numbers['PC'][1]
                            pc_half  = self.cell_types_numbers['PC'][1] // 2
                            io_half  = self.cell_types_numbers['IO'][1] // 2

                            if io_id < io_half:
                                pc_range = range(pc_start, pc_start + pc_half)
                            else:
                                pc_range = range(pc_start + pc_half, pc_end)

                            if pc_id not in pc_range:
                                continue  # skip irrelevant IO–PC pairs

                            # Apply kernel for all GC spikes from previous interval
                            for gc_t in gc_spikes_prev.get(gc_id, []):
                                if io_times:  # IO fired → compute real dt
                                    for io_t in io_times:
                                        dt = gc_t - io_t
                                        delta_w += self.STDP_kernel(dt)
                                else:  # No IO spikes → baseline LTP
                                    dt = np.inf
                                    delta_w += self.STDP_kernel(dt)

                # Clamp and apply weight
                new_weight = max(self.min_weight, min(self.max_weight, last_weight + delta_w))
                idx = self.nc_index_map['GC_to_PC'][key]
                self.ncs['GC_to_PC'][idx].weight[0] = new_weight
                self.weights_over_time[key].append(round(new_weight, 4))

        # --- Save CURRENT GC spikes for use in NEXT interval ---
        gc_spikes_now = {}
        for pop in range(self.cell_types_numbers['GC'][0]):
            for gc_id in range(self.cell_types_numbers['GC'][1]):
                new_spikes = self.get_new_spikes('GC', pop, gc_id)
                if new_spikes:
                    gc_spikes_now[gc_id] = new_spikes
        self.gc_spikes_last_interval = gc_spikes_now

    def update_plots(self, current_time, goal=None, action=None):
        # Update plots
        # Spike raster plot for PN and GC
        y_base_input = self.total_cells_input
        for ct in self.cell_types:
            if ct == 'PN' or ct == 'GC':
                spike_block = self.spike_times[ct][0]
                    
                for k in range(self.cell_types_numbers[ct][1]):
                    y_val_input = y_base_input - k
                    spikes = np.array(spike_block[k].to_python())
                    y_vals = np.ones_like(spikes) * y_val_input
                    self.raster_lines_input[ct][k].set_data(spikes, y_vals)
                y_base_input -= self.cell_types_numbers[ct][1]
        new_xlim = max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time))
        if self.axs_input[1].get_xlim() != new_xlim:
            self.axs_input[1].set_xlim(*new_xlim)

        # Membrane potential plot for PN and GC
        for ct in self.cell_types:
            if ct == 'PN' or ct == 'GC':
                voltages = np.array([(self.recordings[ct][0][j]) for j in range(self.cell_types_numbers[ct][1])])
                for n, voltage in enumerate(voltages):
                    self.mem_lines[ct][n].set_data(np.array(self.t_vec), voltage)
                new_xlim = (max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time)))
                if self.axs_input[0].get_xlim() != new_xlim:
                    self.axs_input[0].set_xlim(*new_xlim)

        for plot_id in range(self.num_of_plots):
            # Membrane potential plot
            for ct in self.cell_types:
                if ct != 'PN' and ct != 'GC':
                    voltages = np.array([(self.recordings[ct][plot_id][j]) for j in range(self.cell_types_numbers[ct][1])])
                    voltages_pos = voltages[:len(voltages) // 2]
                    voltages_neg = voltages[len(voltages) // 2:]
                    for n, voltage in enumerate(voltages_pos):
                        self.mem_lines_pos[ct][plot_id][n].set_data(np.array(self.t_vec), voltage)
                    for n, voltage in enumerate(voltages_neg):
                        self.mem_lines_neg[ct][plot_id][n].set_data(np.array(self.t_vec), voltage)
                    new_xlim = (max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time)))
                    if self.axs_plot[self.row_potential][plot_id].get_xlim() != new_xlim:
                        self.axs_plot[self.row_potential][plot_id].set_xlim(*new_xlim)
            
            # Spike raster plot
            y_base = self.total_cells
            grouped_order = []

            # Determine halves for positive and negative correction
            for half in ['pos', 'neg']:
                for ct in self.cell_types:
                    if ct != 'PN' and ct != 'GC':
                        num_cells = self.cell_types_numbers[ct][1]
                        if half == 'pos':
                            indices = range(num_cells // 2)
                        else:
                            indices = range(num_cells // 2, num_cells)
                        grouped_order.append((ct, indices))
            
            # Plot spikes according to the grouped_order
            y_current = y_base
            for ct, indices in grouped_order:
                spike_block = self.spike_times[ct][plot_id]
                all_spikes = []

                for k in indices:
                    spikes = np.array(spike_block[k].to_python())
                    y_vals = np.ones_like(spikes) * y_current
                    self.raster_lines[ct][plot_id][k].set_data(spikes, y_vals)
                    all_spikes.extend(spikes)
                    y_current -= 1  # decrement y for next neuron

            # Set xlim
            new_xlim = max(0, int(current_time) - self.plot_interval), max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_spike][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_spike][plot_id].set_xlim(*new_xlim)

            # Weight plot
            for pc_idx in range(self.cell_types_numbers['PC'][1]):
                pc_id = plot_id * self.cell_types_numbers['PC'][1] + pc_idx
                for gc_id, line in enumerate(self.weight_lines[plot_id][pc_idx]):
                    key = (gc_id, pc_id)
                    if key in self.weights_over_time:
                        line.set_data(self.weight_times, self.weights_over_time[key])
            new_xlim = 0, max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_weights][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_weights][plot_id].set_xlim(*new_xlim)

            # Error plot
            self.error_lines[plot_id].set_data(self.error_times, self.error_over_time[plot_id])
            new_xlim = 0, max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_errors][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_errors][plot_id].set_xlim(*new_xlim)

            # Correction plot
            self.correction_lines[plot_id].set_data(self.correction_times, self.correction_over_time[plot_id])
            self.correction_lines[plot_id].axes.relim()
            new_xlim = 0, max(self.plot_interval, int(current_time))
            if self.axs_plot[self.row_corrections][plot_id].get_xlim() != new_xlim:
                self.axs_plot[self.row_corrections][plot_id].set_xlim(*new_xlim)
        # Compute global limits from all axes
        limits = [(ax.dataLim.ymin, ax.dataLim.ymax) for ax in self.axs_plot[self.row_corrections]]
        ymins, ymaxs = zip(*limits)
        global_ymin, global_ymax = min(ymins), max(ymaxs)
        # Apply to all axes
        absolute_max = 1.1 * max(abs(global_ymin), abs(global_ymax))
        if absolute_max > 10:
            for ax in self.axs_plot[self.row_corrections]:
                ax.set_ylim(-absolute_max, absolute_max)
    
    def analyze_firing_rate(self, cell, window=None, diff=True):
        """Returns a list of firing rates (Hz) for each action's cell population."""
        current_time = h.t
        rates_diff = []
        rates = []
        if window == None:
            window = self.plot_interval
        for p in range(self.cell_types_numbers[cell][0]):
            spikes_diff = 0
            spikes = []
            for i in range(self.cell_types_numbers[cell][1]):
                spike_vec = self.spike_times[cell][p][i]
                # Count spikes in the last `window` ms
                recent_spikes = [t for t in spike_vec if current_time - window <= t <= current_time]
                if diff:
                    if i < self.cell_types_numbers[cell][1] // 2:
                        spikes_diff += len(recent_spikes)
                    else:
                        spikes_diff -= len(recent_spikes)
                else:
                    spikes.append(len(recent_spikes))
            if diff:
                rate_diff = spikes_diff / (window / 1000.0)  # spikes/sec per neuron
                rates_diff.append(rate_diff)
            else:
                rate = [spike / (window / 1000.0) for spike in spikes]
                rates.append(rate)
        
        if diff:
            return rates_diff
        else:
            return rates
        
    def calculate_correction(self, current_time):
        self.correction_times.append(current_time)
        self.DCN_diff_rates = self.analyze_firing_rate('DCN')
        for joint_idx, corr in enumerate(self.DCN_diff_rates):
            self.correction_over_time[joint_idx].append(int(corr))
    
    def save_data(self, path):
        # Workbook
        wb = Workbook()
        path_extented = f"{path}_Cerebellum"

        # Worksheet for General Details
        ws_globals = wb.active
        ws_globals.title = "GlobalVariables"

        row = 1
        row = write_dict(ws_globals, "cell_types_numbers", self.cell_types_numbers, row)
        row = write_dict(ws_globals, "stim_intervals", self.stim_intervals, row)
        row = write_dict(ws_globals, "stim_weights", self.stim_weights, row)
        row = write_dict(ws_globals, "stim_active", self.stim_active, row)
        row = write_tuples(ws_globals, "connection_specs", self.connection_specs, row)
    
        # --- Scalars ---
        scalars = {
            "min_weight": self.min_weight,
            "max_weight": self.max_weight,
            "touch_threshold": self.touch_threshold,
            "touch_overload_threshold": self.touch_overload_threshold,
            "drop_threshold": self.drop_threshold,
            "STDP_A_pos": self.STDP_A_pos,
            "STDP_A_neg": self.STDP_A_neg,
            "STDP_center": self.STDP_center,
            "STDP_sigma": self.STDP_sigma
        }
        row = write_dict(ws_globals, "Scalars", scalars, row)

        # Worksheet for Weights
        ws_weights = wb.create_sheet(title="WeightsOverTime")

        # Header
        header = ['time']
        keys = sorted(self.weights_over_time.keys())
        header.extend(f"gc{gc_id}_pc{pc_id}" for gc_id, pc_id in keys)
        ws_weights.append(header)

        # Weights
        max_len = len(self.weight_times)
        for t_idx in range(max_len):
            row = [self.weight_times[t_idx]]
            for key in keys:
                val = self.weights_over_time[key][t_idx] if t_idx < len(self.weights_over_time[key]) else None
                row.append(val)
            ws_weights.append(row)

        # Worksheets
        data_list = [
            ("error_over_time", self.error_over_time), 
            ("correction_over_time", self.correction_over_time)
            ]
        ws_list = []
        for name, data in data_list:
            ws = wb.create_sheet(title=f"{name}")
            
            # Header
            header = ['time']
            keys = sorted(data.keys())
            header.extend(f"{key}" for key in keys)
            ws.append(header)

            # Rows
            max_len = len(self.error_times)
            for t_idx in range(max_len):
                row = [self.error_times[t_idx]]
                for key in keys:
                    val = data[key][t_idx] if t_idx < len(data[key]) else None
                    row.append(val)
                ws.append(row)

            ws_list.append(ws)

        # Save
        wb.save(f"{path_extented}.xlsx") # Excel
        print(path_extented)

#--- Functions ------------------------------------------------------------------------------------------------------------------------------------------------#

def create_cell(name):
    sec = h.Section(name=name)
    sec.insert('hh')
    return sec

# Stimulus helper
def create_stim(cell, start=0, number=1e9, interval=10, e=0, tau=1, weight=2, noise=0):
    stim = h.NetStim()
    stim.start = start
    stim.number = number
    stim.interval = interval
    stim.noise = noise
    syn = h.ExpSyn(cell(0.5))
    syn.e = e
    syn.tau = tau
    nc = h.NetCon(stim, syn)
    nc.weight[0] = weight
    nc.delay = 1
    return stim, syn, nc

def create_goal_action_table(indices_mapping, goals, actions):
    goal_action_table = {}
    all_goal_combinations = [''.join(bits) for bits in product("01", repeat=len(goals))]

    for goal_combo in all_goal_combinations:
        # Get the corresponding action string, or default to "0" * len(actions)
        action_combo = indices_mapping.get(goal_combo, "0" * len(actions))
        
        # Build the key: a tuple of (goal, True/False) tuples
        key = tuple((goal, bit == "1") for goal, bit in zip(goals, goal_combo))
        
        # Build the value: a dict of {action: True/False}
        value = {action: bit == "1" for action, bit in zip(actions, action_combo)}
        
        # Add to final mapping
        goal_action_table[key] = value

    return goal_action_table

def create_goal_action_indices_mapping(indices, goals, actions):
    goal_action_indices_mapping = {}

    for joint_bits in product('10', repeat=len(goals)):
        if '1' not in joint_bits:
            continue

        actuator_bits = ['0'] * len(actions)
        for joint_index, bit in enumerate(joint_bits):
            if bit == '1':
                for actuator_index in indices[joint_index]:
                    actuator_bits[actuator_index] = '1'

        joint_key = ''.join(joint_bits)
        actuator_value = ''.join(actuator_bits)
        goal_action_indices_mapping[joint_key] = actuator_value

    return goal_action_indices_mapping

# Formatter function: 1000 ms → 1 s
ms_to_s = FuncFormatter(lambda x, _: f'{x/1000}' if x % 100 == 0 else '')

# --- Dictionaries ---
def write_dict(ws, name, data, row):
    ws.cell(row=row, column=1, value=name)
    row += 1
    for k, v in data.items():
        ws.cell(row=row, column=1, value=str(k))
        if isinstance(v, (list, tuple)):  # if value is a list or tuple
            ws.cell(row=row, column=2, value=", ".join(map(str, v)))
        elif isinstance(v, dict):  # handle nested dicts
            ws.cell(row=row, column=2, value=json.dumps(v))
        else:  # if value is a number, string, etc.
            ws.cell(row=row, column=2, value=v)
        row += 1
    row += 1
    return row

# --- Lists ---
def write_list(ws, name, lst, row):
    ws.cell(row=row, column=1, value=name)
    for i, val in enumerate(lst):
        ws.cell(row=row + 1 + i, column=1, value=val)
    row += len(lst) + 2
    return row

# --- Tuples/List of Tuples ---
def write_tuples(ws, name, tuples_list, row):
    ws.cell(row=row, column=1, value=name)
    for i, tup in enumerate(tuples_list):
        for j, val in enumerate(tup):
            ws.cell(row=row + 1 + i, column=1 + j, value=val)
    row += len(tuples_list) + 2
    return row


#--- Motor Learning ---------------------------------------------------------------------------------------------------------------------------------------------------#

grasp_types = [
    "Palmar pinch", 
    "Prismatic 2-finger pinch",
    "Lateral pinch"
]
all_joints = [
    "Thumb opposition", 
    "Thumb flexion", 
    "Index finger flexion", 
    "Middle finger flexion", 
    "Ring finger flexion", 
    "Pinky finger flexion"
]
all_actuators = [
    "Thumb oppositor", 
    "Thumb flexor", 
    "Index finger flexor", 
    "Middle finger flexor", 
    "Ring finger flexor", 
    "Pinky finger flexor"
]
all_actuators_flexors = [
    "Actuator 4", 
    "Actuator 1", 
    "Actuator 6", 
    "Actuator 8", 
    "Actuator 10", 
    "Actuator 12"
]
all_actuators_extensors = [
    "Actuator 2", 
    "Actuator 7", 
    "Actuator 9", 
    "Actuator 11", 
    "Actuator 13"
]
no_of_joints = 4
shuffle = False

grasp_type_joint_indices_mapping = {
    "100": "1" * min(3, no_of_joints) + "0" * (no_of_joints - min(3, no_of_joints)),
    "010": "1" * min(4, no_of_joints) + "0" * (no_of_joints - min(4, no_of_joints)),
    "001": "0" + "1" * min(2, no_of_joints-1) + "0" * (no_of_joints-1 - min(2, no_of_joints-1)),
}

MotorLearning(no_of_joints, basal_ganglia_required=True, cerebellum_required=True, user_feedback=True, grasp_types=grasp_types, grasp_type_joint_indices_mapping=grasp_type_joint_indices_mapping, all_joints=all_joints, all_actuators_flexors=all_actuators_flexors, all_actuators_extensors=all_actuators_extensors, shuffle_flexors=shuffle)
